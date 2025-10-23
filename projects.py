import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
import sqlite3
import re
import os
import subprocess
import sys
import shutil
from database_setup import DatabaseManager
from date_picker import DateEntry
from directory_picker import DirectoryPicker, FilePicker
from scroll_utils import bind_mousewheel_to_treeview
from notes_utils import open_add_note_dialog, append_job_note
from app_nav import add_app_bar
from help_utils import add_help_button
from duration_utils import (
    parse_duration_to_minutes,
    format_minutes_compact,
    ceil_minutes_to_business_days,
)

class CollapsibleFrame(ttk.Frame):
    """A collapsible frame widget"""
    def __init__(self, parent, text="", **kwargs):
        super().__init__(parent, **kwargs)
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        
        # Header frame with toggle button
        self.header = ttk.Frame(self, relief='raised', style='CollapsibleHeader.TFrame')
        self.header.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 2))
        self.header.columnconfigure(1, weight=1)
        
        # Toggle button (arrow)
        self.toggle_var = tk.StringVar(value="▼")
        self.toggle_btn = ttk.Button(self.header, textvariable=self.toggle_var, 
                                     width=3, command=self.toggle)
        self.toggle_btn.grid(row=0, column=0, padx=(5, 5))
        
        # Title label
        self.title_label = ttk.Label(self.header, text=text, 
                                     font=('Arial', 11, 'bold'))
        self.title_label.grid(row=0, column=1, sticky=tk.W, pady=5)
        
        # Content frame
        self.content = ttk.Frame(self, padding="10")
        self.content.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.content.columnconfigure(0, weight=1)
        
        self.collapsed = False
    
    def toggle(self):
        """Toggle the collapsed state"""
        if self.collapsed:
            self.content.grid()
            self.toggle_var.set("▼")
            self.collapsed = False
        else:
            self.content.grid_remove()
            self.toggle_var.set("▶")
            self.collapsed = True

class ProjectsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Project Management - Drafting Tools")
        
        # Maximize window (not fullscreen, keeps controls)
        self.root.state('zoomed')  # Windows maximized
        
        # Set minimum size
        self.root.minsize(1200, 800)

        # App bar
        try:
            add_app_bar(self.root, current_app='projects')
        except Exception:
            pass
        
        # Initialize database
        self.db_manager = DatabaseManager()
        
        # Initialize current project tracking
        self.current_project = None
        
        # Create a content container so we don't mix pack/grid on root
        self.content = ttk.Frame(self.root)
        self.content.pack(fill=tk.BOTH, expand=True)
        # Configure content grid weights for full expansion
        self.content.columnconfigure(0, weight=1)
        self.content.rowconfigure(0, weight=0)  # Title row (fixed)
        self.content.rowconfigure(1, weight=100)  # Main content (expands)
        self.content.rowconfigure(2, weight=0)  # Separator (fixed)
        self.content.rowconfigure(3, weight=0)  # Footer (fixed)
        
        self.create_widgets()
        self.load_projects()

        # Ensure table to track file timestamps per project
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS file_timestamps (
                    job_number TEXT NOT NULL,
                    path TEXT NOT NULL,
                    last_mtime REAL NOT NULL,
                    acknowledged INTEGER NOT NULL DEFAULT 1,
                    PRIMARY KEY(job_number, path)
                )
                """
            )
            conn.commit()
            conn.close()
        except Exception:
            pass
        
        # Add keyboard shortcuts for fullscreen toggle
        self.root.bind('<F11>', lambda e: self.toggle_fullscreen())
        self.root.bind('<Escape>', lambda e: self.exit_fullscreen() if self.root.attributes('-fullscreen') else None)
        
        # Bind window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    
    def create_widgets(self):
        """Create all GUI widgets with proper layout"""
        # Initialize redline update BooleanVars FIRST
        for cycle in range(1, 5):
            setattr(self, f"redline_update_{cycle}_var", tk.BooleanVar())
        
        # Row 0: Title
        title_label = ttk.Label(self.content, text="Project Management - Complete Workflow", 
                               font=('Arial', 18, 'bold'))
        title_label.grid(row=0, column=0, columnspan=2, pady=(10, 10), padx=20, sticky=(tk.W, tk.E))
        
        # Row 1: Main content area (expands to fill space)
        main_paned = ttk.PanedWindow(self.content, orient=tk.HORIZONTAL)
        main_paned.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=(0, 10))
        
        # Container frames for each section (workflow removed from Projects app)
        self.project_list_container = ttk.Frame(main_paned)
        self.project_details_container = ttk.Frame(main_paned)
        self.quick_access_container = ttk.Frame(main_paned)
        
        # Add containers to paned window - they will expand vertically
        main_paned.add(self.project_list_container, weight=1)
        main_paned.add(self.project_details_container, weight=1)
        main_paned.add(self.quick_access_container, weight=1)
        
        # Create panels inside containers
        self.create_project_list_panel()
        self.create_project_details_panel()
        self.create_quick_access_panel()
        
        # Row 2: Fixed footer with action buttons
        self.create_action_buttons()
        
        # Load dropdown data after all widgets are created
        self.load_dropdown_data()
    
    def create_project_list_panel(self):
        """Create the project list panel"""
        list_frame = ttk.LabelFrame(self.project_list_container, text="Projects", padding="10")
        list_frame.pack(fill=tk.BOTH, expand=True, padx=(0, 5))
        list_frame.columnconfigure(0, weight=1)
        list_frame.columnconfigure(1, weight=0)
        # Place help via grid to avoid pack/grid conflicts
        try:
            add_help_button(list_frame, 'Projects Pane', 'Search/sort jobs, right‑click for actions (including Add Note).').grid(row=0, column=1, sticky='ne')
        except Exception:
            pass
        list_frame.rowconfigure(1, weight=1)
        
        # Search and sort frame - combined for compact layout
        search_sort_frame = ttk.Frame(list_frame)
        search_sort_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        search_sort_frame.columnconfigure(1, weight=1)
        
        # Search row
        ttk.Label(search_sort_frame, text="Search:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(search_sort_frame, textvariable=self.search_var, width=25)
        self.search_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 0))
        self.search_var.trace('w', self.filter_projects)
        
        # Sort buttons row - directly under search (toggle functionality)
        self.job_sort_ascending = True  # Track sort direction for job numbers
        self.customer_sort_ascending = True  # Track sort direction for customers
        self.due_date_sort_ascending = True  # Track sort direction for due dates
        # Track visibility of completed projects
        self.show_completed = False
        # Save column widths on close
        try:
            self.root.protocol("WM_DELETE_WINDOW", self._on_close)
            # Global mouse release to catch column resize releases anywhere
            self.root.bind('<ButtonRelease-1>', lambda e: self._debounce_save_columns())
        except Exception:
            pass
        
        sort_btn_frame = ttk.Frame(search_sort_frame)
        sort_btn_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(3, 0))
        
        self.sort_job_btn = ttk.Button(sort_btn_frame, text="Job # ↑", command=self.sort_by_job_number, width=10)
        self.sort_job_btn.grid(row=0, column=0, padx=(0, 3), sticky=tk.W)
        
        self.sort_customer_btn = ttk.Button(sort_btn_frame, text="Customer ↑", command=self.sort_by_customer, width=12)
        self.sort_customer_btn.grid(row=0, column=1, padx=(0, 3), sticky=tk.W)
        
        self.sort_due_date_btn = ttk.Button(sort_btn_frame, text="Due Date ↑", command=self.sort_by_due_date, width=12)
        self.sort_due_date_btn.grid(row=0, column=2, padx=(0, 6), sticky=tk.W)

        # Toggle show/hide completed button
        self.toggle_completed_btn = ttk.Button(sort_btn_frame, text="Show Completed", command=self.toggle_completed, width=16)
        self.toggle_completed_btn.grid(row=0, column=3, padx=(0, 0), sticky=tk.W)
        
        # Treeview for projects - show due date, days until due, and status
        columns = ('Job Number', 'Customer', 'Due Date', 'Due in', 'Status')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=18)
        
        # Configure tags for row highlighting
        self.tree.tag_configure('selected', background='#E8F0FE', font=('Arial', 9, 'bold'))
        
        # Set column widths
        self.tree.heading('Job Number', text='Job Number')
        self.tree.column('Job Number', width=80, minwidth=80)
        
        self.tree.heading('Customer', text='Customer')
        self.tree.column('Customer', width=150, minwidth=100)
        
        self.tree.heading('Due Date', text='Due Date')
        self.tree.column('Due Date', width=100, minwidth=80)
        
        self.tree.heading('Due in', text='Due in (Days)')
        self.tree.column('Due in', width=100, minwidth=80)
        
        self.tree.heading('Status', text='Status')
        self.tree.column('Status', width=100, minwidth=80)
        
        # Scrollbar for treeview
        tree_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scrollbar.set)

        self.tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        try:
            bind_mousewheel_to_treeview(self.tree)
        except Exception:
            pass
        
        # Bind selection event
        self.tree.bind('<<TreeviewSelect>>', self.on_project_select)
        
        # Create right-click context menu for job numbers
        self.create_job_context_menu()
        
        # Job Notes area below the projects table
        notes_frame = ttk.LabelFrame(list_frame, text="Job Notes", padding="5")
        notes_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        notes_frame.columnconfigure(0, weight=1)
        notes_frame.rowconfigure(0, weight=1)
        
        # Notes text area
        self.notes_text = tk.Text(notes_frame, height=8, wrap=tk.WORD, font=('Arial', 10))
        notes_scrollbar = ttk.Scrollbar(notes_frame, orient=tk.VERTICAL, command=self.notes_text.yview)
        self.notes_text.configure(yscrollcommand=notes_scrollbar.set)
        
        self.notes_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        notes_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Save/Print notes buttons
        btns = ttk.Frame(notes_frame)
        btns.grid(row=1, column=0, sticky=tk.W, pady=(5, 0))
        save_notes_btn = ttk.Button(btns, text="Save Notes", command=self.save_job_notes)
        save_notes_btn.pack(side=tk.LEFT)
        print_notes_btn = ttk.Button(btns, text="Print Notes", command=self.print_job_notes)
        print_notes_btn.pack(side=tk.LEFT, padx=(8, 0))
        
        # Initialize notes
        self.current_job_notes = ""

        # Restore saved tree column widths and persist on change (per-user)
        try:
            self.apply_tree_column_widths()
        except Exception:
            pass
        # Save on left-button release and after resize motion
        self.tree.bind('<ButtonRelease-1>', lambda e: self._debounce_save_columns())
        self.tree.bind('<B1-Motion>', lambda e: self._debounce_save_columns())

    def _get_user_prefs_path(self):
        import json
        base = os.path.join(os.path.expanduser('~'), '.drafting_tools')
        os.makedirs(base, exist_ok=True)
        return os.path.join(base, 'ui_prefs.json')

    def apply_tree_column_widths(self):
        import json
        path = self._get_user_prefs_path()
        if not os.path.exists(path):
            return
        try:
            with open(path, 'r') as f:
                data = json.load(f)
            widths = data.get('projects_tree_columns')
            if widths:
                for col, w in widths.items():
                    if col in self.tree['columns']:
                        try:
                            self.tree.column(col, width=int(w))
                        except Exception:
                            continue
        except Exception:
            pass

    def _debounce_save_columns(self):
        # Debounce rapid events
        if hasattr(self, '_save_cols_after_id') and self._save_cols_after_id:
            try:
                self.root.after_cancel(self._save_cols_after_id)
            except Exception:
                pass
        self._save_cols_after_id = self.root.after(300, self.save_tree_column_widths)

    def save_tree_column_widths(self):
        import json
        widths = {}
        for col in self.tree['columns']:
            try:
                widths[col] = int(self.tree.column(col, option='width'))
            except Exception:
                continue
        path = self._get_user_prefs_path()
        data = {}
        if os.path.exists(path):
            try:
                with open(path, 'r') as f:
                    data = json.load(f)
            except Exception:
                data = {}
        data['projects_tree_columns'] = widths
        try:
            with open(path, 'w') as f:
                json.dump(data, f)
        except Exception:
            pass
        self._last_saved_widths = widths

    def _on_close(self):
        # Ensure widths are saved before exit
        try:
            self.save_tree_column_widths()
        except Exception:
            pass
        try:
            self.root.destroy()
        except Exception:
            self.root.quit()

    def print_job_notes(self):
        """Create an Excel sheet with project header and notes, then open it"""
        try:
            # Determine selected job
            selection = self.tree.selection()
            if not selection:
                messagebox.showwarning("Warning", "Select a project first")
                return
            values = self.tree.item(selection[0])['values']
            job_number = str(values[0])

            # Fetch project details
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.job_number, p.customer_name, p.customer_location,
                       p.job_directory,
                       d.name as designer, e.name as engineer,
                       p.assignment_date, p.start_date, p.due_date,
                       p.completion_date, COALESCE(p.released_to_dee, rd.release_date) as released_to_dee
                FROM projects p
                LEFT JOIN designers d ON p.assigned_to_id = d.id
                LEFT JOIN engineers e ON p.project_engineer_id = e.id
                LEFT JOIN release_to_dee rd ON rd.project_id = p.id
                WHERE p.job_number = ?
            """, (job_number,))
            proj = cursor.fetchone()
            if not proj:
                conn.close()
                messagebox.showerror("Error", "Project not found")
                return
            (p_job, cust, loc, job_dir, designer, engineer,
             assign_dt, start_dt, due_dt, comp_dt, rel_dt) = proj

            # Load notes
            cursor.execute("CREATE TABLE IF NOT EXISTS job_notes (job_number TEXT PRIMARY KEY, notes TEXT)")
            cursor.execute("SELECT notes FROM job_notes WHERE job_number = ?", (job_number,))
            row = cursor.fetchone()
            notes_text = row[0] if row and row[0] else self.notes_text.get("1.0", tk.END).strip()

            # Optional: list drawings for this job
            drawings = []
            try:
                cursor.execute("SELECT drawing_name, added_date FROM drawings WHERE job_number = ? ORDER BY drawing_name", (job_number,))
                drawings = cursor.fetchall()
            except Exception:
                drawings = []
            conn.close()

            # Create Excel
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            wb = Workbook()
            ws = wb.active
            ws.title = "Notes"

            # Header
            header = f"{p_job} — {cust or ''} — {loc or ''}"
            ws["A1"] = header
            ws["A1"].font = Font(name="Calibri", size=16, bold=True)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

            # Project fields
            fields = [
                ("Assigned To", designer or ""),
                ("Project Engineer", engineer or ""),
                ("Assignment Date", assign_dt or ""),
                ("Start Date", start_dt or ""),
                ("Due Date", due_dt or ""),
                ("Completion Date", comp_dt or ""),
                ("Released to Dee", rel_dt or ""),
            ]
            row_i = 3
            ws["A2"] = "Project Info"
            ws["A2"].font = Font(name="Calibri", size=12, bold=True)
            for label, val in fields:
                ws.cell(row=row_i, column=1, value=label)
                ws.cell(row=row_i, column=2, value=val)
                row_i += 1

            # Notes section
            row_i += 1
            ws.cell(row=row_i, column=1, value="Notes").font = Font(name="Calibri", size=12, bold=True)
            row_i += 1
            ws.cell(row=row_i, column=1, value=notes_text)
            ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i+10, end_column=4)
            ws.cell(row=row_i, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            row_i += 12

            # Drawings section
            ws.cell(row=row_i, column=1, value="Drawings").font = Font(name="Calibri", size=12, bold=True)
            row_i += 1
            if drawings:
                ws.cell(row=row_i, column=1, value="Name")
                ws.cell(row=row_i, column=2, value="Added/Updated")
                row_i += 1
                for name, added in drawings:
                    ws.cell(row=row_i, column=1, value=name)
                    ws.cell(row=row_i, column=2, value=added or "")
                    row_i += 1
            else:
                ws.cell(row=row_i, column=1, value="No drawings on record")

            # Column widths
            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 24
            ws.column_dimensions['D'].width = 24

            # Save to job directory
            base_dir = job_dir or os.path.join(os.path.expanduser('~'), 'Documents')
            target_dir = os.path.join(base_dir, 'Status Reports')
            os.makedirs(target_dir, exist_ok=True)
            from datetime import datetime as _dt
            fname = f"{job_number}-Notes-{_dt.now().strftime('%Y-%m-%d_%I%M%p')}.xlsx"
            out_path = os.path.join(target_dir, fname)
            wb.save(out_path)

            # Open for printing
            try:
                os.startfile(out_path)
            except Exception:
                messagebox.showinfo("Saved", f"Notes saved to\n{out_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to print notes: {str(e)}")
    
    def create_project_details_panel(self):
        """Create the project details panel"""
        details_frame = ttk.LabelFrame(self.project_details_container, text="Project Details", padding="10")
        details_frame.pack(fill=tk.BOTH, expand=True, padx=(0, 5))
        details_frame.columnconfigure(1, weight=1)
        # Keep a durable reference for other methods
        self.project_details_frame = details_frame
        try:
            add_help_button(details_frame, 'Details Pane', 'Edit job directory, customer, dates and assignments. Values autosave.').grid(row=0, column=2, sticky='ne')
        except Exception:
            pass
        
        # Job Number
        ttk.Label(details_frame, text="Job Number:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.job_number_var = tk.StringVar()
        self.job_number_entry = ttk.Entry(details_frame, textvariable=self.job_number_var, width=25)
        self.job_number_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Job Directory
        ttk.Label(details_frame, text="Job Directory:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.job_directory_picker = DirectoryPicker(details_frame, width=25)
        self.job_directory_picker.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Customer Name
        ttk.Label(details_frame, text="Customer Name:").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.customer_name_var = tk.StringVar()
        self.customer_name_entry = ttk.Entry(details_frame, textvariable=self.customer_name_var, width=50)
        self.customer_name_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Customer Name Path (moved below)
        self.customer_name_picker = DirectoryPicker(details_frame, width=50)
        self.customer_name_picker.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Customer Location
        ttk.Label(details_frame, text="Customer Location:").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.customer_location_var = tk.StringVar()
        self.customer_location_entry = ttk.Entry(details_frame, textvariable=self.customer_location_var, width=50)
        self.customer_location_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Customer Location Path (moved below)
        self.customer_location_picker = DirectoryPicker(details_frame, width=50)
        self.customer_location_picker.grid(row=5, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Assigned To
        ttk.Label(details_frame, text="Assigned to:").grid(row=6, column=0, sticky=tk.W, pady=2)
        self.assigned_to_var = tk.StringVar()
        self.assigned_to_combo = ttk.Combobox(details_frame, textvariable=self.assigned_to_var, 
                                            state="readonly", width=22)
        self.assigned_to_combo.grid(row=6, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Project Engineer
        ttk.Label(details_frame, text="Project Engineer:").grid(row=7, column=0, sticky=tk.W, pady=2)
        self.project_engineer_var = tk.StringVar()
        self.project_engineer_combo = ttk.Combobox(details_frame, textvariable=self.project_engineer_var, 
                                                 state="readonly", width=22)
        self.project_engineer_combo.grid(row=7, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Assignment Date
        ttk.Label(details_frame, text="Assignment Date:").grid(row=8, column=0, sticky=tk.W, pady=2)
        self.assignment_date_entry = DateEntry(details_frame, width=25)
        self.assignment_date_entry.grid(row=8, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Due Date
        ttk.Label(details_frame, text="Due Date:").grid(row=9, column=0, sticky=tk.W, pady=2)
        self.due_date_entry = DateEntry(details_frame, width=25)
        self.due_date_entry.grid(row=9, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Start Date
        ttk.Label(details_frame, text="Start Date:").grid(row=10, column=0, sticky=tk.W, pady=2)
        self.start_date_entry = DateEntry(details_frame, width=25)
        self.start_date_entry.grid(row=10, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Completion Date
        ttk.Label(details_frame, text="Completion Date:").grid(row=11, column=0, sticky=tk.W, pady=2)
        self.completion_date_entry = DateEntry(details_frame, width=25)
        self.completion_date_entry.grid(row=11, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Total Duration
        ttk.Label(details_frame, text="Total Project Duration:").grid(row=12, column=0, sticky=tk.W, pady=2)
        self.duration_var = tk.StringVar()
        self.duration_label = ttk.Label(details_frame, textvariable=self.duration_var, 
                                       foreground="blue", font=('Arial', 10, 'bold'))
        self.duration_label.grid(row=12, column=1, sticky=tk.W, pady=2, padx=(5, 0))
        
        # Released to Dee
        ttk.Label(details_frame, text="Released to Dee:").grid(row=13, column=0, sticky=tk.W, pady=2)
        self.released_to_dee_entry = DateEntry(details_frame, width=25)
        self.released_to_dee_entry.grid(row=13, column=1, sticky=(tk.W, tk.E), pady=2, padx=(5, 0))
        
        # Bind events
        self.start_date_entry.var.trace('w', self.calculate_duration)
        self.completion_date_entry.var.trace('w', self.calculate_duration)
        self.assignment_date_entry.var.trace('w', self.set_start_date)
        
        # Auto-save on any field change
        self.job_number_var.trace('w', self.auto_save)
        self.customer_name_var.trace('w', self.auto_save)
        self.customer_location_var.trace('w', self.auto_save)
        self.assigned_to_var.trace('w', self.auto_save)
        self.project_engineer_var.trace('w', self.auto_save)
        self.start_date_entry.var.trace('w', self.auto_save)
        self.completion_date_entry.var.trace('w', self.auto_save)
        self.due_date_entry.var.trace('w', self.auto_save)
        self.released_to_dee_entry.var.trace('w', self.auto_save)
        
        # Auto-save for directory pickers
        self.job_directory_picker.var.trace('w', self.auto_extract_and_save)
        self.customer_name_picker.var.trace('w', self.auto_save)
        self.customer_location_picker.var.trace('w', self.auto_save)
        
        # Add specifications section below project details
        self.create_specifications_section(details_frame)
    
    def create_specifications_section(self, parent_frame):
        """Create the specifications section below project details"""
        # Add separator
        separator = ttk.Separator(parent_frame, orient='horizontal')
        separator.grid(row=14, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 5))
        
        # Configure parent frame columns to use full width
        parent_frame.columnconfigure(0, weight=3)  # Buttons take 3/4 of space
        parent_frame.columnconfigure(1, weight=1)  # Inputs take 1/4 of space
        
        # Specifications label
        specs_label = ttk.Label(parent_frame, text="Specifications", font=('Arial', 12, 'bold'))
        specs_label.grid(row=15, column=0, columnspan=2, sticky=tk.W, pady=(5, 10))
        
        # Initialize specifications buttons list and input fields
        self.spec_buttons = []
        self.spec_input_fields = {}
        
        # Create specifications content
        self.update_specifications(parent_frame)
    
    def update_specifications(self, parent_frame=None):
        """Update the specifications panel based on available files"""
        if parent_frame is None:
            return
            
        # Clear existing buttons and input fields
        for button in self.spec_buttons:
            button.destroy()
        self.spec_buttons.clear()
        
        for field in self.spec_input_fields.values():
            field.destroy()
        self.spec_input_fields.clear()
        
        # Check if we have a job directory
        if not hasattr(self, 'job_directory_picker') or not self.job_directory_picker.get():
            no_data_label = ttk.Label(parent_frame, 
                                    text="No project selected", 
                                    foreground="gray", justify="center")
            no_data_label.grid(row=14, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=20)
            self.spec_buttons.append(no_data_label)
            return
        
        # Look for Heater Design file in Quick Access
        heater_design_file = None
        if hasattr(self, 'engineering_general_docs') and self.engineering_general_docs:
            for file_path in self.engineering_general_docs:
                filename = os.path.basename(file_path).upper()
                # Check if filename contains "HEATER DESIGN" OR if file has a "Heater Design" sheet inside
                if "HEATER DESIGN" in filename or self.has_heater_design_sheet(file_path):
                    heater_design_file = file_path
                    print(f"Found Heater Design file: {filename}")
                    break
        
        # Can Size specification
        can_size_value = "No Heater Design"
        can_size_button_state = "disabled"
        button_color = "#FFB6C1"  # Light pink for no file
        button_text_color = "black"
        
        if heater_design_file and os.path.exists(heater_design_file):
            try:
                # Read Excel file to get Can Size from "Heater Design" sheet
                can_size_value = self.read_excel_can_size(heater_design_file)
                if can_size_value:
                    can_size_button_state = "normal"
                    button_color = "#90EE90"  # Light green for file with value
                    button_text_color = "black"
                else:
                    button_color = "#FFB6C1"  # Light pink for file but no value
                    button_text_color = "black"
            except Exception as e:
                print(f"Error reading Heater Design file: {e}")
                can_size_value = "Error reading file"
                button_color = "#FFB6C1"  # Light pink for error
                button_text_color = "black"
        
        # Add Heater Specs group (now includes Can Size at the top)
        self.create_heater_specs_group(parent_frame, heater_design_file, 
                                       can_size_value, can_size_button_state, 
                                       button_color, button_text_color)
    
    def create_heater_specs_group(self, parent_frame, heater_design_file, 
                                  can_size_value=None, can_size_button_state="disabled",
                                  can_color="#FFB6C1", can_text_color="black"):
        """Create the Heater Specs group with dimension buttons"""
        # Heater Specs label
        heater_specs_label = ttk.Label(parent_frame, text="Heater Specs", font=('Arial', 11, 'bold'), foreground="darkblue")
        heater_specs_label.grid(row=17, column=0, columnspan=2, sticky=tk.W, pady=(15, 5))
        self.spec_buttons.append(heater_specs_label)
        
        # Can Size button at top of Heater Specs
        if can_size_value is not None:
            can_size_btn = tk.Button(parent_frame, 
                                    text=f"Can Size: {can_size_value}",
                                    state=can_size_button_state,
                                    command=lambda: self.open_heater_design_file(heater_design_file) if heater_design_file else None,
                                    width=60, height=1,
                                    font=('Arial', 9),
                                    relief='raised', bd=1, 
                                    cursor='hand2' if can_size_button_state == "normal" else 'arrow',
                                    bg=can_color, fg=can_text_color,
                                    activebackground=can_color, activeforeground=can_text_color)
            can_size_btn.grid(row=18, column=0, sticky=(tk.W, tk.E), pady=1, padx=(0, 10))
            self.spec_buttons.append(can_size_btn)
        
        # Define the heater dimension specifications
        heater_specs = [
            ("Heater Diameter", "H13"),
            ("Heater Height", "B31"),
            ("Can Height", "C33"),
            ("Packing Rings Bottom", "D41"),
            ("Packing Rings Height", "D31"),
            ("Packing Rings to Spray Nozzle", "D25"),
            ("Stack Diameter", "H14")
        ]
        
        # Add Spray Nozzle Size and Length from Engineering Design file
        spray_nozzle_size = self.get_spray_nozzle_size_from_engineering_design()
        spray_nozzle_length = self.get_spray_nozzle_length_from_engineering_design()
        
        # Check for manual specs first, then use found values
        manual_size = self.get_saved_manual_spec("Spray Nozzle Size")
        if spray_nozzle_size:
            heater_specs.append(("Spray Nozzle Size", spray_nozzle_size))
        elif manual_size:
            heater_specs.append(("Spray Nozzle Size", f"Manual: {manual_size}"))
        else:
            heater_specs.append(("Spray Nozzle Size", "No Size Found"))
            
        manual_length = self.get_saved_manual_spec("Spray Nozzle Length")
        if spray_nozzle_length:
            heater_specs.append(("Spray Nozzle Length", spray_nozzle_length))
        elif manual_length:
            heater_specs.append(("Spray Nozzle Length", f"Manual: {manual_length}"))
        else:
            heater_specs.append(("Spray Nozzle Length", "No Length Found"))
        
        # Add Spray Nozzle P/N from Spray Nozzles file
        spray_nozzle_pn = self.read_spray_nozzle_pn_from_files()
        manual_pn = self.get_saved_manual_spec("Spray Nozzle P/N")
        if spray_nozzle_pn:
            heater_specs.append(("Spray Nozzle P/N", spray_nozzle_pn))
        elif manual_pn:
            heater_specs.append(("Spray Nozzle P/N", f"Manual: {manual_pn}"))
        else:
            # Add "No Spray Nozzle Found" if no match found
            heater_specs.append(("Spray Nozzle P/N", "No Spray Nozzle Found"))
        
        # Create buttons for each spec
        for i, (spec_name, cell_ref_or_value) in enumerate(heater_specs):
            # Check if this is a pre-formatted value (Spray Nozzle specs) or a cell reference
            if spec_name in ["Spray Nozzle P/N", "Spray Nozzle Size", "Spray Nozzle Length"]:
                spec_value = cell_ref_or_value  # This is already the formatted value
            else:
                spec_value = self.read_heater_spec_value(heater_design_file, cell_ref_or_value)
                if not spec_value:
                    spec_value = "No Data"
                # Prefer saved manual value if present (applies to all heater specs)
                manual_override = self.get_saved_manual_spec(spec_name)
                if manual_override:
                    spec_value = f"Manual: {manual_override}"
            
            # Determine button color and state
            if spec_value and spec_value not in ["No Spray Nozzle Found", "No Size Found", "No Length Found", "No Data"]:
                if spec_value.startswith("Manual:"):
                    button_color = "#FFE082"  # Light yellow for manual values
                else:
                    button_color = "#90EE90"  # Light green for found values
                button_state = "normal"
                cursor_type = "hand2"
            elif spec_value in ["No Spray Nozzle Found", "No Size Found", "No Length Found", "No Data"]:
                button_color = "#FFB6C1"  # Light pink for "No [item] Found" or "No Data"
                button_state = "disabled"
                cursor_type = "arrow"
            else:
                button_color = "#FFB6C1"  # Light pink for no value
                button_state = "disabled"
                cursor_type = "arrow"
            
            # Create button (much wider to fill all available space)
            spec_btn = tk.Button(parent_frame, 
                               text=f"{spec_name}: {spec_value or 'No Data'}",
                               state=button_state,
                               command=lambda file=heater_design_file: self.open_heater_design_file(file) if file else None,
                               width=60, height=1,
                               font=('Arial', 9),
                               relief='raised', bd=1,
                               cursor=cursor_type,
                               bg=button_color, fg="black",
                               activebackground=button_color, activeforeground="black")
            spec_btn.grid(row=18+i, column=0, sticky=(tk.W, tk.E), pady=1, padx=(0, 10))
            
            # Add right-click context menu
            self.create_spec_context_menu(spec_btn, spec_name, parent_frame, 18+i)
            
            self.spec_buttons.append(spec_btn)
            
            # Add input field for missing values (all "No Data" or "No [Item] Found" cases)
            if spec_value in ["No Spray Nozzle Found", "No Size Found", "No Length Found", "No Data"] or not spec_value:
                self.create_spec_input_field(parent_frame, spec_name, 18+i, 1)  # Column 1 for right side
    
    def create_spec_context_menu(self, button, spec_name, parent_frame, row):
        """Create right-click context menu for specification buttons"""
        def show_context_menu(event):
            context_menu = tk.Menu(self.root, tearoff=0)
            
            # Add Edit option
            context_menu.add_command(label="Edit", 
                                   command=lambda: self.edit_spec_value(spec_name, parent_frame, row))
            
            # Add Delete option (only if manual value exists)
            if self.get_saved_manual_spec(spec_name):
                context_menu.add_command(label="Delete", 
                                       command=lambda: self.delete_manual_spec(spec_name))
            
            # Add separator and refresh option
            context_menu.add_separator()
            context_menu.add_command(label="Refresh", 
                                   command=lambda: self.update_specifications(parent_frame))
            
            try:
                context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                context_menu.grab_release()
        
        button.bind("<Button-3>", show_context_menu)  # Right-click
    
    def edit_spec_value(self, spec_name, parent_frame, row):
        """Edit a specification value"""
        current_value = self.get_saved_manual_spec(spec_name)
        if current_value:
            # Remove "Manual:" prefix if present
            if current_value.startswith("Manual: "):
                current_value = current_value[8:]
        else:
            current_value = ""
        
        # Create edit dialog
        edit_window = tk.Toplevel(self.root)
        edit_window.title(f"Edit {spec_name}")
        edit_window.geometry("400x150")
        edit_window.transient(self.root)
        edit_window.grab_set()
        
        # Center the window
        edit_window.update_idletasks()
        x = (edit_window.winfo_screenwidth() // 2) - (400 // 2)
        y = (edit_window.winfo_screenheight() // 2) - (150 // 2)
        edit_window.geometry(f"400x150+{x}+{y}")
        
        # Label
        ttk.Label(edit_window, text=f"Enter new value for {spec_name}:", font=('Arial', 10)).pack(pady=10)
        
        # Input field
        input_var = tk.StringVar(value=current_value)
        input_entry = ttk.Entry(edit_window, textvariable=input_var, width=40, font=('Arial', 10))
        input_entry.pack(pady=10)
        input_entry.focus()
        input_entry.select_range(0, tk.END)
        
        # Buttons
        button_frame = tk.Frame(edit_window)
        button_frame.pack(pady=10)
        
        tk.Button(button_frame, text="Save", 
                command=lambda: self.save_edit_spec(spec_name, input_var.get(), edit_window),
                width=10, bg='#4CAF50', fg='white').pack(side=tk.LEFT, padx=5)
        
        tk.Button(button_frame, text="Cancel", 
                command=edit_window.destroy,
                width=10, bg='#f44336', fg='white').pack(side=tk.LEFT, padx=5)
        
        # Bind Enter key to save
        input_entry.bind('<Return>', lambda e: self.save_edit_spec(spec_name, input_var.get(), edit_window))
    
    def save_edit_spec(self, spec_name, value, window):
        """Save edited specification value"""
        if not value.strip():
            messagebox.showwarning("Warning", f"Please enter a value for {spec_name}")
            return
        
        self.save_manual_spec(spec_name, value)
        window.destroy()
    
    def delete_manual_spec(self, spec_name):
        """Delete a manual specification value for the current job"""
        job_number = str(self.job_number_var.get()).strip()
        if not job_number:
            return
        conn = sqlite3.connect(self.db_manager.db_path)
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS manual_specs (
                job_number TEXT NOT NULL,
                spec_name TEXT NOT NULL,
                value TEXT NOT NULL,
                PRIMARY KEY(job_number, spec_name)
            )
        """)
        cur.execute("DELETE FROM manual_specs WHERE job_number = ? AND spec_name = ?", (job_number, spec_name))
        conn.commit()
        conn.close()
        messagebox.showinfo("Deleted", f"{spec_name} manual value deleted")
        # Refresh the specifications
        # Refresh specifications using stable reference
        if hasattr(self, 'project_details_frame'):
            self.update_specifications(self.project_details_frame)
    
    def create_spec_input_field(self, parent_frame, spec_name, row, column=1):
        """Create an input field for manual entry of missing specifications"""
        # Create input frame for right side - positioned next to the button
        input_frame = tk.Frame(parent_frame)
        input_frame.grid(row=row, column=column, sticky=(tk.W, tk.E), pady=1, padx=(0, 0))
        
        # Input field (wider to use more space)
        input_var = tk.StringVar()
        input_entry = ttk.Entry(input_frame, textvariable=input_var, width=25, font=('Arial', 9))
        input_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # Save button (wider for better visibility)
        save_btn = tk.Button(input_frame, text="Save", 
                           command=lambda: self.save_manual_spec(spec_name, input_var.get()),
                           width=8, height=1, font=('Arial', 9, 'bold'),
                           bg='#4CAF50', fg='white', relief='raised', bd=1)
        save_btn.pack(side=tk.LEFT, padx=(0, 0))
        
        # Store the input field for later reference
        self.spec_input_fields[spec_name] = input_frame
        
        # Load any previously saved value
        saved_value = self.get_saved_manual_spec(spec_name)
        if saved_value:
            input_var.set(saved_value)
        else:
            # Add placeholder text
            input_entry.insert(0, f"Enter value...")
            input_entry.bind('<FocusIn>', lambda e: input_entry.delete(0, tk.END) if input_entry.get() == f"Enter value..." else None)
            input_entry.bind('<FocusOut>', lambda e: input_entry.insert(0, f"Enter value...") if not input_entry.get() else None)
    
    def save_manual_spec(self, spec_name, value):
        """Save a manually entered specification value for the current job"""
        if not value.strip():
            messagebox.showwarning("Warning", f"Please enter a value for {spec_name}")
            return
        job_number = str(self.job_number_var.get()).strip()
        if not job_number:
            messagebox.showwarning("Warning", "No job number selected")
            return
        conn = sqlite3.connect(self.db_manager.db_path)
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS manual_specs (
                job_number TEXT NOT NULL,
                spec_name TEXT NOT NULL,
                value TEXT NOT NULL,
                PRIMARY KEY(job_number, spec_name)
            )
        """)
        cur.execute("INSERT OR REPLACE INTO manual_specs (job_number, spec_name, value) VALUES (?, ?, ?)",
                    (job_number, spec_name, value.strip()))
        conn.commit()
        conn.close()
        messagebox.showinfo("Saved", f"{spec_name} saved as: {value.strip()}")
        # Update the specifications to show the saved value
        if hasattr(self, 'project_details_frame'):
            self.update_specifications(self.project_details_frame)
    
    def get_saved_manual_spec(self, spec_name):
        """Get a previously saved manual specification value for the current job"""
        job_number = str(self.job_number_var.get()).strip()
        if not job_number:
            return None
        conn = sqlite3.connect(self.db_manager.db_path)
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS manual_specs (
                job_number TEXT NOT NULL,
                spec_name TEXT NOT NULL,
                value TEXT NOT NULL,
                PRIMARY KEY(job_number, spec_name)
            )
        """)
        cur.execute("SELECT value FROM manual_specs WHERE job_number = ? AND spec_name = ?", (job_number, spec_name))
        row = cur.fetchone()
        conn.close()
        return row[0] if row else None
    
    def read_heater_spec_value(self, file_path, cell_ref):
        """Read a specific cell value from the Heater Cross Section sheet"""
        if not file_path or not os.path.exists(file_path):
            return None
            
        try:
            import openpyxl
            
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Look for "Heater Cross Section" sheet
            cross_section_sheet = None
            for sheet_name in workbook.sheetnames:
                if "Heater Cross Section" in sheet_name:
                    cross_section_sheet = workbook[sheet_name]
                    break
            
            if not cross_section_sheet:
                print(f"No 'Heater Cross Section' sheet found in {file_path}")
                return None
            
            # Get the value from the specified cell
            cell_value = cross_section_sheet[cell_ref].value
            if cell_value is not None:
                print(f"Found {cell_ref}: {cell_value}")
                return str(cell_value)
            else:
                print(f"No value found in cell {cell_ref}")
                return None
            
        except Exception as e:
            print(f"Error reading cell {cell_ref} from {file_path}: {e}")
            return None
    
    def read_spray_nozzle_pn_from_files(self):
        """Read Spray Nozzle P/N by getting nozzle specs from Engineering Design file and looking up in Spray Nozzles file"""
        # First, find the Engineering Design file to get nozzle size and length
        engineering_design_file = None
        spray_nozzles_file = None
        
        if hasattr(self, 'engineering_general_docs') and self.engineering_general_docs:
            for file_path in self.engineering_general_docs:
                filename = os.path.basename(file_path).upper()
                if "ENGINEERING DESIGN" in filename:
                    engineering_design_file = file_path
                    print(f"Found Engineering Design file: {filename}")
                elif "SPRAY NOZZLES" in filename:
                    spray_nozzles_file = file_path
                    print(f"Found Spray Nozzles file: {filename}")
        
        if not engineering_design_file:
            print("No Engineering Design file found")
            return None
            
        if not spray_nozzles_file:
            print("No Spray Nozzles file found")
            return None
        
        # Get nozzle size and length from Engineering Design file
        nozzle_size = self.get_nozzle_size_from_heater_design(engineering_design_file)
        nozzle_length = self.get_nozzle_length_from_heater_design(engineering_design_file)
        
        if not nozzle_size or not nozzle_length:
            print(f"Could not get Nozzle Size or Length from Engineering Design file")
            return None
        
        # Now look up the part numbers in the Spray Nozzles file
        return self.lookup_spray_nozzle_pn(spray_nozzles_file, nozzle_size, nozzle_length)
    
    def get_spray_nozzle_size_from_engineering_design(self):
        """Get Spray Nozzle Size from Engineering Design file"""
        if hasattr(self, 'engineering_general_docs') and self.engineering_general_docs:
            for file_path in self.engineering_general_docs:
                filename = os.path.basename(file_path).upper()
                if "ENGINEERING DESIGN" in filename:
                    return self.get_nozzle_size_from_heater_design(file_path)
        return None
    
    def get_spray_nozzle_length_from_engineering_design(self):
        """Get Spray Nozzle Length from Engineering Design file"""
        if hasattr(self, 'engineering_general_docs') and self.engineering_general_docs:
            for file_path in self.engineering_general_docs:
                filename = os.path.basename(file_path).upper()
                if "ENGINEERING DESIGN" in filename:
                    return self.get_nozzle_length_from_heater_design(file_path)
        return None
    
    def lookup_spray_nozzle_pn(self, spray_nozzles_file, nozzle_size, nozzle_length):
        """Look up Spray Nozzle P/N in the Spray Nozzles file using size and length"""
        if not spray_nozzles_file or not os.path.exists(spray_nozzles_file):
            return None
            
        try:
            import openpyxl
            
            # Load the Spray Nozzles workbook
            workbook = openpyxl.load_workbook(spray_nozzles_file, data_only=True)
            
            # Look for "Spray Nozzles" sheet or "Nozzle Selection" sheet or use Sheet1
            spray_nozzles_sheet = None
            for sheet_name in workbook.sheetnames:
                if "Spray Nozzles" in sheet_name or "Nozzle Selection" in sheet_name:
                    spray_nozzles_sheet = workbook[sheet_name]
                    print(f"Found sheet: {sheet_name}")
                    break
            
            # If no "Spray Nozzles" or "Nozzle Selection" sheet found, try Sheet1
            if not spray_nozzles_sheet and "Sheet1" in workbook.sheetnames:
                spray_nozzles_sheet = workbook["Sheet1"]
                print(f"Using Sheet1 for Spray Nozzles data")
            
            if not spray_nozzles_sheet:
                print(f"No 'Spray Nozzles', 'Nozzle Selection', or 'Sheet1' sheet found in {spray_nozzles_file}")
                return None
            
            # Search for matching row based on Nozzle Size (Column A) and Nozzle Length (Column P)
            print(f"Searching for matching row in {spray_nozzles_sheet.max_row} rows...")
            print(f"Looking for Nozzle Size: {nozzle_size}, Nozzle Length: {nozzle_length}")
            
            for row in range(1, spray_nozzles_sheet.max_row + 1):
                cell_a = spray_nozzles_sheet[f'A{row}']  # Nozzle Size
                cell_p = spray_nozzles_sheet[f'P{row}']  # Nozzle Length
                
                # Debug: Print what we find in each row
                if cell_a.value or cell_p.value:
                    print(f"Row {row}: A='{cell_a.value}', P='{cell_p.value}'")
                
                # Check if both cells have values and match our criteria
                if cell_a.value and cell_p.value:
                    # Convert to strings and compare
                    size_match = str(cell_a.value).strip() == str(nozzle_size).strip()
                    length_match = str(cell_p.value).strip() == str(nozzle_length).strip()
                    
                    print(f"Row {row} - Size match: {size_match}, Length match: {length_match}")
                    
                    if size_match and length_match:
                        # Found the matching row! Get the part numbers from columns B and C
                        cell_b = spray_nozzles_sheet[f'B{row}']
                        cell_c = spray_nozzles_sheet[f'C{row}']
                        
                        if cell_b.value and cell_c.value:
                            column_b_value = str(cell_b.value).strip()
                            column_c_value = str(cell_c.value).strip()
                            spray_pn = f"{column_b_value}-{column_c_value}"
                            print(f"Found Spray Nozzle P/N: {spray_pn} in matching row {row}")
                            return spray_pn
            
            print(f"No matching row found for Nozzle Size: {nozzle_size}, Length: {nozzle_length}")
            return None
            
        except Exception as e:
            print(f"Error looking up Spray Nozzle P/N: {e}")
            return None
    
    def read_spray_nozzle_pn(self, file_path):
        """Read Spray Nozzle P/N by matching Nozzle Size and Length from Heater Design sheet"""
        if not file_path or not os.path.exists(file_path):
            return None
            
        try:
            import openpyxl
            
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # First, get the Nozzle Size and Length from the Heater Design sheet
            nozzle_size = self.get_nozzle_size_from_heater_design(file_path)
            nozzle_length = self.get_nozzle_length_from_heater_design(file_path)
            
            if not nozzle_size or not nozzle_length:
                print(f"Could not get Nozzle Size or Length from Heater Design sheet")
                return None
            
            print(f"Looking for Nozzle Size: {nozzle_size}, Nozzle Length: {nozzle_length}")
            
            # Look for "Spray Nozzles" sheet or "Nozzle Selection" sheet or use Sheet1
            spray_nozzles_sheet = None
            for sheet_name in workbook.sheetnames:
                if "Spray Nozzles" in sheet_name or "Nozzle Selection" in sheet_name:
                    spray_nozzles_sheet = workbook[sheet_name]
                    print(f"Found sheet: {sheet_name}")
                    break
            
            # If no "Spray Nozzles" or "Nozzle Selection" sheet found, try Sheet1
            if not spray_nozzles_sheet and "Sheet1" in workbook.sheetnames:
                spray_nozzles_sheet = workbook["Sheet1"]
                print(f"Using Sheet1 for Spray Nozzles data")
            
            if not spray_nozzles_sheet:
                print(f"No 'Spray Nozzles', 'Nozzle Selection', or 'Sheet1' sheet found in {file_path}")
                return None
            
            # Search for matching row based on Nozzle Size (Column A) and Nozzle Length (Column P)
            print(f"Searching for matching row in {spray_nozzles_sheet.max_row} rows...")
            
            for row in range(1, spray_nozzles_sheet.max_row + 1):
                cell_a = spray_nozzles_sheet[f'A{row}']  # Nozzle Size
                cell_p = spray_nozzles_sheet[f'P{row}']  # Nozzle Length
                
                # Debug: Print what we find in each row
                if cell_a.value or cell_p.value:
                    print(f"Row {row}: A='{cell_a.value}', P='{cell_p.value}'")
                
                # Check if both cells have values and match our criteria
                if cell_a.value and cell_p.value:
                    # Convert to strings and compare
                    size_match = str(cell_a.value).strip() == str(nozzle_size).strip()
                    length_match = str(cell_p.value).strip() == str(nozzle_length).strip()
                    
                    print(f"Row {row} - Size match: {size_match}, Length match: {length_match}")
                    
                    if size_match and length_match:
                        # Found the matching row! Get the part numbers from columns B and C
                        cell_b = spray_nozzles_sheet[f'B{row}']
                        cell_c = spray_nozzles_sheet[f'C{row}']
                        
                        if cell_b.value and cell_c.value:
                            column_b_value = str(cell_b.value).strip()
                            column_c_value = str(cell_c.value).strip()
                            spray_pn = f"{column_b_value}-{column_c_value}"
                            print(f"Found Spray Nozzle P/N: {spray_pn} in matching row {row}")
                            return spray_pn
            
            print(f"No matching row found for Nozzle Size: {nozzle_size}, Length: {nozzle_length}")
            return None
            
        except Exception as e:
            print(f"Error reading Spray Nozzle P/N from {file_path}: {e}")
            return None
    
    def get_nozzle_size_from_heater_design(self, file_path):
        """Get Nozzle Size from L22 in Heater Design sheet"""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Look for "Heater Design" sheet
            heater_sheet = None
            for sheet_name in workbook.sheetnames:
                if "Heater Design" in sheet_name:
                    heater_sheet = workbook[sheet_name]
                    break
            
            if not heater_sheet:
                print(f"No 'Heater Design' sheet found for nozzle size")
                return None
            
            # Get value from L22
            cell_l22 = heater_sheet['L22']
            if cell_l22.value:
                nozzle_size = str(cell_l22.value).strip()
                print(f"Found Nozzle Size: {nozzle_size}")
                return nozzle_size
            
            return None
            
        except Exception as e:
            print(f"Error getting nozzle size: {e}")
            return None
    
    def get_nozzle_length_from_heater_design(self, file_path):
        """Get Nozzle Length from L21 in Heater Design sheet"""
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Look for "Heater Design" sheet
            heater_sheet = None
            for sheet_name in workbook.sheetnames:
                if "Heater Design" in sheet_name:
                    heater_sheet = workbook[sheet_name]
                    break
            
            if not heater_sheet:
                print(f"No 'Heater Design' sheet found for nozzle length")
                return None
            
            # Get value from L21
            cell_l21 = heater_sheet['L21']
            if cell_l21.value:
                nozzle_length = str(cell_l21.value).strip()
                print(f"Found Nozzle Length: {nozzle_length}")
                return nozzle_length
            
            return None
            
        except Exception as e:
            print(f"Error getting nozzle length: {e}")
            return None
    
    def has_heater_design_sheet(self, file_path):
        """Check if an Excel file has a 'Heater Design' sheet inside it"""
        try:
            import openpyxl
            
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Check if any sheet name contains "Heater Design"
            for sheet_name in workbook.sheetnames:
                if "Heater Design" in sheet_name:
                    print(f"Found 'Heater Design' sheet in file: {os.path.basename(file_path)}")
                    return True
            
            return False
            
        except Exception as e:
            print(f"Error checking for Heater Design sheet in {file_path}: {e}")
            return False
    
    def read_excel_can_size(self, file_path):
        """Read Can Size from Excel file (column I, value from column L) on 'Heater Design' sheet"""
        try:
            import openpyxl
            
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Look for "Heater Design" sheet specifically
            heater_sheet = None
            for sheet_name in workbook.sheetnames:
                if "Heater Design" in sheet_name:
                    heater_sheet = workbook[sheet_name]
                    break
            
            if not heater_sheet:
                print(f"No 'Heater Design' sheet found in {file_path}")
                return None
            
            # Method 1: Look for "Can Size:" in column I and get value from column L
            for row in range(1, heater_sheet.max_row + 1):
                cell_i = heater_sheet[f'I{row}']
                if cell_i.value and "Can Size" in str(cell_i.value):
                    # Get the value from column L in the same row
                    cell_l = heater_sheet[f'L{row}']
                    if cell_l.value:
                        print(f"Found Can Size: {cell_l.value} in row {row} (method 1)")
                        return str(cell_l.value)
            
            # Method 2: Check cell L48 directly (fallback)
            cell_l48 = heater_sheet['L48']
            if cell_l48.value:
                print(f"Found Can Size: {cell_l48.value} in cell L48 (method 2)")
                return str(cell_l48.value)
            
            print(f"No 'Can Size:' found in column I or L48 of Heater Design sheet")
            return None
            
        except Exception as e:
            print(f"Error reading Excel file {file_path}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def open_heater_design_file(self, file_path):
        """Open the Heater Design file"""
        if file_path and os.path.exists(file_path):
            self.open_path(file_path)
        else:
            messagebox.showwarning("Warning", "Heater Design file not found!")
    
    def create_workflow_panel(self):
        """Create the workflow tracking panel with collapsible sections"""
        # Main container
        main_container = ttk.Frame(self.workflow_container)
        main_container.pack(fill=tk.BOTH, expand=True, padx=(0, 5))
        main_container.columnconfigure(0, weight=1)
        
        # Toolbar at the top
        self.create_workflow_toolbar(main_container)
        
        # Scrollable workflow content
        canvas = tk.Canvas(main_container, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        workflow_frame = ttk.Frame(canvas)
        
        workflow_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas_window = canvas.create_window((0, 0), window=workflow_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        
        # Make canvas expand to fill width
        def _configure_canvas(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind('<Configure>', _configure_canvas)
        
        # Enable mouse wheel scrolling
        self._bind_mousewheel(canvas, workflow_frame)
        
        main_container.rowconfigure(1, weight=1)
        workflow_frame.columnconfigure(0, weight=1)
        
        # Section 0: Template-driven workflow (new)
        self.template_workflow_section = CollapsibleFrame(workflow_frame, "Standard Workflow (Template)")
        self.template_workflow_section.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        self.create_template_workflow_content(self.template_workflow_section.content)
        # Legacy hardcoded workflow sections removed per new template-driven workflow
    
    def create_workflow_toolbar(self, parent):
        """Create toolbar with print button"""
        toolbar = ttk.Frame(parent, relief='raised', padding="5")
        toolbar.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        toolbar.columnconfigure(0, weight=1)
        
        # Title on the left
        title_label = ttk.Label(toolbar, text="Project Workflow", 
                               font=('Arial', 12, 'bold'))
        title_label.grid(row=0, column=0, sticky=tk.W)
        
        # Right-side actions
        actions_frame = ttk.Frame(toolbar)
        actions_frame.grid(row=0, column=1, sticky=tk.E)

        settings_btn = ttk.Button(actions_frame, text="⚙️ Workflow Settings", command=self.open_workflow_settings)
        settings_btn.pack(side=tk.LEFT, padx=(0, 6))

        apply_btn = ttk.Button(actions_frame, text="↺ Apply Standard Workflow", command=self.apply_standard_workflow_to_current_project)
        apply_btn.pack(side=tk.LEFT, padx=(0, 6))

        self.cover_sheet_btn = ttk.Button(actions_frame, text="🖨️ Print Status Report", 
                                         command=self.print_cover_sheet,
                                         style='Accent.TButton')
        self.cover_sheet_btn.pack(side=tk.LEFT)

    def create_template_workflow_content(self, parent):
        """Render the template-driven workflow steps for the current project in a vertical layout"""
        for child in parent.winfo_children():
            child.destroy()
        parent.columnconfigure(0, weight=1)

        self.workflow_row_widgets = {}

        steps = self._load_project_workflow_steps()
        if not steps:
            msg = "No workflow steps defined. Use Workflow Settings to define and apply the Standard Workflow."
            ttk.Label(parent, text=msg, foreground='gray').grid(row=0, column=0, sticky=tk.W, pady=(6, 0))
            return

        people = self._get_people_list()

        row_i = 0
        for step in steps:
            sid = step['id']
            step_frame = ttk.LabelFrame(parent, text=step['department'])
            step_frame.grid(row=row_i, column=0, sticky=(tk.W, tk.E), padx=2, pady=4)
            step_frame.columnconfigure(1, weight=1)

            # Group and Title
            ttk.Label(step_frame, text=f"Group: {step['group_name'] or ''}", font=('Arial', 9, 'italic')).grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(2, 0))
            ttk.Label(step_frame, text=f"Title: {step['title']}", font=('Arial', 10, 'bold')).grid(row=1, column=0, columnspan=2, sticky=tk.W)

            # Detailed fields laid out vertically
            r = 2
            # Start
            start_var = tk.BooleanVar(value=bool(step['start_flag']))
            ttk.Label(step_frame, text="Start:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            start_chk = ttk.Checkbutton(step_frame, variable=start_var, command=lambda s=sid, v=start_var: self._on_start_toggle(s, v))
            start_chk.grid(row=r, column=1, sticky=tk.W)
            r += 1
            ttk.Label(step_frame, text="Started At:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            started_lbl = ttk.Label(step_frame, text=(step['start_ts'] or ''))
            started_lbl.grid(row=r, column=1, sticky=tk.W)
            r += 1

            # Completed
            comp_var = tk.BooleanVar(value=bool(step['completed_flag']))
            ttk.Label(step_frame, text="Completed:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            comp_chk = ttk.Checkbutton(step_frame, variable=comp_var, command=lambda s=sid, v=comp_var: self._on_completed_toggle(s, v))
            comp_chk.grid(row=r, column=1, sticky=tk.W)
            r += 1
            ttk.Label(step_frame, text="Completed At:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            completed_lbl = ttk.Label(step_frame, text=(step['completed_ts'] or ''))
            completed_lbl.grid(row=r, column=1, sticky=tk.W)
            r += 1

            # Transfer To / Received From
            to_var = tk.StringVar(value=step['transfer_to_name'] or '')
            ttk.Label(step_frame, text="Transfer To:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            to_combo = ttk.Combobox(step_frame, textvariable=to_var, values=people, width=18, state='readonly')
            to_combo.grid(row=r, column=1, sticky=tk.W)
            to_combo.bind('<<ComboboxSelected>>', lambda e, s=sid, v=to_var: self._on_transfer_set(s, v))
            r += 1
            ttk.Label(step_frame, text="Transfer Set At:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            to_ts_lbl = ttk.Label(step_frame, text=(step.get('transfer_to_ts') or ''))
            to_ts_lbl.grid(row=r, column=1, sticky=tk.W)
            r += 1

            from_var = tk.StringVar(value=step['received_from_name'] or '')
            ttk.Label(step_frame, text="Received From:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            from_combo = ttk.Combobox(step_frame, textvariable=from_var, values=people, width=18, state='readonly')
            from_combo.grid(row=r, column=1, sticky=tk.W)
            from_combo.bind('<<ComboboxSelected>>', lambda e, s=sid, v=from_var: self._on_received_set(s, v))
            r += 1
            ttk.Label(step_frame, text="Received Set At:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            from_ts_lbl = ttk.Label(step_frame, text=(step.get('received_from_ts') or ''))
            from_ts_lbl.grid(row=r, column=1, sticky=tk.W)
            r += 1

            # Due and Duration
            ttk.Label(step_frame, text="Due:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            due_lbl = ttk.Label(step_frame, text=(step['planned_due_date'] or ''))
            due_lbl.grid(row=r, column=1, sticky=tk.W)
            r += 1

            ttk.Label(step_frame, text="Duration:").grid(row=r, column=0, sticky=tk.W, padx=(2, 6))
            _adur_min = step.get('actual_duration_minutes')
            _pdur_min = step.get('planned_duration_minutes') or (int(step.get('planned_duration_days') or 0) * 1440)
            dur_lbl = ttk.Label(step_frame, text=format_minutes_compact(_adur_min if _adur_min is not None else _pdur_min))
            dur_lbl.grid(row=r, column=1, sticky=tk.W)

            self.workflow_row_widgets[sid] = {
                'start_var': start_var, 'start_lbl': started_lbl,
                'comp_var': comp_var, 'comp_lbl': completed_lbl,
                'to_var': to_var, 'from_var': from_var,
                'due_lbl': due_lbl, 'dur_lbl': dur_lbl
            }
            row_i += 1

    def _get_people_list(self):
        """Return a consolidated list of names from designers, engineers, plus Production 'Larry W.'"""
        names = set()
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute("SELECT name FROM designers")
            for r in cur.fetchall():
                if r and r[0]:
                    names.add(r[0])
            cur.execute("SELECT name FROM engineers")
            for r in cur.fetchall():
                if r and r[0]:
                    names.add(r[0])
            conn.close()
        except Exception:
            pass
        # Ensure Production name present
        names.add('Larry W.')
        return sorted(names)

    def _load_project_workflow_steps(self):
        """Load workflow steps for current project (with planned durations from template)."""
        if not self.current_project:
            return []
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute("SELECT id FROM projects WHERE job_number = ?", (self.current_project,))
            pid_row = cur.fetchone()
            if not pid_row:
                conn.close()
                return []
            project_id = pid_row[0]
            try:
                cur.execute(
                    """
                    SELECT pws.id, pws.order_index, pws.department, pws.group_name, pws.title,
                           pws.start_flag, pws.start_ts, pws.completed_flag, pws.completed_ts,
                           pws.transfer_to_name, pws.received_from_name, pws.transfer_to_ts, pws.received_from_ts,
                           pws.planned_due_date, pws.actual_completed_date, pws.actual_duration_days, pws.actual_duration_minutes,
                           wts.planned_duration_days, wts.planned_duration_minutes
                    FROM project_workflow_steps pws
                    LEFT JOIN workflow_template_steps wts ON pws.template_step_id = wts.id
                    WHERE pws.project_id = ?
                    ORDER BY pws.order_index
                    """,
                    (project_id,)
                )
                rows = cur.fetchall()
            except Exception:
                # Fallback if timestamp columns are missing on older DBs
                cur.execute(
                    """
                    SELECT pws.id, pws.order_index, pws.department, pws.group_name, pws.title,
                           pws.start_flag, pws.start_ts, pws.completed_flag, pws.completed_ts,
                           pws.transfer_to_name, pws.received_from_name,
                           pws.planned_due_date, pws.actual_completed_date, pws.actual_duration_days,
                           wts.planned_duration_days
                    FROM project_workflow_steps pws
                    LEFT JOIN workflow_template_steps wts ON pws.template_step_id = wts.id
                    WHERE pws.project_id = ?
                    ORDER BY pws.order_index
                    """,
                    (project_id,)
                )
                rows = cur.fetchall()
            conn.close()
            steps = []
            for r in rows:
                # Support both shapes depending on DB schema
                if len(r) >= 19:
                    steps.append({
                        'id': r[0], 'order_index': r[1], 'department': r[2], 'group_name': r[3], 'title': r[4],
                        'start_flag': r[5], 'start_ts': r[6], 'completed_flag': r[7], 'completed_ts': r[8],
                        'transfer_to_name': r[9], 'received_from_name': r[10], 'transfer_to_ts': r[11], 'received_from_ts': r[12],
                        'planned_due_date': r[13], 'actual_completed_date': r[14], 'actual_duration_days': r[15], 'actual_duration_minutes': r[16],
                        'planned_duration_days': r[17], 'planned_duration_minutes': r[18]
                    })
                else:
                    steps.append({
                        'id': r[0], 'order_index': r[1], 'department': r[2], 'group_name': r[3], 'title': r[4],
                        'start_flag': r[5], 'start_ts': r[6], 'completed_flag': r[7], 'completed_ts': r[8],
                        'transfer_to_name': r[9], 'received_from_name': r[10],
                        'planned_due_date': r[11], 'actual_completed_date': r[12], 'actual_duration_days': r[13],
                        'planned_duration_days': r[14]
                    })
            return steps
        except Exception as e:
            print(f"Error loading project workflow steps: {e}")
            return []

    def _on_start_toggle(self, step_id, var):
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            # Guard: ensure the step belongs to the selected project
            cur.execute("SELECT project_id FROM project_workflow_steps WHERE id = ?", (step_id,))
            row = cur.fetchone()
            if not row:
                print(f"DEBUG[_on_start_toggle]: step {step_id} not found")
                conn.close(); return
            step_pid = row[0]
            cur.execute("SELECT id FROM projects WHERE job_number = ?", (self.current_project,))
            prow = cur.fetchone()
            if not prow or prow[0] != step_pid:
                print(f"DEBUG[_on_start_toggle]: mismatch current_pid={prow[0] if prow else None} step_pid={step_pid} current_project={self.current_project}")
                conn.close(); return
            # Retain timestamp if exists when unchecking
            if var.get():
                cur.execute(
                    "UPDATE project_workflow_steps SET start_flag = 1, start_ts = COALESCE(start_ts, ?) WHERE id = ? AND project_id = ?",
                    (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), step_id, step_pid)
                )
            else:
                cur.execute("UPDATE project_workflow_steps SET start_flag = 0 WHERE id = ? AND project_id = ?", (step_id, step_pid))
            conn.commit()
            print(f"DEBUG[_on_start_toggle]: updated step_id={step_id} project_id={step_pid} new_value={int(var.get())}")
            conn.close()
        except Exception as e:
            print(f"Error updating start flag: {e}")
        self._refresh_template_workflow_after_change()

    def _on_completed_toggle(self, step_id, var):
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute("SELECT project_id FROM project_workflow_steps WHERE id = ?", (step_id,))
            row = cur.fetchone()
            if not row:
                conn.close(); return
            step_pid = row[0]
            cur.execute("SELECT id FROM projects WHERE job_number = ?", (self.current_project,))
            prow = cur.fetchone()
            if not prow or prow[0] != step_pid:
                conn.close(); return
            if var.get():
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                # Compute actual duration if start exists
                cur.execute("SELECT start_ts FROM project_workflow_steps WHERE id = ? AND project_id = ?", (step_id, step_pid))
                row = cur.fetchone()
                actual_dur_min = None
                try:
                    if row and row[0]:
                        sdt = datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S")
                        delta = (datetime.now() - sdt)
                        actual_dur_min = int(delta.total_seconds() // 60)
                except Exception:
                    actual_dur_min = None
                actual_dur_days = int(round((actual_dur_min or 0) / (24 * 60))) if actual_dur_min is not None else None
                cur.execute(
                    """
                    UPDATE project_workflow_steps
                    SET completed_flag = 1,
                        completed_ts = COALESCE(completed_ts, ?),
                        actual_completed_date = COALESCE(actual_completed_date, DATE(?)),
                        actual_duration_days = COALESCE(actual_duration_days, ?),
                        actual_duration_minutes = COALESCE(actual_duration_minutes, ?)
                    WHERE id = ? AND project_id = ?
                    """,
                    (now, now, actual_dur_days, actual_dur_min, step_id, step_pid)
                )
            else:
                cur.execute("UPDATE project_workflow_steps SET completed_flag = 0 WHERE id = ? AND project_id = ?", (step_id, step_pid))
            conn.commit()
            print(f"DEBUG[_on_completed_toggle]: updated step_id={step_id} project_id={step_pid} new_value={int(var.get())}")
            conn.close()
        except Exception as e:
            print(f"Error updating completed flag: {e}")
        self._refresh_template_workflow_after_change()

    def _on_transfer_set(self, step_id, var):
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute("SELECT project_id FROM project_workflow_steps WHERE id = ?", (step_id,))
            row = cur.fetchone()
            if not row:
                conn.close(); return
            step_pid = row[0]
            cur.execute(
                "UPDATE project_workflow_steps SET transfer_to_name = ?, transfer_to_ts = COALESCE(transfer_to_ts, ?) WHERE id = ? AND project_id = ?",
                (var.get(), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), step_id, step_pid)
            )
            conn.commit(); conn.close()
            print(f"DEBUG[_on_transfer_set]: step_id={step_id} project_id={step_pid} name='{var.get()}'")
        except Exception as e:
            print(f"Error setting transfer_to: {e}")
        self._refresh_template_workflow_after_change(recalc=False)

    def _on_received_set(self, step_id, var):
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute("SELECT project_id FROM project_workflow_steps WHERE id = ?", (step_id,))
            row = cur.fetchone()
            if not row:
                conn.close(); return
            step_pid = row[0]
            cur.execute(
                "UPDATE project_workflow_steps SET received_from_name = ?, received_from_ts = COALESCE(received_from_ts, ?) WHERE id = ? AND project_id = ?",
                (var.get(), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), step_id, step_pid)
            )
            conn.commit(); conn.close()
            print(f"DEBUG[_on_received_set]: step_id={step_id} project_id={step_pid} name='{var.get()}'")
        except Exception as e:
            print(f"Error setting received_from: {e}")
        self._refresh_template_workflow_after_change(recalc=False)

    def _refresh_template_workflow_after_change(self, recalc=True):
        # Recompute due dates if needed and refresh UI
        if recalc:
            try:
                self._recompute_workflow_due_dates_for_current_project()
            except Exception:
                pass
        if hasattr(self, 'template_workflow_section'):
            self.create_template_workflow_content(self.template_workflow_section.content)

    def _recompute_workflow_due_dates_for_current_project(self):
        if not self.current_project:
            return
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            # Get project id and due date
            cur.execute("SELECT id, due_date FROM projects WHERE job_number = ?", (self.current_project,))
            row = cur.fetchone()
            if not row:
                conn.close(); return
            project_id, proj_due = row[0], row[1]
            if not proj_due:
                conn.close(); return
            # Load steps and durations
            cur.execute(
                """
                SELECT pws.id, pws.order_index, pws.start_ts,
                       wts.planned_duration_minutes
                FROM project_workflow_steps pws
                LEFT JOIN workflow_template_steps wts ON pws.template_step_id = wts.id
                WHERE pws.project_id = ?
                ORDER BY pws.order_index
                """,
                (project_id,),
            )
            steps = cur.fetchall()
            if not steps:
                conn.close(); return

            # Helper to subtract business days (Mon-Fri)
            def subtract_business_days(date_obj, days):
                import datetime as _dt
                d = date_obj
                remaining = int(days or 0)
                while remaining > 0:
                    d = d - _dt.timedelta(days=1)
                    if d.weekday() < 5:  # Mon-Fri
                        remaining -= 1
                return d

            from datetime import datetime as _dt
            next_start_planned = _dt.strptime(proj_due, "%Y-%m-%d").date()

            # Build list for reverse calc
            steps_rev = list(reversed(steps))
            updates = []
            for sid, order_i, start_ts, dur_min in steps_rev:
                # If next step has actual start, use that as this due
                planned_due = next_start_planned
                # planned start based on duration
                days_to_subtract = ceil_minutes_to_business_days(dur_min or 0)
                planned_start = subtract_business_days(planned_due, days_to_subtract)
                # For the next iteration, determine next_start_planned:
                # If this step has actual start, that's the next planned for previous step; else planned_start
                if start_ts:
                    try:
                        next_start_planned = _dt.strptime(start_ts, "%Y-%m-%d %H:%M:%S").date()
                    except Exception:
                        next_start_planned = planned_start
                else:
                    next_start_planned = planned_start
                updates.append((sid, planned_due.strftime("%Y-%m-%d")))

            # Apply updates
            for sid, due in updates:
                cur.execute("UPDATE project_workflow_steps SET planned_due_date = ? WHERE id = ?", (due, sid))
            conn.commit(); conn.close()
        except Exception as e:
            print(f"Error recomputing due dates: {e}")

    def open_workflow_settings(self):
        """Open the Standard Workflow settings editor (versioned templates)."""
        win = tk.Toplevel(self.root)
        win.title("Standard Workflow Settings")
        win.geometry("900x500")
        win.transient(self.root)
        win.grab_set()

        main = ttk.Frame(win, padding=10)
        main.pack(fill=tk.BOTH, expand=True)
        main.columnconfigure(0, weight=1)

        # Header
        ttk.Label(main, text="Standard Workflow (Active Template)", font=('Arial', 12, 'bold')).grid(row=0, column=0, sticky=tk.W)

        # Steps tree
        cols = ("Order", "Department", "Group", "Title", "Duration (m/h/d/w)")
        tree = ttk.Treeview(main, columns=cols, show='headings', height=16)
        for c in cols:
            tree.heading(c, text=c)
        tree.column("Order", width=60, anchor='center')
        tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(6, 6))
        main.rowconfigure(1, weight=1)

        # Controls
        controls = ttk.Frame(main)
        controls.grid(row=2, column=0, sticky=tk.W)
        add_before_btn = ttk.Button(controls, text="Add Before", command=lambda: self._wf_add_step(tree, before=True))
        add_after_btn = ttk.Button(controls, text="Add After", command=lambda: self._wf_add_step(tree, before=False))
        delete_btn = ttk.Button(controls, text="Delete", command=lambda: self._wf_delete_step(tree))
        up_btn = ttk.Button(controls, text="Move Up", command=lambda: self._wf_move(tree, -1))
        down_btn = ttk.Button(controls, text="Move Down", command=lambda: self._wf_move(tree, +1))
        save_btn = ttk.Button(controls, text="Save As New Version (Activate)", command=lambda: self._wf_save_new_version(tree, win))
        for b in (add_before_btn, add_after_btn, delete_btn, up_btn, down_btn, save_btn):
            b.pack(side=tk.LEFT, padx=4)

        # Load active template steps
        steps = self._wf_load_active_template_steps()
        for i, s in enumerate(steps):
            minutes = s.get('planned_duration_minutes')
            if minutes is None:
                minutes = int(s.get('planned_duration_days') or 1) * 1440
            tree.insert('', 'end', values=(i+1, s['department'], s['group_name'] or '', s['title'], format_minutes_compact(minutes)))

        # Enable in-place edit on double-click
        def on_double_click(event):
            item = tree.selection()
            if not item:
                return
            col = tree.identify_column(event.x)
            if col in ('#2', '#3', '#4', '#5'):
                self._wf_edit_cell(tree, item[0], col)
        tree.bind('<Double-1>', on_double_click)

    def _wf_load_active_template_steps(self):
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute("SELECT id, version FROM workflow_templates WHERE name = ? AND is_active = 1 ORDER BY version DESC LIMIT 1", ('Standard',))
            t = cur.fetchone()
            if not t:
                conn.close(); return []
            tid = t[0]
            cur.execute(
                """
                SELECT id, order_index, department, group_name, title, planned_duration_days, planned_duration_minutes
                FROM workflow_template_steps WHERE template_id = ? ORDER BY order_index
                """,
                (tid,),
            )
            rows = cur.fetchall(); conn.close()
            return [
                {
                    'id': r[0], 'order_index': r[1], 'department': r[2], 'group_name': r[3], 'title': r[4], 'planned_duration_days': r[5], 'planned_duration_minutes': r[6]
                } for r in rows
            ]
        except Exception as e:
            print(f"Error loading template steps: {e}")
            return []

    def _wf_add_step(self, tree, before=True):
        sel = tree.selection()
        insert_index = 0
        if sel:
            idx = tree.index(sel[0])
            insert_index = idx if before else idx + 1
        vals = (insert_index+1, "Drafting", "", "New Step", "1d")
        tree.insert('', insert_index, values=vals)
        # Re-number
        for i, iid in enumerate(tree.get_children()):
            v = list(tree.item(iid, 'values')); v[0] = i+1; tree.item(iid, values=v)

    def _wf_delete_step(self, tree):
        sel = tree.selection()
        if not sel:
            return
        tree.delete(sel[0])
        for i, iid in enumerate(tree.get_children()):
            v = list(tree.item(iid, 'values')); v[0] = i+1; tree.item(iid, values=v)

    def _wf_move(self, tree, delta):
        sel = tree.selection()
        if not sel:
            return
        iid = sel[0]
        idx = tree.index(iid)
        new_idx = idx + delta
        if new_idx < 0 or new_idx >= len(tree.get_children()):
            return
        tree.move(iid, '', new_idx)
        for i, ci in enumerate(tree.get_children()):
            v = list(tree.item(ci, 'values')); v[0] = i+1; tree.item(ci, values=v)

    def _wf_edit_cell(self, tree, item_id, col_id):
        # Simple inline editor: replace cell with entry and commit on Return
        x, y, w, h = tree.bbox(item_id, col_id)
        value = tree.set(item_id, col_id)
        entry = ttk.Entry(tree)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, value)
        entry.focus_set()

        def on_return(event):
            new_val = entry.get()
            entry.destroy()
            tree.set(item_id, col_id, new_val)
        entry.bind('<Return>', on_return)
        entry.bind('<FocusOut>', lambda e: entry.destroy())

    def _wf_save_new_version(self, tree, win):
        # Persist a new template version and activate it
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            # Compute next version
            cur.execute("SELECT COALESCE(MAX(version),0) FROM workflow_templates WHERE name = ?", ('Standard',))
            next_ver = (cur.fetchone()[0] or 0) + 1
            cur.execute("INSERT INTO workflow_templates (name, version, is_active, created_date) VALUES (?, ?, 1, ?)",
                        ('Standard', next_ver, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            new_tid = cur.lastrowid
            # Deactivate others
            cur.execute("UPDATE workflow_templates SET is_active = 0 WHERE name = ? AND id != ?", ('Standard', new_tid))
            # Insert steps
            for i, iid in enumerate(tree.get_children()):
                v = tree.item(iid, 'values')
                order_i = int(v[0]); dept = str(v[1]); group = str(v[2]); title = str(v[3]); raw = str(v[4])
                try:
                    dur_min = parse_duration_to_minutes(raw)
                except Exception:
                    dur_min = 1440  # default 1d
                dur_days = max(1, int((dur_min + 1440 - 1) // 1440))
                cur.execute(
                    """
                    INSERT INTO workflow_template_steps
                    (template_id, order_index, department, group_name, title, planned_duration_days, planned_duration_minutes)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (new_tid, order_i, dept, group, title, dur_days, dur_min)
                )
            conn.commit(); conn.close()
            messagebox.showinfo("Saved", f"Activated Standard v{next_ver}.")
            win.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save template: {e}")

    def apply_standard_workflow_to_current_project(self):
        """Seed or replace the current project's workflow from the active Standard template."""
        if not self.current_project:
            messagebox.showwarning("No Project", "Please select a project first.")
            return
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            # Get project id
            cur.execute("SELECT id FROM projects WHERE job_number = ?", (self.current_project,))
            row = cur.fetchone()
            if not row:
                conn.close(); messagebox.showerror("Error", "Project not found in database."); return
            project_id = row[0]
            # Count existing steps
            cur.execute("SELECT COUNT(*) FROM project_workflow_steps WHERE project_id = ?", (project_id,))
            existing = cur.fetchone()[0]
            # Load active template id
            cur.execute("SELECT id, version FROM workflow_templates WHERE name = ? AND is_active = 1 ORDER BY version DESC LIMIT 1", ('Standard',))
            trow = cur.fetchone()
            if not trow:
                conn.close(); messagebox.showwarning("No Template", "No active Standard workflow found. Define it in Workflow Settings."); return
            tid, tver = trow
            if existing > 0:
                if not messagebox.askyesno("Replace Workflow", f"This project already has {existing} workflow step(s).\nReplace with Standard v{tver}? This will remove all current step state."):
                    conn.close(); return
                cur.execute("DELETE FROM project_workflow_steps WHERE project_id = ?", (project_id,))
            # Insert steps from template
            cur.execute("""
                SELECT id, order_index, department, group_name, title
                FROM workflow_template_steps WHERE template_id = ? ORDER BY order_index
            """, (tid,))
            steps = cur.fetchall()
            for sid, order_i, dept, group_name, title in steps:
                cur.execute(
                    """
                    INSERT INTO project_workflow_steps
                    (project_id, template_id, template_step_id, order_index, department, group_name, title)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (project_id, tid, sid, order_i, dept, group_name, title)
                )
            conn.commit(); conn.close()
            # Recompute planned due dates and refresh UI
            self._recompute_workflow_due_dates_for_current_project()
            if hasattr(self, 'template_workflow_section'):
                self.create_template_workflow_content(self.template_workflow_section.content)
            messagebox.showinfo("Applied", f"Applied Standard v{tver} to project {self.current_project}.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to apply workflow: {e}")
    
    def create_cover_sheet_button(self, parent):
        """Create the cover sheet print button"""
        button_frame = ttk.Frame(parent)
        button_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        button_frame.columnconfigure(0, weight=1)
        
        # Cover sheet button - will be updated based on project status
        self.cover_sheet_btn = ttk.Button(button_frame, text="📄 Print Status Report", 
                                         command=self.print_cover_sheet, width=25)
        self.cover_sheet_btn.grid(row=0, column=0, pady=5)
        
        # Initialize button state
        self.update_cover_sheet_button()
    
    def print_cover_sheet(self):
        """Print the project cover sheet"""
        if not self.current_project:
            messagebox.showwarning("Warning", "Please select a project first!")
            return
        
        try:
            from project_cover_sheet import print_project_cover_sheet
            print_project_cover_sheet(self.current_project, self.db_manager)
        except ImportError:
            messagebox.showerror("Error", "Cover sheet module not found!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to print cover sheet: {str(e)}")
    
    def update_release_due_display(self, *args):
        """Update the Release to Dee due date display"""
        try:
            due_date_str = self.release_due_date_entry.get()
            if not due_date_str:
                self.release_due_display_var.set("")
                return
            
            # Parse the due date
            due_date = datetime.strptime(due_date_str, "%Y-%m-%d").date()
            today = datetime.now().date()
            
            # Calculate difference
            delta = (due_date - today).days
            
            if delta == 0:
                display_text = "Due TODAY"
                color = "red"
            elif delta > 0:
                display_text = f"Due in {delta} day{'s' if delta != 1 else ''}"
                color = "blue" if delta > 3 else "orange"
            else:
                display_text = f"Due {abs(delta)} day{'s' if abs(delta) != 1 else ''} ago"
                color = "red"
            
            self.release_due_display_var.set(display_text)
            self.release_due_display_label.config(foreground=color)
            
        except (ValueError, AttributeError):
            self.release_due_display_var.set("")
    
    def update_cover_sheet_button(self):
        """Update the cover sheet button appearance based on project status"""
        if not hasattr(self, 'cover_sheet_btn'):
            return
            
        if not self.current_project:
            self.cover_sheet_btn.config(text="📄 Print Status Report", 
                                      style="TButton")
            return
        
        # Check if there are recent updates that warrant a new report
        has_updates = self.check_for_recent_updates()
        
        if has_updates:
            # Make button stand out with different color/text
            self.cover_sheet_btn.config(text="🆕 Print NEW Status Report", 
                                      style="Accent.TButton")
            # Try to make it more visually distinct
            try:
                self.cover_sheet_btn.configure(background='#ff6b6b', foreground='white')
            except:
                pass  # Fallback if styling doesn't work
        else:
            self.cover_sheet_btn.config(text="📄 Print Status Report", 
                                      style="TButton")
            # Reset to default styling
            try:
                self.cover_sheet_btn.configure(background='', foreground='')
            except:
                pass
    
    def check_for_recent_updates(self):
        """Check if there are recent updates since last report"""
        if not self.current_project:
            return False
            
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            
            # Get the last cover sheet date
            cursor.execute("""
                SELECT last_cover_sheet_date FROM projects 
                WHERE job_number = ?
            """, (self.current_project,))
            result = cursor.fetchone()
            last_cover_sheet_date = result[0] if result and result[0] else None
            
            if not last_cover_sheet_date:
                # No cover sheet generated yet, so there are "updates"
                conn.close()
                return True
            
            # Check if any workflow data has been updated since last cover sheet
            from datetime import datetime
            last_date = datetime.strptime(last_cover_sheet_date, "%Y-%m-%d %H:%M:%S")
            
            # Get project ID
            cursor.execute("SELECT id FROM projects WHERE job_number = ?", (self.current_project,))
            project_id = cursor.fetchone()[0]
            
            # Check for recent updates in workflow tables
            tables_to_check = [
                ("initial_redline", "redline_date"),
                ("redline_updates", "update_date"),
                ("ops_review", "review_date"),
                ("peter_weck_review", "fixed_errors_date"),
                ("release_to_dee", "release_date")
            ]
            
            for table, date_column in tables_to_check:
                cursor.execute(f"""
                    SELECT {date_column} FROM {table} 
                    WHERE project_id = ? AND {date_column} IS NOT NULL
                """, (project_id,))
                dates = cursor.fetchall()
                
                for date_row in dates:
                    if date_row[0]:
                        try:
                            update_date = datetime.strptime(date_row[0], "%Y-%m-%d")
                            if update_date > last_date:
                                conn.close()
                                return True
                        except ValueError:
                            # Skip invalid dates
                            continue
            
            conn.close()
            return False
            
        except Exception as e:
            print(f"Error checking for updates: {e}")
            # Default to showing updates available
            return True
    
    def create_initial_redline_section(self, parent, row):
        """Create initial redline section"""
        section_frame = ttk.LabelFrame(parent, text="1. Drafting Drawing Package to Engineering for Initial Review", padding="5")
        section_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        section_frame.columnconfigure(1, weight=1)
        
        # Initial Redline checkbox
        self.initial_redline_var = tk.BooleanVar()
        ttk.Checkbutton(section_frame, text="Initial Redline", 
                       variable=self.initial_redline_var).grid(row=0, column=0, columnspan=2, sticky=tk.W)
        
        # Engineer dropdown
        ttk.Label(section_frame, text="Engineer:").grid(row=1, column=0, sticky=tk.W)
        self.initial_engineer_var = tk.StringVar()
        self.initial_engineer_combo = ttk.Combobox(section_frame, textvariable=self.initial_engineer_var, 
                                                 state="readonly", width=20)
        self.initial_engineer_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        
        # Date
        ttk.Label(section_frame, text="Date:").grid(row=2, column=0, sticky=tk.W)
        self.initial_date_entry = DateEntry(section_frame, width=20)
        self.initial_date_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
    
    def create_redline_updates_section(self, parent, row):
        """Create redline updates section"""
        section_frame = ttk.LabelFrame(parent, text="2. Redline Updates", padding="5")
        section_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        section_frame.columnconfigure(1, weight=1)
        
        # Update 1
        self.create_redline_update(section_frame, 0, "Redline Update 1")
        self.create_redline_update(section_frame, 1, "Redline Update 2")
        self.create_redline_update(section_frame, 2, "Redline Update 3")
        self.create_redline_update(section_frame, 3, "Redline Update 4")
    
    def create_redline_update(self, parent, row, title):
        """Create a single redline update"""
        update_frame = ttk.Frame(parent)
        update_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
        update_frame.columnconfigure(1, weight=1)
        
        # Checkbox
        var_name = f"redline_update_{row+1}_var"
        # Don't create BooleanVar here - it's already created in __init__
        ttk.Checkbutton(update_frame, text=title, 
                       variable=getattr(self, var_name)).grid(row=0, column=0, sticky=tk.W)
        
        # Engineer dropdown
        ttk.Label(update_frame, text="Engineer:").grid(row=1, column=0, sticky=tk.W)
        engineer_var_name = f"redline_update_{row+1}_engineer_var"
        setattr(self, engineer_var_name, tk.StringVar())
        combo = ttk.Combobox(update_frame, textvariable=getattr(self, engineer_var_name), 
                           state="readonly", width=15)
        combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        
        # Store the combo reference for later use
        combo_name = f"redline_update_{row+1}_engineer_combo"
        setattr(self, combo_name, combo)
        
        # Date
        ttk.Label(update_frame, text="Date:").grid(row=2, column=0, sticky=tk.W)
        date_entry_name = f"redline_update_{row+1}_date_entry"
        date_entry = DateEntry(update_frame, width=15)
        date_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        setattr(self, date_entry_name, date_entry)
    
    def create_ops_review_section(self, parent, row):
        """Create OPS review section"""
        section_frame = ttk.LabelFrame(parent, text="3. To Production for OPS Review", padding="5")
        section_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        section_frame.columnconfigure(1, weight=1)
        
        # OPS Review checkbox
        self.ops_review_var = tk.BooleanVar()
        ttk.Checkbutton(section_frame, text="OPS Review Updates", 
                       variable=self.ops_review_var).grid(row=0, column=0, columnspan=2, sticky=tk.W)
        
        # Date
        ttk.Label(section_frame, text="Updated:").grid(row=1, column=0, sticky=tk.W)
        self.ops_review_date_entry = DateEntry(section_frame, width=20)
        self.ops_review_date_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
    
    def create_d365_bom_section(self, parent, row):
        """Create D365 BOM Entry section"""
        section_frame = ttk.LabelFrame(parent, text="4. D365 BOM Entry", padding="5")
        section_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        section_frame.columnconfigure(1, weight=1)
        
        # D365 BOM Entry checkbox
        self.d365_bom_var = tk.BooleanVar()
        ttk.Checkbutton(section_frame, text="D365 BOM Entry", 
                       variable=self.d365_bom_var).grid(row=0, column=0, columnspan=2, sticky=tk.W)
        
        # Date
        ttk.Label(section_frame, text="Date:").grid(row=1, column=0, sticky=tk.W)
        self.d365_bom_date_entry = DateEntry(section_frame, width=20)
        self.d365_bom_date_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
    
    def create_peter_weck_section(self, parent, row):
        """Create Peter Weck review section"""
        section_frame = ttk.LabelFrame(parent, text="5. PETER WECK Review", padding="5")
        section_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        section_frame.columnconfigure(1, weight=1)
        
        # Fixed Errors checkbox
        self.peter_weck_var = tk.BooleanVar()
        ttk.Checkbutton(section_frame, text="Fixed Errors", 
                       variable=self.peter_weck_var).grid(row=0, column=0, columnspan=2, sticky=tk.W)
        
        # Date
        ttk.Label(section_frame, text="Date:").grid(row=1, column=0, sticky=tk.W)
        self.peter_weck_date_entry = DateEntry(section_frame, width=20)
        self.peter_weck_date_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
    
    def create_release_to_dee_section(self, parent, row):
        """Create Release to Dee section"""
        section_frame = ttk.LabelFrame(parent, text="6. Release to Dee", padding="5")
        section_frame.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        section_frame.columnconfigure(1, weight=1)
        
        # Fixed Errors checkbox
        self.release_fixed_errors_var = tk.BooleanVar()
        ttk.Checkbutton(section_frame, text="Fixed Errors", 
                       variable=self.release_fixed_errors_var).grid(row=0, column=0, columnspan=2, sticky=tk.W)
        
        # Missing Prints
        ttk.Label(section_frame, text="Missing Prints:").grid(row=1, column=0, sticky=tk.W)
        self.missing_prints_date_entry = DateEntry(section_frame, width=20)
        self.missing_prints_date_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        
        # D365 Updates
        ttk.Label(section_frame, text="D365 Updates:").grid(row=2, column=0, sticky=tk.W)
        self.d365_updates_date_entry = DateEntry(section_frame, width=20)
        self.d365_updates_date_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        
        # Other
        ttk.Label(section_frame, text="Other:").grid(row=3, column=0, sticky=tk.W)
        self.other_notes_var = tk.StringVar()
        self.other_notes_entry = ttk.Entry(section_frame, textvariable=self.other_notes_var, width=20)
        self.other_notes_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        
        ttk.Label(section_frame, text="Date:").grid(row=4, column=0, sticky=tk.W)
        self.other_date_entry = DateEntry(section_frame, width=20)
        self.other_date_entry.grid(row=4, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        
        # Due Date with auto-updating display
        ttk.Label(section_frame, text="Due Date:").grid(row=5, column=0, sticky=tk.W)
        self.release_due_date_entry = DateEntry(section_frame, width=20)
        self.release_due_date_entry.grid(row=5, column=1, sticky=(tk.W, tk.E), padx=(5, 0))
        
        # Due date display (auto-updating)
        self.release_due_display_var = tk.StringVar()
        self.release_due_display_label = ttk.Label(section_frame, textvariable=self.release_due_display_var, 
                                                   foreground="blue", font=('Arial', 9, 'italic'))
        self.release_due_display_label.grid(row=6, column=0, columnspan=2, sticky=tk.W, pady=(2, 0))
        
        # Set up auto-save traces for all workflow fields
        self.setup_workflow_autosave()
    
    def create_drafting_redline_content(self, parent):
        """Create content for Drafting & Redline Updates section"""
        parent.columnconfigure(1, weight=1)
        row = 0
        
        # Initial Redline
        ttk.Label(parent, text="Initial Redline:", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(0, 6))
        row += 1
        
        self.initial_redline_var = tk.BooleanVar()
        ttk.Checkbutton(parent, text="Completed", variable=self.initial_redline_var).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=6)
        row += 1
        
        ttk.Label(parent, text="Engineer:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.initial_engineer_var = tk.StringVar()
        self.initial_engineer_combo = ttk.Combobox(parent, textvariable=self.initial_engineer_var, state="readonly", width=25)
        self.initial_engineer_combo.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        ttk.Label(parent, text="Date:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.initial_date_entry = DateEntry(parent, width=25)
        self.initial_date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        # Redline Updates
        ttk.Separator(parent, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        row += 1
        ttk.Label(parent, text="Redline Updates:", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(0, 6))
        row += 1
        
        for i in range(1, 5):
            checkbox_var = tk.BooleanVar()
            setattr(self, f"redline_update_{i}_var", checkbox_var)
            ttk.Checkbutton(parent, text=f"Update {i}", variable=checkbox_var).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=6)
            row += 1
            
            ttk.Label(parent, text="Engineer:").grid(row=row, column=0, sticky=tk.W, pady=6)
            engineer_var = tk.StringVar()
            setattr(self, f"redline_update_{i}_engineer_var", engineer_var)
            combo = ttk.Combobox(parent, textvariable=engineer_var, state="readonly", width=25)
            setattr(self, f"redline_update_{i}_engineer_combo", combo)
            combo.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
            row += 1
            
            ttk.Label(parent, text="Date:").grid(row=row, column=0, sticky=tk.W, pady=6)
            date_entry = DateEntry(parent, width=25)
            setattr(self, f"redline_update_{i}_date_entry", date_entry)
            date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
            row += 1
            
            if i < 4:
                ttk.Separator(parent, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=6)
                row += 1
    
    def create_production_ops_content(self, parent):
        """Create content for Production & OPS Review section"""
        parent.columnconfigure(1, weight=1)
        row = 0
        
        # OPS Review
        ttk.Label(parent, text="OPS Review:", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(0, 6))
        row += 1
        
        self.ops_review_var = tk.BooleanVar()
        ttk.Checkbutton(parent, text="OPS Review Updates", variable=self.ops_review_var).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=6)
        row += 1
        
        ttk.Label(parent, text="Updated:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.ops_review_date_entry = DateEntry(parent, width=25)
        self.ops_review_date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        # Peter Weck Review
        ttk.Separator(parent, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        row += 1
        ttk.Label(parent, text="Peter Weck Review:", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(0, 6))
        row += 1
        
        self.peter_weck_var = tk.BooleanVar()
        ttk.Checkbutton(parent, text="Fixed Errors", variable=self.peter_weck_var).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=6)
        row += 1
        
        ttk.Label(parent, text="Date:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.peter_weck_date_entry = DateEntry(parent, width=25)
        self.peter_weck_date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
    
    def create_d365_release_content(self, parent):
        """Create content for D365 & Release section"""
        parent.columnconfigure(1, weight=1)
        row = 0
        
        # D365 BOM Entry
        ttk.Label(parent, text="D365 BOM Entry:", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(0, 6))
        row += 1
        
        self.d365_bom_var = tk.BooleanVar()
        ttk.Checkbutton(parent, text="D365 BOM Entry", variable=self.d365_bom_var).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=6)
        row += 1
        
        ttk.Label(parent, text="Date:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.d365_bom_date_entry = DateEntry(parent, width=25)
        self.d365_bom_date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        # Release to Dee
        ttk.Separator(parent, orient='horizontal').grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        row += 1
        ttk.Label(parent, text="Release to Dee:", font=('Arial', 10, 'bold')).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(0, 6))
        row += 1
        
        self.release_fixed_errors_var = tk.BooleanVar()
        ttk.Checkbutton(parent, text="Fixed Errors", variable=self.release_fixed_errors_var).grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=6)
        row += 1
        
        ttk.Label(parent, text="Missing Prints:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.missing_prints_date_entry = DateEntry(parent, width=25)
        self.missing_prints_date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        ttk.Label(parent, text="D365 Updates:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.d365_updates_date_entry = DateEntry(parent, width=25)
        self.d365_updates_date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        ttk.Label(parent, text="Other:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.other_notes_var = tk.StringVar()
        self.other_notes_entry = ttk.Entry(parent, textvariable=self.other_notes_var, width=25)
        self.other_notes_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        ttk.Label(parent, text="Date:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.other_date_entry = DateEntry(parent, width=25)
        self.other_date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        ttk.Label(parent, text="Due Date:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.release_due_date_entry = DateEntry(parent, width=25)
        self.release_due_date_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=6)
        row += 1
        
        # Due date display
        self.release_due_display_var = tk.StringVar()
        self.release_due_display_label = ttk.Label(parent, textvariable=self.release_due_display_var, 
                                                   foreground="blue", font=('Arial', 9, 'italic'))
        self.release_due_display_label.grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(2, 0))
        
        # Set up auto-save
        self.setup_workflow_autosave()
    
    def setup_workflow_autosave(self):
        """Set up auto-save traces for all workflow fields"""
        # Auto-save for initial redline
        if hasattr(self, 'initial_redline_var'):
            self.initial_redline_var.trace('w', self.auto_save)
        if hasattr(self, 'initial_engineer_var'):
            self.initial_engineer_var.trace('w', self.auto_save)
        if hasattr(self, 'initial_date_entry'):
            self.initial_date_entry.var.trace('w', self.auto_save)
        
        # Auto-save for redline updates
        for i in range(1, 5):
            var_name = f"redline_update_{i}_var"
            engineer_var_name = f"redline_update_{i}_engineer_var"
            date_entry_name = f"redline_update_{i}_date_entry"
            
            if hasattr(self, var_name):
                getattr(self, var_name).trace('w', self.auto_save)
            if hasattr(self, engineer_var_name):
                getattr(self, engineer_var_name).trace('w', self.auto_save)
            if hasattr(self, date_entry_name):
                getattr(self, date_entry_name).var.trace('w', self.auto_save)
        
        # Auto-save for OPS review
        if hasattr(self, 'ops_review_var'):
            self.ops_review_var.trace('w', self.auto_save)
        if hasattr(self, 'ops_review_date_entry'):
            self.ops_review_date_entry.var.trace('w', self.auto_save)
        
        # Auto-save for D365 BOM Entry
        if hasattr(self, 'd365_bom_var'):
            self.d365_bom_var.trace('w', self.auto_save)
        if hasattr(self, 'd365_bom_date_entry'):
            self.d365_bom_date_entry.var.trace('w', self.auto_save)
        
        # Auto-save for Peter Weck review
        if hasattr(self, 'peter_weck_var'):
            self.peter_weck_var.trace('w', self.auto_save)
        if hasattr(self, 'peter_weck_date_entry'):
            self.peter_weck_date_entry.var.trace('w', self.auto_save)
        
        # Auto-save for release to Dee
        if hasattr(self, 'release_fixed_errors_var'):
            self.release_fixed_errors_var.trace('w', self.auto_save)
        if hasattr(self, 'missing_prints_date_entry'):
            self.missing_prints_date_entry.var.trace('w', self.auto_save)
        if hasattr(self, 'd365_updates_date_entry'):
            self.d365_updates_date_entry.var.trace('w', self.auto_save)
        if hasattr(self, 'other_notes_var'):
            self.other_notes_var.trace('w', self.auto_save)
        if hasattr(self, 'other_date_entry'):
            self.other_date_entry.var.trace('w', self.auto_save)
        if hasattr(self, 'release_due_date_entry'):
            self.release_due_date_entry.var.trace('w', self.auto_save)
            self.release_due_date_entry.var.trace('w', self.update_release_due_display)
    
    def create_quick_access_panel(self):
        """Create the quick access panel for files and folders with scrolling"""
        main_container = ttk.LabelFrame(self.quick_access_container, text="Quick Access", padding="5")
        main_container.pack(fill=tk.BOTH, expand=True)
        main_container.rowconfigure(0, weight=1)
        main_container.columnconfigure(0, weight=1)
        
        # Create canvas and scrollbar for scrolling
        canvas = tk.Canvas(main_container, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        
        # Create frame inside canvas for content
        self.access_frame = ttk.Frame(canvas)
        self.access_frame.columnconfigure(0, weight=1)
        
        # Store canvas reference for later updates
        self.quick_access_canvas = canvas
        
        # Configure canvas scrolling with better region calculation
        def update_scroll_region(event=None):
            canvas.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        self.access_frame.bind("<Configure>", update_scroll_region)
        
        canvas_window = canvas.create_window((0, 0), window=self.access_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Grid canvas and scrollbar (always show scrollbar)
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Adjust canvas window width when canvas resizes
        def _configure_canvas(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind('<Configure>', _configure_canvas)
        
        # Enable mouse wheel scrolling
        self._bind_mousewheel(canvas, self.access_frame)
        
        # Initialize empty quick access area
        self.quick_access_buttons = []
        self.update_quick_access()
        
        # Start periodic refresh to check for Project File Monitor changes
        self.refresh_quick_access_periodically()
    
    def refresh_quick_access_periodically(self):
        """Refresh Quick Access panel every 10 seconds to check for Project File Monitor changes"""
        try:
            # Only refresh if we have a job number
            if hasattr(self, 'job_number_var') and self.job_number_var.get():
                self.update_quick_access()
        except Exception as e:
            print(f"Error refreshing quick access: {e}")
        
        # Schedule next refresh in 10 seconds
        self.root.after(10000, self.refresh_quick_access_periodically)
    
    def update_quick_access(self):
        """Update the quick access panel based on current project data"""
        # Clear existing buttons
        for button in self.quick_access_buttons:
            button.destroy()
        self.quick_access_buttons.clear()
        
        row = 0
        # Track paths and new/changed flags for this project
        changed_paths = set()
        
        def get_file_monitor_status(job_number):
            """Check Project File Monitor for file changes"""
            try:
                conn = sqlite3.connect(self.db_manager.db_path)
                cursor = conn.cursor()
                
                # Check for unacknowledged changes in file_changes table
                cursor.execute('''
                    SELECT file_path, change_type, COUNT(*) as count
                    FROM file_changes 
                    WHERE job_number = ? AND acknowledged = 0
                    GROUP BY file_path, change_type
                ''', (job_number,))
                
                changes = cursor.fetchall()
                conn.close()
                
                # Debug output
                print(f"Project Management - Checking file monitor status for job {job_number}:")
                print(f"  Found {len(changes)} unacknowledged changes")
                for file_path, change_type, count in changes:
                    print(f"    {change_type}: {file_path} ({count} records)")
                
                # Return status summary
                status = {
                    'has_changes': len(changes) > 0,
                    'new_files': sum(1 for _, change_type, _ in changes if change_type == 'new'),
                    'updated_files': sum(1 for _, change_type, _ in changes if change_type == 'updated'),
                    'deleted_files': sum(1 for _, change_type, _ in changes if change_type == 'deleted'),
                    'total_changes': len(changes)
                }
                
                return status
            except Exception as e:
                print(f"Error checking file monitor status: {e}")
                return {'has_changes': False, 'new_files': 0, 'updated_files': 0, 'deleted_files': 0, 'total_changes': 0}
        
        def track_path(path):
            if not path:
                return False
            try:
                if os.path.exists(path):
                    mtime = os.path.getmtime(path)
                else:
                    return False
                conn = sqlite3.connect(self.db_manager.db_path)
                cur = conn.cursor()
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS file_timestamps (
                        job_number TEXT NOT NULL,
                        path TEXT NOT NULL,
                        last_mtime REAL NOT NULL,
                        acknowledged INTEGER NOT NULL DEFAULT 1,
                        PRIMARY KEY(job_number, path)
                    )
                """)
                job_num = str(self.job_number_var.get()).strip()
                cur.execute("SELECT last_mtime, acknowledged FROM file_timestamps WHERE job_number=? AND path=?", (job_num, path))
                rowx = cur.fetchone()
                is_changed = False
                if rowx is None:
                    cur.execute("INSERT OR REPLACE INTO file_timestamps(job_number, path, last_mtime, acknowledged) VALUES(?,?,?,0)", (job_num, path, mtime))
                    is_changed = True
                else:
                    prev_mtime, acknowledged = rowx
                    if abs(mtime - prev_mtime) > 1e-6:
                        cur.execute("UPDATE file_timestamps SET last_mtime=?, acknowledged=0 WHERE job_number=? AND path=?", (mtime, job_num, path))
                        is_changed = True
                    elif acknowledged == 0:
                        is_changed = True
                conn.commit(); conn.close()
                if is_changed:
                    changed_paths.add(path)
                return is_changed
            except Exception:
                return False
        
        def style_button(btn, path, job_number=None):
            # Check Project File Monitor status first
            file_monitor_status = get_file_monitor_status(job_number) if job_number else {'has_changes': False}
            
            # Debug output for button styling
            button_text = btn.cget('text')
            print(f"Styling button '{button_text}' for job {job_number}:")
            print(f"  File monitor status: {file_monitor_status}")
            print(f"  Path in changed_paths: {path in changed_paths if path else 'N/A'}")
            
            # Determine button style based on Project File Monitor status
            if file_monitor_status['has_changes']:
                if file_monitor_status['deleted_files'] > 0:
                    # Red for deletions
                    print(f"  -> Applying RED style (deletions)")
                    try:
                        btn.configure(style='Deleted.TButton')
                    except Exception:
                        s = ttk.Style()
                        s.configure('Deleted.TButton', background='#F44336', foreground='white')
                        btn.configure(style='Deleted.TButton')
                elif file_monitor_status['new_files'] > 0 or file_monitor_status['updated_files'] > 0:
                    # Green for new/updated files
                    print(f"  -> Applying GREEN style (new/updated)")
                    try:
                        btn.configure(style='NewChanged.TButton')
                    except Exception:
                        s = ttk.Style()
                        s.configure('NewChanged.TButton', background='#4CAF50', foreground='white')
                        btn.configure(style='NewChanged.TButton')
            elif path and path in changed_paths:
                # Fallback to original change detection
                print(f"  -> Applying ORANGE style (fallback)")
                try:
                    btn.configure(style='Changed.TButton')
                except Exception:
                    s = ttk.Style()
                    s.configure('Changed.TButton', background='#FFB74D')
                    btn.configure(style='Changed.TButton')
            else:
                print(f"  -> No styling applied (normal)")
        
        # Job Directory button - use job number as button text
        job_dir = self.job_directory_picker.get()
        job_number = self.job_number_var.get()
        if job_dir and job_number:
            icon = "📁" if os.path.isdir(job_dir) else "📄"
            button_text = f"{icon} {job_number}"
            button = ttk.Button(self.access_frame, text=button_text, 
                              command=self.open_job_directory)
            button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
            style_button(button, job_dir, job_number)
            self.quick_access_buttons.append(button)
            row += 1
        
        # Customer Name button - use directory picker value if available
        customer_name = self.customer_name_var.get()
        customer_name_dir = self.customer_name_picker.get()
        
        if customer_name_dir:  # Use directory picker value first
            if os.path.exists(customer_name_dir):
                icon = "📁" if os.path.isdir(customer_name_dir) else "📄"
                # Just show the customer name from the text field, not the folder basename
                button_text = f"{icon} {customer_name}"
            else:
                icon = "📁"
                button_text = f"{icon} {customer_name}"
        elif customer_name:  # Fall back to text field value
            if os.path.exists(customer_name):
                icon = "📁" if os.path.isdir(customer_name) else "📄"
                button_text = f"{icon} {customer_name}"
            else:
                icon = "📁"
                button_text = f"{icon} {customer_name}"
        else:
            customer_name_dir = None
            button_text = None
        
        if button_text:
            path0 = customer_name_dir or customer_name
            changed = track_path(path0)
            button = ttk.Button(self.access_frame, text=button_text, 
                              command=lambda p=path0: self.open_customer_name_path(p))
            button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
            style_button(button, path0 if changed else None, job_number)
            self.quick_access_buttons.append(button)
            row += 1
        
        # Customer Location button - use directory picker value if available
        customer_location = self.customer_location_var.get()
        customer_location_dir = self.customer_location_picker.get()
        
        if customer_location_dir:  # Use directory picker value first
            if os.path.exists(customer_location_dir):
                icon = "📁" if os.path.isdir(customer_location_dir) else "📄"
                # Just show the customer location from the text field, not the folder basename
                button_text = f"{icon} {customer_location}"
            else:
                icon = "📁"
                button_text = f"{icon} {customer_location}"
        elif customer_location:  # Fall back to text field value
            if os.path.exists(customer_location):
                icon = "📁" if os.path.isdir(customer_location) else "📄"
                button_text = f"{icon} {customer_location}"
            else:
                icon = "📁"
                button_text = f"{icon} {customer_location}"
        else:
            customer_location_dir = None
            button_text = None
        
        if button_text:
            path1 = customer_location_dir or customer_location
            changed = track_path(path1)
            button = ttk.Button(self.access_frame, text=button_text, 
                              command=lambda p=path1: self.open_customer_location_path(p))
            button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
            style_button(button, path1 if changed else None, job_number)
            self.quick_access_buttons.append(button)
            row += 1
        
        # KOM AND OC FORM section - always show if job directory is loaded
        if hasattr(self, 'job_directory_picker') and self.job_directory_picker.get():
            if hasattr(self, 'kom_oc_form_path') and self.kom_oc_form_path and os.path.exists(self.kom_oc_form_path):
                button_text = f"📊 KOM AND OC FORM"
                button = ttk.Button(self.access_frame, text=button_text, 
                                  command=self.open_kom_oc_form)
                button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                style_button(button, self.kom_oc_form_path, job_number)
                self.quick_access_buttons.append(button)
                row += 1
            else:
                # No KOM file found - show placeholder
                kom_placeholder = ttk.Label(self.access_frame, text="KOM AND OC FORM: NOT FOUND", 
                                         font=('Arial', 9), foreground="gray")
                kom_placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                self.quick_access_buttons.append(kom_placeholder)
                row += 1
        
        # Sales documents section - always show if job directory is loaded
        if hasattr(self, 'job_directory_picker') and self.job_directory_picker.get():
            # Add SALES divider
            sales_label = ttk.Label(self.access_frame, text="SALES", font=('Arial', 10, 'bold'), foreground="blue")
            sales_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(10, 5))
            self.quick_access_buttons.append(sales_label)
            row += 1
            
            # Check if there are any sales files
            has_sales_files = (hasattr(self, 'proposal_docs') and self.proposal_docs) or (hasattr(self, 'other_docs') and self.other_docs)
            
            if has_sales_files:
                # Proposal documents buttons - automatically added when job directory is loaded
                if hasattr(self, 'proposal_docs') and self.proposal_docs:
                    for doc_path in self.proposal_docs:
                        # Use actual filename with truncation
                        filename = os.path.basename(doc_path)
                        # Remove file extension and truncate if too long
                        name_without_ext = os.path.splitext(filename)[0]
                        # Increase max length to 35 to show more of the filename
                        if len(name_without_ext) > 35:
                            display_name = name_without_ext[:32] + "..."
                        else:
                            display_name = name_without_ext
                        
                        button_text = f"📄 {display_name}"
                        button = ttk.Button(self.access_frame, text=button_text, 
                                          command=lambda path=doc_path: self.open_proposal_doc(path))
                        button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        style_button(button, doc_path, job_number)
                        self.quick_access_buttons.append(button)
                        row += 1
                
                # Other important documents buttons - automatically added when job directory is loaded
                if hasattr(self, 'other_docs') and self.other_docs:
                    for icon, filename, file_path in self.other_docs:
                        # Create shorter, consistent button labels
                        button_text = self.create_short_button_text(icon, filename)
                        button = ttk.Button(self.access_frame, text=button_text, 
                                          command=lambda path=file_path: self.open_other_doc(path))
                        button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        style_button(button, file_path, job_number)
                        self.quick_access_buttons.append(button)
                        row += 1
            else:
                # No sales files found - show placeholder
                placeholder_label = ttk.Label(self.access_frame, text="SALES: NOT PROCESSED", 
                                           font=('Arial', 9), foreground="gray", style="Placeholder.TLabel")
                placeholder_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                self.quick_access_buttons.append(placeholder_label)
                row += 1
        
        # Engineering documents section - always show if job directory is loaded
        if hasattr(self, 'job_directory_picker') and self.job_directory_picker.get():
            # Add ENGINEERING divider
            engineering_label = ttk.Label(self.access_frame, text="ENGINEERING", font=('Arial', 10, 'bold'), foreground="green")
            engineering_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(10, 5))
            self.quick_access_buttons.append(engineering_label)
            row += 1
            
            # Check if there are any engineering files
            has_engineering_files = (hasattr(self, 'engineering_general_docs') and self.engineering_general_docs) or (hasattr(self, 'engineering_releases_docs') and self.engineering_releases_docs)
            
            if has_engineering_files:
                # General Design subsection
                if hasattr(self, 'engineering_general_docs') and self.engineering_general_docs:
                    general_label = ttk.Label(self.access_frame, text="General Design", font=('Arial', 9, 'bold'), foreground="darkgreen")
                    general_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(5, 2))
                    self.quick_access_buttons.append(general_label)
                    row += 1
                    
                    for file_path in self.engineering_general_docs:
                        filename = os.path.basename(file_path)
                        button_text = self.create_short_button_text("📊", filename)
                        button = ttk.Button(self.access_frame, text=button_text, 
                                          command=lambda path=file_path: self.open_engineering_doc(path))
                        button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        style_button(button, file_path, job_number)
                        self.quick_access_buttons.append(button)
                        row += 1
                else:
                    # No General Design files - show placeholder
                    general_placeholder = ttk.Label(self.access_frame, text="General Design: NOT PROCESSED", 
                                                 font=('Arial', 8), foreground="gray")
                    general_placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                    self.quick_access_buttons.append(general_placeholder)
                    row += 1
                
                # Releases subsection
                if hasattr(self, 'engineering_releases_docs') and self.engineering_releases_docs:
                    releases_label = ttk.Label(self.access_frame, text="Releases", font=('Arial', 9, 'bold'), foreground="darkgreen")
                    releases_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(5, 2))
                    self.quick_access_buttons.append(releases_label)
                    row += 1
                    
                    for file_path in self.engineering_releases_docs:
                        filename = os.path.basename(file_path)
                        button_text = self.create_short_button_text("📄", filename)
                        button = ttk.Button(self.access_frame, text=button_text, 
                                          command=lambda path=file_path: self.open_engineering_doc(path))
                        button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        style_button(button, file_path, job_number)
                        self.quick_access_buttons.append(button)
                        row += 1
                else:
                    # No Releases files - show placeholder
                    releases_placeholder = ttk.Label(self.access_frame, text="Releases: NOT PROCESSED", 
                                                   font=('Arial', 8), foreground="gray")
                    releases_placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                    self.quick_access_buttons.append(releases_placeholder)
                    row += 1
            else:
                # No engineering files found at all - show main placeholder
                engineering_placeholder = ttk.Label(self.access_frame, text="ENGINEERING: NOT PROCESSED", 
                                                 font=('Arial', 9), foreground="gray")
                engineering_placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                self.quick_access_buttons.append(engineering_placeholder)
                row += 1
        
        # Drafting documents section - always show if job directory is loaded
        if hasattr(self, 'job_directory_picker') and self.job_directory_picker.get():
            # Add DRAFTING divider
            drafting_label = ttk.Label(self.access_frame, text="DRAFTING", font=('Arial', 10, 'bold'), foreground="purple")
            drafting_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(10, 5))
            self.quick_access_buttons.append(drafting_label)
            row += 1
            
            # Check for Systems folder
            job_dir = self.job_directory_picker.get()
            systems_dir = os.path.join(job_dir, "4. Drafting", "Systems")
            
            if os.path.exists(systems_dir) and os.path.isdir(systems_dir):
                # Systems subsection
                systems_label = ttk.Label(self.access_frame, text="Systems", font=('Arial', 9, 'bold'), foreground="darkviolet")
                systems_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(5, 2))
                self.quick_access_buttons.append(systems_label)
                row += 1
                
                # Scan for .dwg files
                dwg_files = []
                try:
                    for file in os.listdir(systems_dir):
                        if file.lower().endswith('.dwg'):
                            dwg_files.append(os.path.join(systems_dir, file))
                except Exception as e:
                    print(f"Error scanning drafting systems directory: {e}")
                
                if dwg_files:
                    # Sort by name
                    dwg_files.sort(key=lambda x: os.path.basename(x).lower())
                    
                    for file_path in dwg_files:
                        filename = os.path.basename(file_path)
                        button_text = self.create_short_button_text("📐", filename)
                        button = ttk.Button(self.access_frame, text=button_text, 
                                          command=lambda path=file_path: self.open_drafting_doc(path))
                        button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        style_button(button, file_path, job_number)
                        self.quick_access_buttons.append(button)
                        row += 1
                else:
                    # No .dwg files found
                    placeholder = ttk.Label(self.access_frame, text="Systems: No DWG files found", 
                                         font=('Arial', 8), foreground="gray")
                    placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                    self.quick_access_buttons.append(placeholder)
                    row += 1
                
                # Package subsection
                package_dir = os.path.join(job_dir, "4. Drafting", "Package")
                if os.path.exists(package_dir) and os.path.isdir(package_dir):
                    package_label = ttk.Label(self.access_frame, text="Package", font=('Arial', 9, 'bold'), foreground="darkviolet")
                    package_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(5, 2))
                    self.quick_access_buttons.append(package_label)
                    row += 1
                    
                    # Scan for .dwf and .dwg files
                    package_files = []
                    try:
                        for file in os.listdir(package_dir):
                            if file.lower().endswith(('.dwf', '.dwg')):
                                package_files.append(os.path.join(package_dir, file))
                    except Exception as e:
                        print(f"Error scanning drafting package directory: {e}")
                    
                    if package_files:
                        # Sort by name
                        package_files.sort(key=lambda x: os.path.basename(x).lower())
                        
                        for file_path in package_files:
                            filename = os.path.basename(file_path)
                            # Use different icon for .dwf vs .dwg
                            icon = "📦" if filename.lower().endswith('.dwf') else "📐"
                            button_text = self.create_short_button_text(icon, filename)
                            button = ttk.Button(self.access_frame, text=button_text, 
                                              command=lambda path=file_path: self.open_drafting_doc(path))
                            button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                            self.quick_access_buttons.append(button)
                            row += 1
                    else:
                        # No package files found
                        placeholder = ttk.Label(self.access_frame, text="Package: No files found", 
                                             font=('Arial', 8), foreground="gray")
                        placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        self.quick_access_buttons.append(placeholder)
                        row += 1
                
                # Fabs subsection
                fabs_dir = os.path.join(job_dir, "4. Drafting", "Fabs")
                if os.path.exists(fabs_dir) and os.path.isdir(fabs_dir):
                    fabs_label = ttk.Label(self.access_frame, text="Fabs", font=('Arial', 9, 'bold'), foreground="darkviolet")
                    fabs_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(5, 2))
                    self.quick_access_buttons.append(fabs_label)
                    row += 1
                    
                    # Get customer and location info from job_dir path
                    # Parse: F:\Customer\Location\JobNum\4. Drafting\Fabs
                    # to get: C:\$WorkingFolder\Jobs F\Customer\Location\JobNum\4. Drafting\Fabs
                    customer_name = self.customer_name_var.get()
                    customer_location = self.customer_location_var.get()
                    job_number = self.job_number_var.get()
                    
                    # Build working folder path
                    working_fabs_dir = None
                    if customer_name and customer_location and job_number:
                        working_fabs_dir = os.path.join(r"C:\$WorkingFolder\Jobs F", 
                                                       customer_name, customer_location, 
                                                       job_number, "4. Drafting", "Fabs")
                    
                    # Scan for files in specific order: .dwf (for .idw lookup), then .dwg, then excel files
                    idw_files = []  # Changed from dwf_files
                    dwg_files = []
                    excel_files = []
                    
                    try:
                        for file in os.listdir(fabs_dir):
                            file_lower = file.lower()
                            file_path = os.path.join(fabs_dir, file)
                            
                            if file_lower.endswith('.dwf'):
                                # For .dwf files, look for corresponding .idw in working folder
                                base_name = os.path.splitext(file)[0]
                                idw_name = base_name + '.idw'
                                
                                if working_fabs_dir and os.path.exists(working_fabs_dir):
                                    idw_path = os.path.join(working_fabs_dir, idw_name)
                                    if os.path.exists(idw_path):
                                        idw_files.append((file, idw_path))  # Store display name and actual path
                                    else:
                                        # .idw not found, still add but will open .dwf
                                        idw_files.append((file, file_path))
                                else:
                                    # Working folder not available, use .dwf
                                    idw_files.append((file, file_path))
                            elif file_lower.endswith('.dwg'):
                                dwg_files.append(file_path)
                            elif file_lower.endswith(('.xls', '.xlsx', '.xlsm')):
                                excel_files.append(file_path)
                    except Exception as e:
                        print(f"Error scanning drafting fabs directory: {e}")
                    
                    # Sort each category
                    idw_files.sort(key=lambda x: x[0].lower())  # Sort by display name
                    dwg_files.sort(key=lambda x: os.path.basename(x).lower())
                    excel_files.sort(key=lambda x: os.path.basename(x).lower())
                    
                    # Display .idw files first
                    if idw_files:
                        for display_name, actual_path in idw_files:
                            # Show .idw in the button text if it's actually an .idw file
                            if actual_path.lower().endswith('.idw'):
                                button_filename = os.path.splitext(display_name)[0] + '.idw'
                                icon = "🔧"  # Inventor icon
                            else:
                                button_filename = display_name
                                icon = "📦"  # .dwf fallback
                            
                            button_text = self.create_short_button_text(icon, button_filename)
                            button = ttk.Button(self.access_frame, text=button_text, 
                                              command=lambda path=actual_path: self.open_drafting_doc(path))
                            button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                            self.quick_access_buttons.append(button)
                            row += 1
                    
                    # Then .dwg files
                    if dwg_files:
                        for file_path in dwg_files:
                            filename = os.path.basename(file_path)
                            button_text = self.create_short_button_text("📐", filename)
                            button = ttk.Button(self.access_frame, text=button_text, 
                                              command=lambda path=file_path: self.open_drafting_doc(path))
                            button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                            self.quick_access_buttons.append(button)
                            row += 1
                    
                    # Check for D365 Import file
                    has_d365_import = False
                    if excel_files:
                        for file_path in excel_files:
                            filename = os.path.basename(file_path).upper()
                            if "D365 IMPORT" in filename:
                                has_d365_import = True
                                break
                    
                    # Show "NEW D365 Import" button if file doesn't exist
                    if not has_d365_import:
                        new_d365_btn = tk.Button(self.access_frame, text="📊 NEW D365 Import", 
                                                bg='#28a745', fg='white',
                                                font=('Arial', 9, 'bold'),
                                                relief='raised', bd=2, cursor='hand2',
                                                activebackground='#218838', activeforeground='white',
                                                command=lambda: self.create_d365_import(fabs_dir))
                        new_d365_btn.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        self.quick_access_buttons.append(new_d365_btn)
                        row += 1
                    
                    # Check for Transmittal Notice DWG file
                    has_transmittal = False
                    if dwg_files:
                        for file_path in dwg_files:
                            filename = os.path.basename(file_path).upper()
                            if "TRANSMITTAL NOTICE" in filename or "TRANMITTAL NOTICE" in filename:
                                has_transmittal = True
                                break
                    
                    # Show "NEW Transmittal Notice" button if file doesn't exist
                    if not has_transmittal:
                        new_transmittal_btn = tk.Button(self.access_frame, text="📐 NEW Transmittal Notice", 
                                                bg='#28a745', fg='white',
                                                font=('Arial', 9, 'bold'),
                                                relief='raised', bd=2, cursor='hand2',
                                                activebackground='#218838', activeforeground='white',
                                                command=lambda: self.create_transmittal_notice(fabs_dir))
                        new_transmittal_btn.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        self.quick_access_buttons.append(new_transmittal_btn)
                        row += 1
                    
                    # Display Excel files
                    if excel_files:
                        for file_path in excel_files:
                            filename = os.path.basename(file_path)
                            button_text = self.create_short_button_text("📊", filename)
                            button = ttk.Button(self.access_frame, text=button_text, 
                                              command=lambda path=file_path: self.open_drafting_doc(path))
                            button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                            self.quick_access_buttons.append(button)
                            row += 1
                    
                    if not idw_files and not dwg_files and not excel_files:
                        # No fabs files found
                        placeholder = ttk.Label(self.access_frame, text="Fabs: No files found", 
                                             font=('Arial', 8), foreground="gray")
                        placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        self.quick_access_buttons.append(placeholder)
                        row += 1
                
                # Burn Table Files subsection
                burn_dir = os.path.join(job_dir, "4. Drafting", "Burn Table Files")
                if os.path.exists(burn_dir) and os.path.isdir(burn_dir):
                    burn_label = ttk.Label(self.access_frame, text="Burn Table Files", font=('Arial', 9, 'bold'), foreground="darkviolet")
                    burn_label.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(5, 2))
                    self.quick_access_buttons.append(burn_label)
                    row += 1
                    
                    # Scan for .dwg files only
                    burn_files = []
                    try:
                        for file in os.listdir(burn_dir):
                            if file.lower().endswith('.dwg'):
                                burn_files.append(os.path.join(burn_dir, file))
                    except Exception as e:
                        print(f"Error scanning burn table files directory: {e}")
                    
                    if burn_files:
                        # Sort by name
                        burn_files.sort(key=lambda x: os.path.basename(x).lower())
                        
                        for file_path in burn_files:
                            filename = os.path.basename(file_path)
                            button_text = self.create_short_button_text("🔥", filename)
                            button = ttk.Button(self.access_frame, text=button_text, 
                                              command=lambda path=file_path: self.open_drafting_doc(path))
                            button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                            self.quick_access_buttons.append(button)
                            row += 1
                    else:
                        # No burn table files found
                        placeholder = ttk.Label(self.access_frame, text="Burn Table Files: No DWG files found", 
                                             font=('Arial', 8), foreground="gray")
                        placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                        self.quick_access_buttons.append(placeholder)
                        row += 1
            else:
                # Systems folder doesn't exist
                placeholder = ttk.Label(self.access_frame, text="DRAFTING: NOT PROCESSED", 
                                     font=('Arial', 9), foreground="gray")
                placeholder.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=2)
                self.quick_access_buttons.append(placeholder)
                row += 1
        
        # Print Package Review button - only show if job directory is loaded
        if hasattr(self, 'job_directory_picker') and self.job_directory_picker.get() and job_number:
            # Add separator
            separator = ttk.Separator(self.access_frame, orient='horizontal')
            separator.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=(10, 5))
            self.quick_access_buttons.append(separator)
            row += 1
            
            # Check if Print Package Review already exists
            pp_review_exists = self.check_print_package_review_exists(job_number)
            
            if pp_review_exists:
                # Show "Open Print Package Folder" button
                button_text = "📁 Open Print Package Folder"
                button = ttk.Button(self.access_frame, text=button_text, 
                                  command=self.open_print_package_folder)
                button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=5)
                self.quick_access_buttons.append(button)
            else:
                # Show "Initialize Print Package Review" button
                button_text = "🚀 Initialize Print Package Review"
                button = ttk.Button(self.access_frame, text=button_text, 
                                  command=self.initialize_print_package_review)
                button.grid(row=row, column=0, sticky=(tk.W, tk.E), pady=5)
                self.quick_access_buttons.append(button)
            
            row += 1
        
        # If no quick access items, show a message
        if not self.quick_access_buttons:
            label = ttk.Label(self.access_frame, text="No quick access items\navailable for this project", 
                            foreground="gray", justify="center")
            label.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=20)
            self.quick_access_buttons.append(label)
        
        # Update scroll region after all buttons are added
        if hasattr(self, 'quick_access_canvas'):
            self.access_frame.update_idletasks()
            self.quick_access_canvas.configure(scrollregion=self.quick_access_canvas.bbox("all"))
    
    def initialize_print_package_review(self):
        """Initialize Print Package Review workflow for the current project"""
        job_number = self.job_number_var.get()
        job_directory = self.job_directory_picker.get()
        
        if not job_number:
            messagebox.showwarning("Warning", "Please select a project first")
            return
        
        if not job_directory:
            messagebox.showwarning("Warning", "Please set the job directory first")
            return
        
        # Check if Print Package Review already exists
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT COUNT(*) FROM print_package_reviews 
                WHERE job_number = ?
            ''', (job_number,))
            
            if cursor.fetchone()[0] > 0:
                conn.close()
                messagebox.showinfo("Info", f"Print Package Review already initialized for job {job_number}")
                return
            
            conn.close()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to check existing reviews: {str(e)}")
            return
        
        # Show file picker for print package files
        from tkinter import filedialog
        file_paths = filedialog.askopenfilenames(
            title="Select Print Package PDF Files",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        if not file_paths:
            messagebox.showinfo("Info", "No files selected. Print Package Review initialization cancelled.")
            return
        
        # Initialize the Print Package Review
        try:
            self.create_print_package_structure(job_number, job_directory, file_paths)
            messagebox.showinfo("Success", f"Print Package Review initialized for job {job_number}!\n\n{len(file_paths)} files added to Stage 0.\n\nButton will now switch to 'Open Print Package Folder'.")
            self.update_quick_access()  # Refresh to show new button
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to initialize Print Package Review: {str(e)}")
    
    def create_print_package_structure(self, job_number, job_directory, file_paths):
        """Create the Print Package Review folder structure and database records"""
        import uuid
        import shutil
        from datetime import datetime
        
        # Generate unique review ID
        review_id = str(uuid.uuid4())
        
        # Create folder structure
        pp_base_path = os.path.join(job_directory, "4. Drafting", "PP-Print Packages")
        
        # Define the 8 stages
        stages = [
            "0-Drafting-Print Package",
            "1-Engineer Review", 
            "2-Engineering QC Review",
            "3-Drafting Updates (ENG)",
            "4-Lead Designer Review",
            "5-Production OPS Review", 
            "6-Drafting Updates (OPS)",
            "7-FINAL Print Package (Approved)"
        ]
        
        # Create all stage directories
        for stage in stages:
            stage_path = os.path.join(pp_base_path, stage)
            os.makedirs(stage_path, exist_ok=True)
        
        # Copy files to Stage 0
        stage_0_path = os.path.join(pp_base_path, "0-Drafting-Print Package")
        copied_files = []
        
        for file_path in file_paths:
            file_name = os.path.basename(file_path)
            dest_path = os.path.join(stage_0_path, file_name)
            shutil.copy2(file_path, dest_path)
            copied_files.append((file_name, file_path, dest_path))
        
        # Save to database
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        # Create review record
        cursor.execute('''
            INSERT INTO print_package_reviews 
            (job_number, review_id, status, current_stage, initialized_by, initialized_date, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (job_number, review_id, 'initialized', 0, 'System', datetime.now().isoformat(), 
              f'Initialized with {len(file_paths)} files'))
        
        # Create file records
        for file_name, original_path, stage_0_path in copied_files:
            file_size = os.path.getsize(stage_0_path)
            
            cursor.execute('''
                INSERT INTO print_package_files 
                (review_id, job_number, file_name, original_path, stage_0_path, file_size, created_date)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (review_id, job_number, file_name, original_path, stage_0_path, file_size, 
                  datetime.now().isoformat()))
        
        # Create workflow records for all stages
        stage_names = [
            "Drafting Print Package",
            "Engineer Review",
            "Engineering QC Review", 
            "Drafting Updates (ENG)",
            "Lead Designer Review",
            "Production OPS Review",
            "Drafting Updates (OPS)",
            "FINAL Print Package (Approved)"
        ]
        
        for i, (stage, stage_name) in enumerate(zip(stages, stage_names)):
            cursor.execute('''
                INSERT INTO print_package_workflow 
                (review_id, job_number, stage, stage_name, status, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (review_id, job_number, i, stage_name, 'pending' if i > 0 else 'completed', 
                  f'Stage {i}: {stage_name}'))
        
        conn.commit()
        conn.close()
        
        print(f"Print Package Review structure created for job {job_number}")
        print(f"Review ID: {review_id}")
        print(f"Files copied to Stage 0: {len(copied_files)}")
    
    def check_print_package_review_exists(self, job_number):
        """Check if a Print Package Review already exists for the given job"""
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT COUNT(*) FROM print_package_reviews 
                WHERE job_number = ?
            ''', (job_number,))
            
            count = cursor.fetchone()[0]
            conn.close()
            
            return count > 0
            
        except Exception as e:
            print(f"Error checking Print Package Review existence: {e}")
            return False
    
    def open_print_package_folder(self):
        """Open the Print Package Review folder for the current project"""
        job_number = self.job_number_var.get()
        job_directory = self.job_directory_picker.get()
        
        if not job_number:
            messagebox.showwarning("Warning", "Please select a project first")
            return
        
        if not job_directory:
            messagebox.showwarning("Warning", "Please set the job directory first")
            return
        
        # Construct the Print Package Review folder path
        pp_folder_path = os.path.join(job_directory, "4. Drafting", "PP-Print Packages")
        
        if os.path.exists(pp_folder_path):
            try:
                os.startfile(pp_folder_path)
                print(f"Opened Print Package Review folder for job {job_number}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open Print Package Review folder: {str(e)}")
        else:
            messagebox.showerror("Error", f"Print Package Review folder not found:\n{pp_folder_path}")

    def create_action_buttons(self):
        """Create compact footer toolbar with uniform buttons"""
        # Separator line above footer
        parent = getattr(self, 'content', self.root)
        separator = ttk.Separator(parent, orient='horizontal')
        separator.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        # Footer toolbar frame - pinned to bottom
        footer_frame = tk.Frame(parent, bg='#f5f5f5', relief='flat', bd=0, height=45)
        footer_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E))
        footer_frame.grid_propagate(False)
        
        # Inner container for buttons
        button_container = tk.Frame(footer_frame, bg='#f5f5f5')
        button_container.pack(side=tk.LEFT, padx=10, pady=6)
        
        # Dashboard button with accent style
        dashboard_btn = tk.Button(button_container, text="🏠 Dashboard", 
                                  command=self.open_dashboard,
                                  width=12, height=1,
                                  bg='#2196F3', fg='white',
                                  font=('Arial', 9),
                                  relief='raised', bd=1, cursor='hand2',
                                  activebackground='#1976D2', activeforeground='white')
        dashboard_btn.pack(side=tk.LEFT, padx=(0, 10))
        self._add_button_hover_effect(dashboard_btn, '#2196F3', '#1976D2')
        
        # All other buttons with consistent styling
        buttons = [
            ("New Project", self.new_project),
            ("Save Project", self.save_project),
            ("Duplicate", self.duplicate_project),
            ("Delete Project", self.delete_project),
            ("Clean & Fix Data", self.clean_duplicates),
            ("Reset Database", self.reset_database),
            ("Refresh", self.load_projects),
            ("Refresh People", self.load_dropdown_data),
            ("Export JSON", self.export_data),
            ("Import JSON", self.import_data)
        ]
        
        for text, command in buttons:
            btn = tk.Button(button_container, text=text, command=command,
                          width=12, height=1,
                          bg='#ffffff', fg='#333333',
                          font=('Arial', 9),
                          relief='raised', bd=1, cursor='hand2',
                          activebackground='#E3F2FD', activeforeground='#1976D2')
            btn.pack(side=tk.LEFT, padx=(0, 5))
            self._add_button_hover_effect(btn, '#ffffff', '#E3F2FD')
    
    def _add_button_hover_effect(self, button, normal_bg, hover_bg):
        """Add subtle hover effect to a button"""
        def on_enter(e):
            button.config(bg=hover_bg)
        
        def on_leave(e):
            button.config(bg=normal_bg)
        
        button.bind('<Enter>', on_enter)
        button.bind('<Leave>', on_leave)
    
    def open_job_directory(self):
        """Open the job directory"""
        directory = self.job_directory_picker.get()
        if directory and os.path.exists(directory):
            self.open_path(directory)
        else:
            messagebox.showwarning("Warning", "No valid job directory selected!")
    
    def open_customer_name_path(self, path):
        """Open customer name path (from directory picker or text field)"""
        print(f"DEBUG: Opening customer name path: '{path}'")
        if path and os.path.exists(path):
            print(f"DEBUG: Opening direct path: {path}")
            self.open_path(path)
        elif path:
            print(f"DEBUG: Path doesn't exist, searching for folder: {path}")
            self.search_and_open_folder(path)
        else:
            print("DEBUG: No customer name path provided")
            messagebox.showwarning("Warning", "No customer name path provided!")
    
    def open_customer_location_path(self, path):
        """Open customer location path (from directory picker or text field)"""
        print(f"DEBUG: Opening customer location path: '{path}'")
        if path and os.path.exists(path):
            print(f"DEBUG: Opening direct path: {path}")
            self.open_path(path)
        elif path:
            print(f"DEBUG: Path doesn't exist, searching for folder: {path}")
            self.search_and_open_folder(path)
        else:
            print("DEBUG: No customer location path provided")
            messagebox.showwarning("Warning", "No customer location path provided!")
    
    def open_customer_name(self):
        """Open customer name path (legacy method)"""
        customer_name = self.customer_name_var.get()
        print(f"DEBUG: Customer name = '{customer_name}'")
        if customer_name:
            if os.path.exists(customer_name):
                print(f"DEBUG: Opening direct path: {customer_name}")
                self.open_path(customer_name)
            else:
                print(f"DEBUG: Searching for folder: {customer_name}")
                self.search_and_open_folder(customer_name)
        else:
            print("DEBUG: No customer name entered")
            messagebox.showwarning("Warning", "No customer name entered!")
    
    def open_customer_location(self):
        """Open customer location path (legacy method)"""
        customer_location = self.customer_location_var.get()
        print(f"DEBUG: Customer location = '{customer_location}'")
        if customer_location:
            if os.path.exists(customer_location):
                print(f"DEBUG: Opening direct path: {customer_location}")
                self.open_path(customer_location)
            else:
                print(f"DEBUG: Searching for folder: {customer_location}")
                self.search_and_open_folder(customer_location)
        else:
            print("DEBUG: No customer location entered")
            messagebox.showwarning("Warning", "No customer location entered!")
    
    def open_path(self, path):
        """Open a file or directory path"""
        try:
            # Mark acknowledged for this file so the orange state clears
            try:
                job_num = str(self.job_number_var.get()).strip()
                conn = sqlite3.connect(self.db_manager.db_path)
                cur = conn.cursor()
                cur.execute("UPDATE file_timestamps SET acknowledged=1 WHERE job_number=? AND path=?", (job_num, path))
                conn.commit(); conn.close()
            except Exception:
                pass
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", path])
            else:
                subprocess.run(["xdg-open", path])
            # Refresh to update button styles
            self.update_quick_access()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open path: {str(e)}")
    
    def search_and_open_folder(self, name):
        """Search for and open a folder by name"""
        print(f"DEBUG: Searching for folder: '{name}'")
        
        # Common search locations
        search_paths = [
            os.path.join(os.path.expanduser("~"), "Documents", "Projects", name),
            os.path.join(os.path.expanduser("~"), "Desktop", name),
            os.path.join("C:", "Projects", name),
            os.path.join("C:", "Users", os.getenv("USERNAME", ""), "Documents", name),
            os.path.join(os.path.expanduser("~"), "Documents", name),
            os.path.join(os.path.expanduser("~"), "OneDrive", "Documents", name),
            os.path.join("C:", "Users", os.getenv("USERNAME", ""), "OneDrive", "Documents", name),
            os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", name),
            os.path.join("C:", "Users", os.getenv("USERNAME", ""), "OneDrive", "Desktop", name)
        ]
        
        print(f"DEBUG: Checking {len(search_paths)} search paths...")
        for i, path in enumerate(search_paths):
            print(f"DEBUG: Checking path {i+1}: {path}")
            if os.path.exists(path):
                print(f"DEBUG: Found folder at: {path}")
                self.open_path(path)
                return
        
        # If not found, try to create a folder and open it
        print(f"DEBUG: Folder not found, creating new folder...")
        new_folder = os.path.join(os.path.expanduser("~"), "Documents", "Projects", name)
        try:
            os.makedirs(new_folder, exist_ok=True)
            print(f"DEBUG: Created new folder: {new_folder}")
            self.open_path(new_folder)
        except Exception as e:
            print(f"DEBUG: Failed to create folder: {e}")
            messagebox.showinfo("Info", f"Folder for '{name}' not found. Please create it manually or use the directory picker to select an existing folder.")
    
    def clean_duplicates(self):
        """Remove duplicate projects and clean job numbers"""
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        try:
            print("DEBUG: Starting database cleanup...")
            
            # First, get all projects
            cursor.execute("SELECT id, job_number FROM projects")
            projects = cursor.fetchall()
            print(f"DEBUG: Found {len(projects)} projects to check")
            
            cleaned_count = 0
            for project_id, job_number in projects:
                original = str(job_number)
                print(f"DEBUG: Processing project {project_id}: '{original}'")
                
                # Clean the job number, preserving suffixed copies like '12345 (1)'
                clean_job_number = str(job_number).strip()
                if re.match(r"^\d{5}( \(\d+\))?$", clean_job_number):
                    pass  # already normalized format
                else:
                    # Extract just the first 5-digit sequence as a fallback
                    m = re.search(r"\d{5}", clean_job_number)
                    if m:
                        clean_job_number = m.group(0)
                
                print(f"DEBUG: Cleaned '{original}' to '{clean_job_number}'")
                
                # Update if different
                if clean_job_number != original:
                    print(f"DEBUG: Updating project {project_id} from '{original}' to '{clean_job_number}'")
                    cursor.execute("UPDATE projects SET job_number = ? WHERE id = ?", 
                                 (clean_job_number, project_id))
                    cleaned_count += 1
                else:
                    print(f"DEBUG: No change needed for project {project_id}")
            
            print(f"DEBUG: Updated {cleaned_count} job numbers")
            
            # Find duplicates after cleaning
            cursor.execute("""
                SELECT job_number, COUNT(*) as count
                FROM projects 
                GROUP BY job_number 
                HAVING COUNT(*) > 1
            """)
            duplicates = cursor.fetchall()
            print(f"DEBUG: Found {len(duplicates)} duplicate groups")
            
            duplicate_count = 0
            if duplicates:
                # Remove duplicates, keeping the one with the highest ID (most recent)
                for job_number, count in duplicates:
                    print(f"DEBUG: Removing {count-1} duplicates for job {job_number}")
                    cursor.execute("""
                        DELETE FROM projects 
                        WHERE job_number = ? AND id NOT IN (
                            SELECT MAX(id) FROM projects WHERE job_number = ?
                        )
                    """, (job_number, job_number))
                    duplicate_count += count - 1
            
            conn.commit()
            print(f"DEBUG: Cleanup complete - {cleaned_count} cleaned, {duplicate_count} duplicates removed")
            messagebox.showinfo("Success", f"Cleaned {cleaned_count} job numbers and removed {duplicate_count} duplicate project(s)!")
            self.load_projects()
            
        except Exception as e:
            print(f"DEBUG: Error during cleanup: {e}")
            messagebox.showerror("Error", f"Failed to clean duplicates: {str(e)}")
        finally:
            conn.close()
    
    def reset_database(self):
        """Reset the database by recreating it"""
        if messagebox.askyesno("Confirm Reset", 
                              "Are you sure you want to reset the database? This will delete ALL data!"):
            try:
                # Delete the database file
                import os
                if os.path.exists(self.db_manager.db_path):
                    os.remove(self.db_manager.db_path)
                    print(f"DEBUG: Deleted database file: {self.db_manager.db_path}")
                
                # Recreate the database
                self.db_manager = DatabaseManager()
                print("DEBUG: Recreated database")
                
                # Reload projects (will be empty)
                self.load_projects()
                self.new_project()
                
                messagebox.showinfo("Success", "Database reset successfully!")
                
            except Exception as e:
                print(f"DEBUG: Error resetting database: {e}")
                messagebox.showerror("Error", f"Failed to reset database: {str(e)}")
    
    def auto_extract_and_save(self, *args):
        """Auto-extract customer info from job directory and save"""
        job_dir = self.job_directory_picker.get()
        if job_dir and os.path.exists(job_dir):
            self.extract_customer_info_from_path(job_dir)
        
        # Also auto-save
        self.auto_save()
    
    def extract_customer_info_from_path(self, job_dir):
        """Extract customer name and location from job directory path"""
        try:
            # Normalize the path and split into parts
            normalized_path = os.path.normpath(job_dir)
            path_parts = normalized_path.split(os.sep)
            
            print(f"DEBUG: Extracting from path: {normalized_path}")
            print(f"DEBUG: Path parts: {path_parts}")
            
            # Find the job number (should be the last part)
            job_number = path_parts[-1] if path_parts else ""
            
            # Customer location is one level up from job number
            customer_location = path_parts[-2] if len(path_parts) >= 2 else ""
            
            # Customer name is two levels up from job number
            customer_name = path_parts[-3] if len(path_parts) >= 3 else ""
            
            print(f"DEBUG: Extracted - Job: {job_number}, Location: {customer_location}, Name: {customer_name}")
            
            # Set the extracted values
            if customer_name:
                self.customer_name_var.set(customer_name.upper())
                self.customer_name_picker.set(os.path.dirname(os.path.dirname(job_dir)))
            
            if customer_location:
                self.customer_location_var.set(customer_location.upper())
                self.customer_location_picker.set(os.path.dirname(job_dir))
            
            # Automatically find and add KOM AND OC FORM Excel file
            self.find_and_add_kom_oc_form(job_dir)
            
            # Automatically find and add Proposal Word documents
            self.find_and_add_proposal_docs(job_dir)
            
            # Automatically find and add other important files
            self.find_and_add_other_docs(job_dir)
            
            # Automatically find and add engineering files
            self.find_and_add_engineering_docs(job_dir)
            
            # Update quick access panel
            self.update_quick_access()
            
        except Exception as e:
            print(f"DEBUG: Error extracting customer info: {e}")
    
    def find_and_add_kom_oc_form(self, job_dir):
        """Find and add KOM AND OC FORM Excel file to quick access"""
        try:
            print(f"DEBUG: Looking for KOM AND OC FORM file in: {job_dir}")
            
            # Get job number from the directory path
            job_number = os.path.basename(job_dir)
            
            # Look for Excel files with "KOM AND OC FORM" in the filename
            for file in os.listdir(job_dir):
                if file.endswith('.xlsx') and 'KOM AND OC FORM' in file.upper():
                    kom_file_path = os.path.join(job_dir, file)
                    print(f"DEBUG: Found KOM AND OC FORM file: {kom_file_path}")
                    
                    # Store the file path for quick access
                    self.kom_oc_form_path = kom_file_path
                    return
            
            print(f"DEBUG: No KOM AND OC FORM file found in {job_dir}")
            self.kom_oc_form_path = None
            
        except Exception as e:
            print(f"DEBUG: Error finding KOM AND OC FORM file: {e}")
            self.kom_oc_form_path = None
    
    def open_kom_oc_form(self):
        """Open the KOM AND OC FORM Excel file"""
        if hasattr(self, 'kom_oc_form_path') and self.kom_oc_form_path and os.path.exists(self.kom_oc_form_path):
            print(f"DEBUG: Opening KOM AND OC FORM file: {self.kom_oc_form_path}")
            self.open_path(self.kom_oc_form_path)
        else:
            messagebox.showwarning("Warning", "KOM AND OC FORM file not found!")
    
    def find_and_add_proposal_docs(self, job_dir):
        """Find and add Proposal Word documents from 1. Sales\\Order folder"""
        try:
            print(f"DEBUG: Looking for Proposal documents in: {job_dir}")
            
            # Look for 1. Sales\Order folder
            sales_order_path = os.path.join(job_dir, "1. Sales", "Order")
            
            if not os.path.exists(sales_order_path):
                print(f"DEBUG: Sales\\Order folder not found: {sales_order_path}")
                self.proposal_docs = []
                return
            
            print(f"DEBUG: Found Sales\\Order folder: {sales_order_path}")
            
            # Look for Word documents with "Proposal" in the filename
            proposal_files = []
            print(f"DEBUG: Listing all files in {sales_order_path}:")
            for file in os.listdir(sales_order_path):
                print(f"DEBUG: Found file: '{file}'")
                if (file.endswith('.docx') or file.endswith('.doc')):
                    print(f"DEBUG: File is Word document: {file}")
                    if 'Proposal' in file.upper():
                        proposal_file_path = os.path.join(sales_order_path, file)
                        proposal_files.append(proposal_file_path)
                        print(f"DEBUG: Found Proposal document: {proposal_file_path}")
                    else:
                        print(f"DEBUG: File does not contain 'Proposal': {file}")
                else:
                    print(f"DEBUG: File is not Word document: {file}")
            
            # Store the proposal files for quick access
            self.proposal_docs = proposal_files
            print(f"DEBUG: Found {len(proposal_files)} Proposal documents")
            
        except Exception as e:
            print(f"DEBUG: Error finding Proposal documents: {e}")
            self.proposal_docs = []
    
    def open_proposal_doc(self, doc_path):
        """Open a specific Proposal Word document"""
        if doc_path and os.path.exists(doc_path):
            print(f"DEBUG: Opening Proposal document: {doc_path}")
            self.open_path(doc_path)
        else:
            messagebox.showwarning("Warning", "Proposal document not found!")
    
    def find_and_add_other_docs(self, job_dir):
        """Find and add other important files from 1. Sales\\Order folder"""
        try:
            print(f"DEBUG: Looking for other important files in: {job_dir}")
            
            # Look for 1. Sales\Order folder
            sales_order_path = os.path.join(job_dir, "1. Sales", "Order")
            
            if not os.path.exists(sales_order_path):
                print(f"DEBUG: Sales\\Order folder not found: {sales_order_path}")
                self.other_docs = []
                return
            
            print(f"DEBUG: Found Sales\\Order folder: {sales_order_path}")
            
            # Look for other important files
            other_files = []
            print(f"DEBUG: Listing all files in {sales_order_path}:")
            for file in os.listdir(sales_order_path):
                print(f"DEBUG: Found file: '{file}'")
                file_path = os.path.join(sales_order_path, file)
                
                # Check for Excel files with "Cost" or "Template" in filename
                if (file.endswith('.xlsx') or file.endswith('.xls')) and ('Cost' in file.upper() or 'Template' in file.upper()):
                    other_files.append(('📊', file, file_path))
                    print(f"DEBUG: Found Cost/Template Excel file: {file}")
                
                # Check for PDF files
                elif file.endswith('.pdf'):
                    other_files.append(('📄', file, file_path))
                    print(f"DEBUG: Found PDF file: {file}")
                
                # Check for any other Word documents (not already captured as proposals)
                elif (file.endswith('.docx') or file.endswith('.doc')) and 'Proposal' not in file.upper():
                    # Double-check that this file wasn't already captured as a proposal
                    is_proposal = False
                    if hasattr(self, 'proposal_docs') and self.proposal_docs:
                        for proposal_path in self.proposal_docs:
                            if os.path.basename(proposal_path) == file:
                                is_proposal = True
                                break
                    
                    if not is_proposal:
                        other_files.append(('📄', file, file_path))
                        print(f"DEBUG: Found other Word document: {file}")
                    else:
                        print(f"DEBUG: Skipping {file} - already captured as proposal")
            
            # Store the other files for quick access
            self.other_docs = other_files
            print(f"DEBUG: Found {len(other_files)} other important files")
            
        except Exception as e:
            print(f"DEBUG: Error finding other files: {e}")
            self.other_docs = []
    
    def open_other_doc(self, doc_path):
        """Open a specific other document"""
        if doc_path and os.path.exists(doc_path):
            print(f"DEBUG: Opening document: {doc_path}")
            self.open_path(doc_path)
        else:
            messagebox.showwarning("Warning", "Document not found!")
    
    def find_and_add_engineering_docs(self, job_dir):
        """Find and add engineering documents from 3. Engineering folders"""
        try:
            print(f"DEBUG: Looking for engineering documents in: {job_dir}")
            
            # Look for 3. Engineering folder
            engineering_path = os.path.join(job_dir, "3. Engineering")
            
            if not os.path.exists(engineering_path):
                print(f"DEBUG: Engineering folder not found: {engineering_path}")
                self.engineering_general_docs = []
                self.engineering_releases_docs = []
                return
            
            print(f"DEBUG: Found Engineering folder: {engineering_path}")
            
            # Find General Design files
            general_design_path = os.path.join(engineering_path, "General Design")
            self.engineering_general_docs = []
            
            if os.path.exists(general_design_path):
                print(f"DEBUG: Found General Design folder: {general_design_path}")
                for file in os.listdir(general_design_path):
                    if file.endswith('.xlsx') or file.endswith('.xls'):
                        file_path = os.path.join(general_design_path, file)
                        self.engineering_general_docs.append(file_path)
                        print(f"DEBUG: Found General Design file: {file}")
            
            # Find Releases files
            releases_path = os.path.join(engineering_path, "Releases")
            self.engineering_releases_docs = []
            
            if os.path.exists(releases_path):
                print(f"DEBUG: Found Releases folder: {releases_path}")
                for file in os.listdir(releases_path):
                    file_path = os.path.join(releases_path, file)
                    self.engineering_releases_docs.append(file_path)
                    print(f"DEBUG: Found Releases file: {file}")
            
            print(f"DEBUG: Found {len(self.engineering_general_docs)} General Design files")
            print(f"DEBUG: Found {len(self.engineering_releases_docs)} Releases files")
            
        except Exception as e:
            print(f"DEBUG: Error finding engineering documents: {e}")
            self.engineering_general_docs = []
            self.engineering_releases_docs = []
    
    def open_engineering_doc(self, doc_path):
        """Open a specific engineering document"""
        if doc_path and os.path.exists(doc_path):
            print(f"DEBUG: Opening engineering document: {doc_path}")
            self.open_path(doc_path)
        else:
            messagebox.showwarning("Warning", "Engineering document not found!")
    
    def open_drafting_doc(self, doc_path):
        """Open a specific drafting document (.dwg file)"""
        if doc_path and os.path.exists(doc_path):
            print(f"DEBUG: Opening drafting document: {doc_path}")
            self.open_path(doc_path)
        else:
            messagebox.showwarning("Warning", "Drafting document not found!")
    
    def create_d365_import(self, fabs_dir):
        """Create a new D365 Import file by copying the Excel file and renaming it"""
        try:
            # Source Excel file (not template)
            source_file = r"C:\excel\templates\XXXXX D365 IMPORT.xlsx"
            
            print(f"DEBUG: Source file: {source_file}")
            print(f"DEBUG: Source exists: {os.path.exists(source_file)}")
            
            # Check if source file exists
            if not os.path.exists(source_file):
                messagebox.showerror("File Not Found", 
                                   f"D365 Import file not found at:\n{source_file}")
                return
            
            # Get job number
            job_number = self.job_number_var.get()
            print(f"DEBUG: Job number: {job_number}")
            
            if not job_number:
                messagebox.showerror("Error", "Job number is required to create D365 Import file.")
                return
            
            # Create new filename
            new_filename = f"{job_number} D365 IMPORT.xlsx"
            new_file_path = os.path.join(fabs_dir, new_filename)
            
            print(f"DEBUG: Target path: {new_file_path}")
            print(f"DEBUG: Target exists: {os.path.exists(new_file_path)}")
            
            # Check if file already exists
            if os.path.exists(new_file_path):
                messagebox.showinfo("File Exists", 
                                  f"D365 Import file already exists:\n{new_filename}")
                return
            
            # Copy the Excel file to the new location with new name
            print(f"DEBUG: About to copy from {source_file} to {new_file_path}")
            shutil.copy2(source_file, new_file_path)
            print(f"DEBUG: Copy completed")
            print(f"DEBUG: New file exists: {os.path.exists(new_file_path)}")
            
            # Refresh Quick Access to remove the green button and show the new file
            self.update_quick_access()
            
            messagebox.showinfo("Success", 
                              f"D365 Import file created successfully!\n\n{new_filename}\n\nThe file is ready in the Fabs folder.\nClick the file button in Quick Access to open it.")
            
        except Exception as e:
            messagebox.showerror("Error", 
                               f"Failed to create D365 Import file:\n{str(e)}")
            print(f"ERROR creating D365 Import file: {e}")
            import traceback
            traceback.print_exc()
    
    def create_transmittal_notice(self, fabs_dir):
        """Create a new Transmittal Notice DWG by copying the template and renaming it"""
        try:
            # Source DWG file
            source_file = r"C:\Users\llaing\OneDrive - CECO Environmental Corp\Drafting Standards\Release Process\Drawing Release Form (Blue).dwg"
            
            print(f"DEBUG: Source file: {source_file}")
            print(f"DEBUG: Source exists: {os.path.exists(source_file)}")
            
            # Check if source file exists
            if not os.path.exists(source_file):
                messagebox.showerror("File Not Found", 
                                   f"Transmittal Notice template not found at:\n{source_file}")
                return
            
            # Get job number
            job_number = self.job_number_var.get()
            print(f"DEBUG: Job number: {job_number}")
            
            if not job_number:
                messagebox.showerror("Error", "Job number is required to create Transmittal Notice.")
                return
            
            # Create new filename
            new_filename = f"{job_number} TRANMITTAL NOTICE.dwg"
            new_file_path = os.path.join(fabs_dir, new_filename)
            
            print(f"DEBUG: Target path: {new_file_path}")
            print(f"DEBUG: Target exists: {os.path.exists(new_file_path)}")
            
            # Check if file already exists
            if os.path.exists(new_file_path):
                messagebox.showinfo("File Exists", 
                                  f"Transmittal Notice already exists:\n{new_filename}")
                return
            
            # Copy the DWG file to the new location with new name
            print(f"DEBUG: About to copy from {source_file} to {new_file_path}")
            shutil.copy2(source_file, new_file_path)
            print(f"DEBUG: Copy completed")
            print(f"DEBUG: New file exists: {os.path.exists(new_file_path)}")
            
            # Refresh Quick Access to remove the green button and show the new file
            self.update_quick_access()
            
            messagebox.showinfo("Success", 
                              f"Transmittal Notice created successfully!\n\n{new_filename}\n\nThe file is ready in the Fabs folder.\nClick the file button in Quick Access to open it.")
            
        except Exception as e:
            messagebox.showerror("Error", 
                               f"Failed to create Transmittal Notice:\n{str(e)}")
            print(f"ERROR creating Transmittal Notice: {e}")
            import traceback
            traceback.print_exc()
    
    def create_short_button_text(self, icon, filename):
        """Create short, consistent button text for files"""
        # Remove file extension
        name_without_ext = os.path.splitext(filename)[0]
        file_ext = os.path.splitext(filename)[1].upper()
        
        # Consistent labels for specific file types
        if 'PROPOSAL' in filename.upper():
            return f"{icon} Proposal"
        elif 'ENGINEERING DESIGN' in filename.upper():
            return f"{icon} Engineering Design"
        elif 'PRESSURE DROP CALCULATOR' in filename.upper():
            return f"{icon} Pressure Drop Calculator"
        elif 'SPRAY NOZZLES' in filename.upper():
            return f"{icon} Spray Nozzles"
        elif 'ELECTRICAL RELEASE' in filename.upper():
            return f"{icon} Electrical Release{file_ext}"
        elif 'GAS TRAIN RELEASE' in filename.upper():
            return f"{icon} Gas Train Release{file_ext}"
        elif 'MECHANICAL RELEASE' in filename.upper():
            return f"{icon} Mechanical Release{file_ext}"
        elif 'HEATER RELEASE' in filename.upper():
            return f"{icon} Heater Release{file_ext}"
        elif 'TANK RELEASE' in filename.upper():
            return f"{icon} Tank Release{file_ext}"
        else:
            # For all other files, show filename (truncated if too long)
            if len(name_without_ext) > 25:
                return f"{icon} {name_without_ext[:22]}..."
            else:
                return f"{icon} {name_without_ext}"
    
    def auto_save(self, *args):
        """Auto-save project when any field changes"""
        # Don't auto-save while loading project details
        if hasattr(self, '_loading_project') and self._loading_project:
            return
            
        # Only auto-save if we have a valid 5-digit job number
        job_number = self.job_number_var.get().strip()
        if self.is_valid_job_number(job_number):
            try:
                self.save_project_silent()
                # Update cover sheet button after saving
                self.update_cover_sheet_button()
            except Exception as e:
                print(f"Auto-save failed: {e}")
    
    def is_valid_job_number(self, job_number):
        """Validate that job number is exactly 5 digits"""
        if not job_number:
            return False
        # Allow base 5-digit (e.g., 12345) or suffixed copy (e.g., 12345 (1))
        clean_number = job_number.strip()
        if re.match(r"^\d{5}$", clean_number):
            return True
        if re.match(r"^\d{5} \(\d+\)$", clean_number):
            return True
        return False

    def normalize_job_number(self, job_number: str) -> str:
        """Normalize job number to '12345' or '12345 (n)' format, trimming stray spaces.

        Falls back to the first 5-digit sequence if non-standard input is provided.
        """
        if not job_number:
            return ""
        s = str(job_number).strip()
        m = re.match(r"^(\d{5})(?:\s*\(\s*(\d+)\s*\))?$", s)
        if m:
            base = m.group(1)
            suf = m.group(2)
            return f"{base} ({suf})" if suf else base
        # Fallback: extract first 5-digit sequence
        m2 = re.search(r"\d{5}", s)
        return m2.group(0) if m2 else s
    
    def save_project_silent(self):
        """Save project without showing success message"""
        job_number = self.normalize_job_number(self.job_number_var.get())
        # reflect normalized value in UI
        if job_number and job_number != self.job_number_var.get().strip():
            self.job_number_var.set(job_number)
        if not self.is_valid_job_number(job_number):
            return

        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        try:
            # Determine if this is a new project (no existing row before upsert)
            cursor.execute("SELECT id FROM projects WHERE job_number = ?", (job_number,))
            existed_before = cursor.fetchone() is not None
            # Get designer ID
            designer_id = None
            if self.assigned_to_var.get():
                cursor.execute("SELECT id FROM designers WHERE name = ?", (self.assigned_to_var.get(),))
                result = cursor.fetchone()
                if result:
                    designer_id = result[0]
            
            # Get project engineer ID
            project_engineer_id = None
            if self.project_engineer_var.get():
                cursor.execute("SELECT id FROM engineers WHERE name = ?", (self.project_engineer_var.get(),))
                result = cursor.fetchone()
                if result:
                    project_engineer_id = result[0]
            
            # Calculate duration
            duration = None
            if self.start_date_entry.get() and self.completion_date_entry.get():
                try:
                    start = datetime.strptime(self.start_date_entry.get(), "%Y-%m-%d")
                    end = datetime.strptime(self.completion_date_entry.get(), "%Y-%m-%d")
                    duration = (end - start).days
                except ValueError:
                    pass
            
            # Insert or update project
            cursor.execute(
                """
                INSERT INTO projects (
                    job_number, job_directory, customer_name, customer_name_directory,
                    customer_location, customer_location_directory, assigned_to_id, project_engineer_id,
                    assignment_date, start_date, completion_date, total_duration_days, released_to_dee, due_date
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(job_number) DO UPDATE SET
                    job_directory = excluded.job_directory,
                    customer_name = excluded.customer_name,
                    customer_name_directory = excluded.customer_name_directory,
                    customer_location = excluded.customer_location,
                    customer_location_directory = excluded.customer_location_directory,
                    assigned_to_id = excluded.assigned_to_id,
                    project_engineer_id = excluded.project_engineer_id,
                    assignment_date = excluded.assignment_date,
                    start_date = excluded.start_date,
                    completion_date = excluded.completion_date,
                    total_duration_days = excluded.total_duration_days,
                    released_to_dee = excluded.released_to_dee,
                    due_date = excluded.due_date
                """,
                (
                    job_number,
                    self.job_directory_picker.get() or None,
                    self.customer_name_var.get().upper() or None,
                    self.customer_name_picker.get() or None,
                    self.customer_location_var.get().upper() or None,
                    self.customer_location_picker.get() or None,
                    designer_id,
                    project_engineer_id,
                    self.assignment_date_entry.get() or None,
                    self.start_date_entry.get() or None,
                    self.completion_date_entry.get() or None,
                    duration,
                    self.released_to_dee_entry.get() or None,
                    self.due_date_entry.get() or None,
                ),
            )
            
            # Get project ID
            cursor.execute("SELECT id FROM projects WHERE job_number = ?", (job_number,))
            project_id = cursor.fetchone()[0]
            
            # Save workflow data (legacy, fixed sections)
            self.save_workflow_data(cursor, project_id)
            
            conn.commit()

            # Ensure template-driven workflow steps exist only for brand-new projects
            if not existed_before:
                try:
                    self._ensure_project_workflow_seeded(project_id)
                    # Recompute planned due dates chain
                    self._recompute_workflow_due_dates_for_current_project()
                except Exception:
                    pass
            
        except Exception as e:
            print(f"Silent save failed: {e}")
        finally:
            conn.close()

    def _get_active_template_id(self):
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute("SELECT id FROM workflow_templates WHERE name = ? AND is_active = 1 ORDER BY version DESC LIMIT 1", ('Standard',))
            r = cur.fetchone(); conn.close()
            return r[0] if r else None
        except Exception:
            return None

    def _ensure_project_workflow_seeded(self, project_id):
        """Seed project workflow steps from active template if none exist (new projects only)."""
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM project_workflow_steps WHERE project_id = ?", (project_id,))
            if cur.fetchone()[0] > 0:
                conn.close(); return
            tid = self._get_active_template_id()
            if not tid:
                conn.close(); return
            # Load template steps
            cur.execute("""
                SELECT id, order_index, department, group_name, title, planned_duration_days
                FROM workflow_template_steps WHERE template_id = ? ORDER BY order_index
            """, (tid,))
            steps = cur.fetchall()
            for s in steps:
                step_id, order_i, dept, group, title, dur = s
                cur.execute(
                    """
                    INSERT INTO project_workflow_steps
                    (project_id, template_id, template_step_id, order_index, department, group_name, title)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (project_id, tid, step_id, order_i, dept, group, title)
                )
            conn.commit(); conn.close()
        except Exception as e:
            print(f"Error seeding project workflow: {e}")
    
    def load_dropdown_data(self):
        """Load data for dropdown menus"""
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        # Load designers
        cursor.execute("SELECT name FROM designers ORDER BY name")
        designers = [row[0] for row in cursor.fetchall()]
        if hasattr(self, 'assigned_to_combo'):
            self.assigned_to_combo['values'] = designers
        
        # Load engineers
        cursor.execute("SELECT name FROM engineers ORDER BY name")
        engineers = [row[0] for row in cursor.fetchall()]
        if hasattr(self, 'initial_engineer_combo'):
            self.initial_engineer_combo['values'] = engineers
        
        # Set engineers for project engineer combo
        if hasattr(self, 'project_engineer_combo'):
            self.project_engineer_combo['values'] = engineers
        
        # Set engineers for all redline update combos
        for i in range(1, 5):
            combo_name = f"redline_update_{i}_engineer_combo"
            if hasattr(self, combo_name):
                getattr(self, combo_name)['values'] = engineers
        
        conn.close()
    
    def load_projects(self):
        """Load projects from database"""
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        query = """
        SELECT 
            p.job_number,
            p.customer_name,
            p.due_date,
            p.completion_date,
            COALESCE(
                p.released_to_dee,
                (SELECT release_date FROM release_to_dee rd WHERE rd.project_id = p.id ORDER BY rd.id DESC LIMIT 1)
            ) AS release_date,
            CASE 
                WHEN (
                    COALESCE(
                        p.released_to_dee,
                        (SELECT release_date FROM release_to_dee rd2 WHERE rd2.project_id = p.id ORDER BY rd2.id DESC LIMIT 1)
                    ) IS NOT NULL
                    AND COALESCE(
                        p.released_to_dee,
                        (SELECT release_date FROM release_to_dee rd3 WHERE rd3.project_id = p.id ORDER BY rd3.id DESC LIMIT 1)
                    ) != ''
                )
                OR (
                    (SELECT is_completed FROM release_to_dee rd4 WHERE rd4.project_id = p.id ORDER BY rd4.id DESC LIMIT 1) = 1
                )
                OR (p.completion_date IS NOT NULL AND p.completion_date != '') THEN 'Completed'
                WHEN p.start_date IS NOT NULL AND p.start_date != '' THEN 'In Progress'
                WHEN p.assignment_date IS NOT NULL AND p.assignment_date != '' THEN 'Assigned'
                ELSE 'Not Assigned'
            END as status
        FROM projects p
        ORDER BY 
            CASE WHEN p.due_date IS NULL OR p.due_date = '' THEN 1 ELSE 0 END,
            p.due_date ASC
        """
        
        cursor.execute(query)
        projects = cursor.fetchall()
        
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Insert projects
        for project in projects:
            job_number = project[0]
            customer_name = project[1]
            due_date = project[2] if project[2] else ""
            completion_date = project[3]
            status = project[5]
            
            # Calculate days until due
            days_until_due = ""
            if due_date and not completion_date:
                try:
                    due = datetime.strptime(due_date, "%Y-%m-%d")
                    today = datetime.now()
                    days_diff = (due - today).days
                    
                    if days_diff < 0:
                        days_until_due = f"{abs(days_diff)} overdue"
                    elif days_diff == 0:
                        days_until_due = "Today"
                    else:
                        days_until_due = str(days_diff)
                except:
                    days_until_due = ""
            
            self.tree.insert('', 'end', values=(
                job_number,
                customer_name,
                due_date,
                days_until_due,
                status
            ))

        # Apply current visibility (hide completed if needed)
        try:
            self.filter_projects()
        except Exception:
            pass
        
        conn.close()
    
    def filter_projects(self, *args):
        """Filter projects based on search term"""
        search_term = self.search_var.get().lower()
        
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            matches_search = any(search_term in str(value).lower() for value in values)
            is_completed = str(values[4]).lower() == 'completed'
            should_show = matches_search and (self.show_completed or not is_completed)
            if should_show:
                self.tree.reattach(item, '', 'end')
            else:
                self.tree.detach(item)

    def toggle_completed(self):
        """Toggle showing/hiding completed projects in the list"""
        self.show_completed = not getattr(self, 'show_completed', False)
        self.toggle_completed_btn.config(text=('Hide Completed' if self.show_completed else 'Show Completed'))
        # If we're showing completed, repopulate the list so any previously-detached
        # rows are restored immediately. Otherwise, just apply the filter to hide them.
        if self.show_completed:
            try:
                self.load_projects()
            except Exception:
                self.filter_projects()
        else:
            self.filter_projects()
    
    def sort_by_job_number(self):
        """Sort projects by job number (toggle ascending/descending)"""
        # Get all current items and their values
        items = []
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            items.append(values)
        
        # Clear the tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Sort by job number (convert to int for proper numeric sorting)
        # Handle both numeric and non-numeric job numbers
        def job_sort_key(x):
            job_num = str(x[0]).strip()
            if job_num.isdigit():
                return int(job_num)
            else:
                # For non-numeric, try to extract first sequence of digits
                import re
                digits = re.findall(r'\d+', job_num)
                if digits:
                    return int(digits[0])
                else:
                    return 0
        
        # Sort with current direction
        sorted_items = sorted(items, key=job_sort_key, reverse=not self.job_sort_ascending)
        
        # Toggle direction for next time
        self.job_sort_ascending = not self.job_sort_ascending
        
        # Update button text to show current direction
        direction = "↑" if self.job_sort_ascending else "↓"
        self.sort_job_btn.config(text=f"Job # {direction}")
        
        # Add sorted items back
        for item in sorted_items:
            self.tree.insert('', 'end', values=item)
    
    def sort_by_customer(self):
        """Sort projects by customer name (toggle ascending/descending)"""
        # Get all current items and their values
        items = []
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            items.append(values)
        
        # Clear the tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Sort by customer name (case-insensitive)
        sorted_items = sorted(items, key=lambda x: x[1].upper() if x[1] else "", reverse=not self.customer_sort_ascending)
        
        # Toggle direction for next time
        self.customer_sort_ascending = not self.customer_sort_ascending
        
        # Update button text to show current direction
        direction = "↑" if self.customer_sort_ascending else "↓"
        self.sort_customer_btn.config(text=f"Customer {direction}")
        
        # Add sorted items back
        for item in sorted_items:
            self.tree.insert('', 'end', values=item)
    
    def sort_by_due_date(self):
        """Sort projects by due date - earliest on top when ascending"""
        # Get all projects with due dates from database
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT job_number, customer_name, due_date, completion_date,
                   CASE 
                       WHEN completion_date IS NOT NULL AND completion_date != '' THEN 'Completed'
                       WHEN start_date IS NOT NULL AND start_date != '' THEN 'In Progress'
                       WHEN assignment_date IS NOT NULL AND assignment_date != '' THEN 'Assigned'
                       ELSE 'Not Assigned'
                   END as status
            FROM projects
            ORDER BY 
                CASE 
                    WHEN due_date IS NULL OR due_date = '' THEN 1
                    ELSE 0
                END,
                due_date """ + ("ASC" if self.due_date_sort_ascending else "DESC") + """
        """)
        
        projects = cursor.fetchall()
        conn.close()
        
        # Clear the tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Add sorted items back
        for project in projects:
            job_num, customer, due_date, completion_date, status = project
            
            # Calculate days until due
            days_until_due = ""
            if due_date and not completion_date:
                try:
                    due = datetime.strptime(due_date, "%Y-%m-%d")
                    today = datetime.now()
                    days_diff = (due - today).days
                    
                    if days_diff < 0:
                        days_until_due = f"{abs(days_diff)} overdue"
                    elif days_diff == 0:
                        days_until_due = "Today"
                    else:
                        days_until_due = str(days_diff)
                except:
                    days_until_due = ""
            
            self.tree.insert('', 'end', values=(job_num, customer or '', due_date or '', 
                                               days_until_due, status))
        
        # Toggle direction for next time
        self.due_date_sort_ascending = not self.due_date_sort_ascending
        
        # Update button text to show current direction
        direction = "↑" if self.due_date_sort_ascending else "↓"
        self.sort_due_date_btn.config(text=f"Due Date {direction}")
    
    def on_project_select(self, event):
        """Handle project selection with row highlighting"""
        selection = self.tree.selection()
        if not selection:
            print("DEBUG: No selection")
            return
        
        # Remove 'selected' tag from all items
        for item in self.tree.get_children():
            self.tree.item(item, tags=())
        
        # Add 'selected' tag to selected item
        self.tree.item(selection[0], tags=('selected',))
        
        item = self.tree.item(selection[0])
        job_number = item['values'][0]
        print(f"DEBUG: Selected project: {job_number}")
        
        # Clear notes area immediately to avoid showing stale notes while loading
        try:
            if hasattr(self, 'notes_text'):
                self.notes_text.delete('1.0', tk.END)
        except Exception:
            pass

        # Set current project before loading details
        self.current_project = job_number
        
        self.load_project_details(job_number)
    
    def clear_workflow_data(self):
        """Clear all workflow data before loading new project"""
        # Temporarily disable auto-save to prevent saving empty values
        self._loading_project = True
        
        # Clear initial redline (guarded)
        if hasattr(self, 'initial_redline_var'):
            self.initial_redline_var.set(False)
        if hasattr(self, 'initial_engineer_var'):
            self.initial_engineer_var.set("")
        if hasattr(self, 'initial_date_entry'):
            self.initial_date_entry.set("")
        
        # Clear redline updates
        for i in range(1, 5):
            if hasattr(self, f"redline_update_{i}_var"):
                getattr(self, f"redline_update_{i}_var").set(False)
            if hasattr(self, f"redline_update_{i}_engineer_var"):
                getattr(self, f"redline_update_{i}_engineer_var").set("")
            if hasattr(self, f"redline_update_{i}_date_entry"):
                getattr(self, f"redline_update_{i}_date_entry").set("")
        
        # Clear OPS review
        if hasattr(self, 'ops_review_var'):
            self.ops_review_var.set(False)
        if hasattr(self, 'ops_review_date_entry'):
            self.ops_review_date_entry.set("")
        
        # Clear D365 BOM Entry
        if hasattr(self, 'd365_bom_var'):
            self.d365_bom_var.set(False)
        if hasattr(self, 'd365_bom_date_entry'):
            self.d365_bom_date_entry.set("")
        
        # Clear Peter Weck review
        if hasattr(self, 'peter_weck_var'):
            self.peter_weck_var.set(False)
        if hasattr(self, 'peter_weck_date_entry'):
            self.peter_weck_date_entry.set("")
        
        # Clear release to Dee
        if hasattr(self, 'release_fixed_errors_var'):
            self.release_fixed_errors_var.set(False)
        if hasattr(self, 'missing_prints_date_entry'):
            self.missing_prints_date_entry.set("")
        if hasattr(self, 'd365_updates_date_entry'):
            self.d365_updates_date_entry.set("")
        if hasattr(self, 'other_notes_var'):
            self.other_notes_var.set("")
        if hasattr(self, 'other_date_entry'):
            self.other_date_entry.set("")
        if hasattr(self, 'release_due_date_entry'):
            self.release_due_date_entry.set("")
        if hasattr(self, 'release_due_display_var'):
            self.release_due_display_var.set("")

    def load_project_details(self, job_number):
        """Load details for selected project"""
        print(f"DEBUG: Loading project details for: {job_number}")
        
        # Normalize job number for lookup: accept '12345' or '12345 (n)' as-is
        raw_job = str(job_number).strip()
        if re.match(r"^\d{5}( \(\d+\))?$", raw_job):
            clean_job_number = raw_job
        else:
            # Fallback: extract first 5-digit sequence
            m = re.search(r"\d{5}", raw_job)
            clean_job_number = m.group(0) if m else raw_job
        
        print(f"DEBUG: Cleaned job number: {clean_job_number}")
        
        # Clear workflow data first to prevent showing old data
        self.clear_workflow_data()
        
        # Temporarily disable auto-save to prevent interference
        self._loading_project = True
        
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        # Load main project data
        query = """
        SELECT p.job_number, p.job_directory, p.customer_name, p.customer_name_directory,
               p.customer_location, p.customer_location_directory, d.name, e.name,
               p.assignment_date, p.start_date, p.completion_date, 
               p.total_duration_days, p.released_to_dee, p.due_date
        FROM projects p
        LEFT JOIN designers d ON p.assigned_to_id = d.id
        LEFT JOIN engineers e ON p.project_engineer_id = e.id
        WHERE p.job_number = ?
        """
        
        cursor.execute(query, (clean_job_number,))
        project = cursor.fetchone()
        
        print(f"DEBUG: Project data loaded: {project}")
        
        if project:
            self.job_number_var.set(project[0])
            self.job_directory_picker.set(project[1] or "")
            self.customer_name_var.set(project[2] or "")
            self.customer_name_picker.set(project[3] or "")
            self.customer_location_var.set(project[4] or "")
            self.customer_location_picker.set(project[5] or "")
            self.assigned_to_var.set(project[6] or "")
            self.project_engineer_var.set(project[7] or "")
            self.assignment_date_entry.set(project[8] or "")
            self.start_date_entry.set(project[9] or "")
            self.completion_date_entry.set(project[10] or "")
            self.duration_var.set(f"{project[11]} days" if project[11] else "N/A")
            self.released_to_dee_entry.set(project[12] or "")
            self.due_date_entry.set(project[13] or "")
        
        # Load workflow data (guard errors so notes still load)
        try:
            self.load_workflow_data(clean_job_number, cursor)
        except Exception as e:
            print(f"Error loading workflow data: {e}")
        
        # Update quick access panel
        try:
            self.update_quick_access()
        except Exception as e:
            print(f"Error updating quick access: {e}")
        
        # Update specifications panel
        try:
            if hasattr(self, 'project_details_frame'):
                self.update_specifications(self.project_details_frame)
        except Exception as e:
            print(f"Error updating specifications: {e}")
        
        # Update cover sheet button
        try:
            self.update_cover_sheet_button()
        except Exception as e:
            print(f"Error updating cover sheet button: {e}")
        
        # Load job notes last to ensure the UI shows correct per-job notes
        try:
            self.load_job_notes(clean_job_number)
        except Exception as e:
            print(f"Error loading notes: {e}")

        # Re-enable auto-save
        self._loading_project = False

        conn.close()
        
        # Refresh template-driven workflow section after loading a project
        try:
            if hasattr(self, 'template_workflow_section'):
                self.create_template_workflow_content(self.template_workflow_section.content)
        except Exception:
            pass
    
    def load_workflow_data(self, job_number, cursor):
        """Load workflow data for selected project"""
        # Get project ID
        cursor.execute("SELECT id FROM projects WHERE job_number = ?", (job_number,))
        project_result = cursor.fetchone()
        if not project_result:
            return
        
        project_id = project_result[0]
        
        # Load initial redline
        cursor.execute("""
            SELECT ir.redline_date, e.name, ir.is_completed
            FROM initial_redline ir
            LEFT JOIN engineers e ON ir.engineer_id = e.id
            WHERE ir.project_id = ?
        """, (project_id,))
        initial_redline = cursor.fetchone()
        
        if initial_redline:
            self.initial_redline_var.set(bool(initial_redline[2]))
            self.initial_engineer_var.set(initial_redline[1] or "")
            self.initial_date_entry.set(initial_redline[0] or "")
        
        # Load redline updates
        cursor.execute("""
            SELECT ru.update_cycle, ru.update_date, e.name, ru.is_completed
            FROM redline_updates ru
            LEFT JOIN engineers e ON ru.engineer_id = e.id
            WHERE ru.project_id = ?
            ORDER BY ru.update_cycle
        """, (project_id,))
        redline_updates = cursor.fetchall()
        
        for update in redline_updates:
            cycle = update[0]
            if 1 <= cycle <= 4:
                # Only set the values if the widgets exist
                if hasattr(self, f"redline_update_{cycle}_var"):
                    getattr(self, f"redline_update_{cycle}_var").set(bool(update[3]))
                if hasattr(self, f"redline_update_{cycle}_engineer_var"):
                    getattr(self, f"redline_update_{cycle}_engineer_var").set(update[2] or "")
                if hasattr(self, f"redline_update_{cycle}_date_entry"):
                    getattr(self, f"redline_update_{cycle}_date_entry").set(update[1] or "")
        
        # Load OPS review
        cursor.execute("""
            SELECT review_date, is_completed
            FROM ops_review
            WHERE project_id = ?
        """, (project_id,))
        ops_review = cursor.fetchone()
        
        if ops_review:
            self.ops_review_var.set(bool(ops_review[1]))
            self.ops_review_date_entry.set(ops_review[0] or "")
        
        # Load D365 BOM Entry
        cursor.execute("""
            SELECT entry_date, is_completed
            FROM d365_bom_entry
            WHERE project_id = ?
        """, (project_id,))
        d365_bom = cursor.fetchone()
        
        if d365_bom:
            self.d365_bom_var.set(bool(d365_bom[1]))
            self.d365_bom_date_entry.set(d365_bom[0] or "")
        
        # Load Peter Weck review
        cursor.execute("""
            SELECT fixed_errors_date, is_completed
            FROM peter_weck_review
            WHERE project_id = ?
        """, (project_id,))
        peter_weck = cursor.fetchone()
        
        if peter_weck:
            self.peter_weck_var.set(bool(peter_weck[1]))
            self.peter_weck_date_entry.set(peter_weck[0] or "")
        
        # Load release to Dee
        cursor.execute("""
            SELECT release_date, missing_prints_date, d365_updates_date, 
                   other_notes, other_date, due_date, is_completed
            FROM release_to_dee
            WHERE project_id = ?
        """, (project_id,))
        release_data = cursor.fetchone()
        
        if release_data:
            self.release_fixed_errors_var.set(bool(release_data[6]))
            self.missing_prints_date_entry.set(release_data[1] or "")
            self.d365_updates_date_entry.set(release_data[2] or "")
            self.other_notes_var.set(release_data[3] or "")
            self.other_date_entry.set(release_data[4] or "")
            self.release_due_date_entry.set(release_data[5] or "")
            # Update the due date display
            self.update_release_due_display()
            
            # Sync the released_to_dee field in the main projects table
            if release_data[0]:  # If there's a release date
                cursor.execute("""
                    UPDATE projects 
                    SET released_to_dee = ?
                    WHERE id = ?
                """, (release_data[0], project_id))
                # Ensure the update is persisted immediately
                try:
                    cursor.connection.commit()
                except Exception:
                    pass
    
    def new_project(self):
        """Clear form for new project"""
        # Avoid autosave triggers while clearing
        self._loading_project = True

        # Unselect any selected project in the tree and clear highlight
        try:
            for item in self.tree.get_children():
                self.tree.item(item, tags=())
            self.tree.selection_set(())
            self.tree.focus("")
        except Exception:
            pass

        # Reset selection context
        self.current_project = None

        # Clear main project fields (all blank; no defaults)
        self.job_number_var.set("")
        self.job_directory_picker.set("")
        self.customer_name_var.set("")
        self.customer_name_picker.set("")
        self.customer_location_var.set("")
        self.customer_location_picker.set("")
        self.assigned_to_var.set("")
        self.project_engineer_var.set("")
        self.assignment_date_entry.set("")
        self.due_date_entry.set("")
        self.start_date_entry.set("")
        self.completion_date_entry.set("")
        self.duration_var.set("")
        self.released_to_dee_entry.set("")

        # Clear workflow fields (legacy sections guarded)
        if hasattr(self, 'initial_redline_var'):
            self.initial_redline_var.set(False)
        if hasattr(self, 'initial_engineer_var'):
            self.initial_engineer_var.set("")
        if hasattr(self, 'initial_date_entry'):
            self.initial_date_entry.set("")

        for i in range(1, 5):
            # Uncheck, clear engineer dropdown, and clear dates
            if hasattr(self, f"redline_update_{i}_var"):
                getattr(self, f"redline_update_{i}_var").set(False)
            if hasattr(self, f"redline_update_{i}_engineer_var"):
                getattr(self, f"redline_update_{i}_engineer_var").set("")
            if hasattr(self, f"redline_update_{i}_date_entry"):
                getattr(self, f"redline_update_{i}_date_entry").set("")

        if hasattr(self, 'ops_review_var'):
            self.ops_review_var.set(False)
        if hasattr(self, 'ops_review_date_entry'):
            self.ops_review_date_entry.set("")
        if hasattr(self, 'peter_weck_var'):
            self.peter_weck_var.set(False)
        if hasattr(self, 'peter_weck_date_entry'):
            self.peter_weck_date_entry.set("")
        if hasattr(self, 'd365_bom_var'):
            self.d365_bom_var.set(False)
        if hasattr(self, 'd365_bom_date_entry'):
            self.d365_bom_date_entry.set("")
        if hasattr(self, 'release_fixed_errors_var'):
            self.release_fixed_errors_var.set(False)
        if hasattr(self, 'missing_prints_date_entry'):
            self.missing_prints_date_entry.set("")
        if hasattr(self, 'd365_updates_date_entry'):
            self.d365_updates_date_entry.set("")
        if hasattr(self, 'other_notes_var'):
            self.other_notes_var.set("")
        if hasattr(self, 'other_date_entry'):
            self.other_date_entry.set("")
        if hasattr(self, 'release_due_date_entry'):
            self.release_due_date_entry.set("")
        if hasattr(self, 'release_due_display_var'):
            self.release_due_display_var.set("")

        # Clear notes context and text area
        try:
            self.current_job_notes = ""
            if hasattr(self, 'notes_text'):
                self.notes_text.delete('1.0', tk.END)
        except Exception:
            pass

        # Clear KOM file path and all document lists
        if hasattr(self, 'kom_oc_form_path'):
            self.kom_oc_form_path = None
        if hasattr(self, 'proposal_docs'):
            self.proposal_docs = []
        if hasattr(self, 'other_docs'):
            self.other_docs = []
        if hasattr(self, 'engineering_general_docs'):
            self.engineering_general_docs = []
        if hasattr(self, 'engineering_releases_docs'):
            self.engineering_releases_docs = []

        # Update quick access and specifications panels to reflect no selection
        self.update_quick_access()
        try:
            if hasattr(self, 'project_details_frame'):
                self.update_specifications(self.project_details_frame)
        except Exception:
            pass

        # Update cover sheet action state
        try:
            self.update_cover_sheet_button()
        except Exception:
            pass

        # Re-enable autosave after clearing
        self._loading_project = False
    
    def save_project(self):
        """Save project to database"""
        job_number = self.normalize_job_number(self.job_number_var.get())
        if job_number and job_number != self.job_number_var.get().strip():
            self.job_number_var.set(job_number)
        if not job_number:
            messagebox.showerror("Error", "Job number is required!")
            return
        
        if not self.is_valid_job_number(job_number):
            messagebox.showerror("Error", "Job number must be exactly 5 digits (e.g., 12345)!")
            return
        
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        try:
            # Get designer ID
            designer_id = None
            if self.assigned_to_var.get():
                cursor.execute("SELECT id FROM designers WHERE name = ?", (self.assigned_to_var.get(),))
                result = cursor.fetchone()
                if result:
                    designer_id = result[0]
            
            # Get project engineer ID
            project_engineer_id = None
            if self.project_engineer_var.get():
                cursor.execute("SELECT id FROM engineers WHERE name = ?", (self.project_engineer_var.get(),))
                result = cursor.fetchone()
                if result:
                    project_engineer_id = result[0]
            
            # Calculate duration
            duration = None
            if self.start_date_entry.get() and self.completion_date_entry.get():
                try:
                    start = datetime.strptime(self.start_date_entry.get(), "%Y-%m-%d")
                    end = datetime.strptime(self.completion_date_entry.get(), "%Y-%m-%d")
                    duration = (end - start).days
                except ValueError:
                    pass
            
            # Insert or update project (preserve row id)
            cursor.execute(
                """
                INSERT INTO projects (
                    job_number, job_directory, customer_name, customer_name_directory,
                    customer_location, customer_location_directory, assigned_to_id, project_engineer_id,
                    assignment_date, start_date, completion_date, total_duration_days, released_to_dee, due_date
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(job_number) DO UPDATE SET
                    job_directory = excluded.job_directory,
                    customer_name = excluded.customer_name,
                    customer_name_directory = excluded.customer_name_directory,
                    customer_location = excluded.customer_location,
                    customer_location_directory = excluded.customer_location_directory,
                    assigned_to_id = excluded.assigned_to_id,
                    project_engineer_id = excluded.project_engineer_id,
                    assignment_date = excluded.assignment_date,
                    start_date = excluded.start_date,
                    completion_date = excluded.completion_date,
                    total_duration_days = excluded.total_duration_days,
                    released_to_dee = excluded.released_to_dee,
                    due_date = excluded.due_date
                """,
                (
                    job_number,
                    self.job_directory_picker.get() or None,
                    self.customer_name_var.get().upper() or None,
                    self.customer_name_picker.get() or None,
                    self.customer_location_var.get().upper() or None,
                    self.customer_location_picker.get() or None,
                    designer_id,
                    project_engineer_id,
                    self.assignment_date_entry.get() or None,
                    self.start_date_entry.get() or None,
                    self.completion_date_entry.get() or None,
                    duration,
                    self.released_to_dee_entry.get() or None,
                    self.due_date_entry.get() or None,
                ),
            )
            
            # Get project ID
            cursor.execute("SELECT id FROM projects WHERE job_number = ?", (self.job_number_var.get(),))
            project_id = cursor.fetchone()[0]
            
            # Save workflow data
            self.save_workflow_data(cursor, project_id)
            
            conn.commit()
            messagebox.showinfo("Success", "Project saved successfully!")
            self.load_projects()
            self.update_quick_access()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save project: {str(e)}")
        finally:
            conn.close()
    
    def save_workflow_data(self, cursor, project_id):
        """Save workflow data for the project"""
        # Make legacy saves safe if the corresponding UI fields were removed
        # Initial redline
        try:
            engineer_id = None
            if hasattr(self, 'initial_engineer_var') and self.initial_engineer_var.get():
                cursor.execute("SELECT id FROM engineers WHERE name = ?", (self.initial_engineer_var.get(),))
                result = cursor.fetchone()
                if result:
                    engineer_id = result[0]
            cursor.execute(
                """
                INSERT OR REPLACE INTO initial_redline 
                (project_id, engineer_id, redline_date, is_completed)
                VALUES (?, ?, ?, ?)
                """,
                (
                    project_id,
                    engineer_id,
                    self.initial_date_entry.get() if hasattr(self, 'initial_date_entry') else None,
                    self.initial_redline_var.get() if hasattr(self, 'initial_redline_var') else 0,
                ),
            )
        except Exception:
            pass

        # Redline updates
        for i in range(1, 5):
            try:
                var_name = f"redline_update_{i}_var"
                engineer_var_name = f"redline_update_{i}_engineer_var"
                date_entry_name = f"redline_update_{i}_date_entry"

                engineer_id = None
                if hasattr(self, engineer_var_name) and getattr(self, engineer_var_name).get():
                    cursor.execute("SELECT id FROM engineers WHERE name = ?", (getattr(self, engineer_var_name).get(),))
                    result = cursor.fetchone()
                    if result:
                        engineer_id = result[0]

                date_value = getattr(self, date_entry_name).get() if hasattr(self, date_entry_name) else None
                checkbox_value = getattr(self, var_name).get() if hasattr(self, var_name) else 0

                cursor.execute(
                    """
                    INSERT OR REPLACE INTO redline_updates 
                    (project_id, engineer_id, update_date, update_cycle, is_completed)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (project_id, engineer_id, date_value, i, checkbox_value),
                )
            except Exception:
                pass

        # OPS review
        try:
            review_date = self.ops_review_date_entry.get() if hasattr(self, 'ops_review_date_entry') else None
            is_completed = self.ops_review_var.get() if hasattr(self, 'ops_review_var') else 0
            cursor.execute(
                """
                INSERT OR REPLACE INTO ops_review 
                (project_id, review_date, is_completed)
                VALUES (?, ?, ?)
                """,
                (project_id, review_date, is_completed),
            )
        except Exception:
            pass

        # D365 BOM Entry
        try:
            entry_date = self.d365_bom_date_entry.get() if hasattr(self, 'd365_bom_date_entry') else None
            is_completed = self.d365_bom_var.get() if hasattr(self, 'd365_bom_var') else 0
            cursor.execute(
                """
                INSERT OR REPLACE INTO d365_bom_entry 
                (project_id, entry_date, is_completed)
                VALUES (?, ?, ?)
                """,
                (project_id, entry_date, is_completed),
            )
        except Exception:
            pass

        # Peter Weck review
        try:
            fixed_date = self.peter_weck_date_entry.get() if hasattr(self, 'peter_weck_date_entry') else None
            is_completed = self.peter_weck_var.get() if hasattr(self, 'peter_weck_var') else 0
            cursor.execute(
                """
                INSERT OR REPLACE INTO peter_weck_review 
                (project_id, fixed_errors_date, is_completed)
                VALUES (?, ?, ?)
                """,
                (project_id, fixed_date, is_completed),
            )
        except Exception:
            pass

        # Release to Dee
        try:
            release_date = self.released_to_dee_entry.get() if hasattr(self, 'released_to_dee_entry') else None
            missing = self.missing_prints_date_entry.get() if hasattr(self, 'missing_prints_date_entry') else None
            updates = self.d365_updates_date_entry.get() if hasattr(self, 'd365_updates_date_entry') else None
            notes = self.other_notes_var.get() if hasattr(self, 'other_notes_var') else None
            other_date = self.other_date_entry.get() if hasattr(self, 'other_date_entry') else None
            due = self.release_due_date_entry.get() if hasattr(self, 'release_due_date_entry') else None
            is_completed = self.release_fixed_errors_var.get() if hasattr(self, 'release_fixed_errors_var') else 0
            cursor.execute(
                """
                INSERT OR REPLACE INTO release_to_dee 
                (project_id, release_date, missing_prints_date, d365_updates_date, 
                 other_notes, other_date, due_date, is_completed)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (project_id, release_date, missing, updates, notes, other_date, due, is_completed),
            )
            cursor.execute("UPDATE projects SET released_to_dee = ? WHERE id = ?", (release_date, project_id))
        except Exception:
            pass
    
    def delete_project(self):
        """Delete selected project"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a project to delete!")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this project?"):
            item = self.tree.item(selection[0])
            job_number = item['values'][0]
            
            # Clean the job number for lookup: accept '12345' or '12345 (n)' as-is
            raw_job = str(job_number).strip()
            if re.match(r"^\d{5}( \(\d+\))?$", raw_job):
                clean_job_number = raw_job
            else:
                # Fallback: extract first 5-digit sequence
                m = re.search(r"\d{5}", raw_job)
                clean_job_number = m.group(0) if m else raw_job
            
            print(f"DEBUG: Deleting project - Original: {job_number}, Cleaned: {clean_job_number}")
            
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            
            try:
                # Get project ID
                cursor.execute("SELECT id FROM projects WHERE job_number = ?", (clean_job_number,))
                project_result = cursor.fetchone()
                if project_result:
                    project_id = project_result[0]
                    print(f"DEBUG: Found project ID: {project_id}")
                    
                    # Delete all related workflow data
                    print("DEBUG: Deleting workflow data...")
                    cursor.execute("DELETE FROM initial_redline WHERE project_id = ?", (project_id,))
                    cursor.execute("DELETE FROM redline_updates WHERE project_id = ?", (project_id,))
                    cursor.execute("DELETE FROM ops_review WHERE project_id = ?", (project_id,))
                    cursor.execute("DELETE FROM d365_bom_entry WHERE project_id = ?", (project_id,))
                    cursor.execute("DELETE FROM peter_weck_review WHERE project_id = ?", (project_id,))
                    cursor.execute("DELETE FROM release_to_dee WHERE project_id = ?", (project_id,))
                    print("DEBUG: Workflow data deleted")
                
                # Delete project
                print("DEBUG: Deleting main project...")
                cursor.execute("DELETE FROM projects WHERE job_number = ?", (clean_job_number,))
                rows_deleted = cursor.rowcount
                print(f"DEBUG: Rows deleted: {rows_deleted}")
                
                # Commit changes
                conn.commit()
                print("DEBUG: Changes committed")
                
                if rows_deleted > 0:
                    print("DEBUG: Project deleted successfully")
                    messagebox.showinfo("Success", f"Project {clean_job_number} deleted successfully!")
                    self.load_projects()
                    self.new_project()
                else:
                    print("DEBUG: No project found to delete")
                    messagebox.showwarning("Warning", f"No project found with job number: {clean_job_number}")
                    
            except Exception as e:
                print(f"DEBUG: Error during deletion: {e}")
                print(f"DEBUG: Error type: {type(e)}")
                import traceback
                traceback.print_exc()
                messagebox.showerror("Error", f"Failed to delete project: {str(e)}")
            finally:
                conn.close()

    def duplicate_project(self):
        """Duplicate the selected project, creating a suffixed job number and adding a timestamped note."""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select a project to duplicate!")
            return

        values = self.tree.item(selection[0], 'values')
        if not values:
            messagebox.showwarning("Warning", "No job data found!")
            return

        original_display = str(values[0]).strip()

        # Determine base 5-digit job number
        base_match = re.match(r"^(\d{5})(?: \((\d+)\))?$", original_display)
        if base_match:
            base_job = base_match.group(1)
        else:
            m = re.search(r"\d{5}", original_display)
            if not m:
                messagebox.showerror("Error", "Selected job number is not in a recognized format.")
                return
            base_job = m.group(0)

        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cur = conn.cursor()

            # Find a unique next suffix like "12345 (1)", "12345 (2)", ...
            suffix = 1
            while True:
                candidate = f"{base_job} ({suffix})"
                cur.execute("SELECT 1 FROM projects WHERE job_number = ?", (candidate,))
                if not cur.fetchone():
                    new_job_number = candidate
                    break
                suffix += 1

            # Prefer copying from the exact selected job if it exists; else from the base
            src_job = original_display
            cur.execute("SELECT 1 FROM projects WHERE job_number = ?", (src_job,))
            if not cur.fetchone():
                src_job = base_job

            cur.execute(
                """
                SELECT job_directory, customer_name, customer_name_directory,
                       customer_location, customer_location_directory, assigned_to_id,
                       project_engineer_id, assignment_date, start_date, completion_date,
                       total_duration_days, released_to_dee, due_date
                FROM projects WHERE job_number = ?
                """,
                (src_job,)
            )
            row = cur.fetchone()
            if not row:
                conn.close()
                messagebox.showerror("Error", f"Source job '{src_job}' not found in database.")
                return

            # Insert the duplicated project
            cur.execute(
                """
                INSERT INTO projects (
                    job_number, job_directory, customer_name, customer_name_directory,
                    customer_location, customer_location_directory, assigned_to_id, project_engineer_id,
                    assignment_date, start_date, completion_date, total_duration_days, released_to_dee, due_date
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (new_job_number,) + tuple(row)
            )

            # Copy workflow rows and checklist tasks from source project if present
            # Resolve project ids
            cur.execute("SELECT id FROM projects WHERE job_number = ?", (src_job,))
            src_pid_row = cur.fetchone()
            cur.execute("SELECT id FROM projects WHERE job_number = ?", (new_job_number,))
            new_pid_row = cur.fetchone()
            if src_pid_row and new_pid_row:
                src_pid = src_pid_row[0]
                new_pid = new_pid_row[0]

                # Fetch source steps
                cur.execute(
                    """
                    SELECT id, template_id, template_step_id, order_index, department, group_name, title
                    FROM project_workflow_steps
                    WHERE project_id = ?
                    ORDER BY order_index
                    """,
                    (src_pid,)
                )
                steps = cur.fetchall() or []

                step_id_map = {}
                for (old_step_id, template_id, template_step_id, order_index, department, group_name, title) in steps:
                    # Insert a fresh step row with cleared state; planned due recomputed later
                    cur.execute(
                        """
                        INSERT INTO project_workflow_steps (
                            project_id, template_id, template_step_id, order_index, department, group_name, title,
                            start_flag, start_ts, completed_flag, completed_ts,
                            transfer_to_name, transfer_to_ts, received_from_name, received_from_ts,
                            planned_due_date, actual_completed_date, actual_duration_days, actual_duration_minutes
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, 0, NULL, 0, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL, NULL)
                        """,
                        (new_pid, template_id, template_step_id, order_index, department, group_name, title)
                    )
                    new_step_id = cur.lastrowid
                    step_id_map[old_step_id] = new_step_id

                # Copy any checklist tasks, reset checked state
                if step_id_map:
                    for old_sid, new_sid in step_id_map.items():
                        cur.execute(
                            """
                            SELECT template_task_id, order_index, title
                            FROM project_step_tasks
                            WHERE project_step_id = ?
                            ORDER BY order_index
                            """,
                            (old_sid,)
                        )
                        tasks = cur.fetchall() or []
                        for template_task_id, order_index, title in tasks:
                            cur.execute(
                                """
                                INSERT INTO project_step_tasks (
                                    project_step_id, template_task_id, order_index, title, is_checked, checked_ts
                                ) VALUES (?, ?, ?, ?, 0, NULL)
                                """,
                                (new_sid, template_task_id, order_index, title)
                            )

            conn.commit()
            conn.close()

            # Add a stamped note to the new job
            ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            note = f"Duplicated from {src_job} on {ts}"
            try:
                append_job_note(new_job_number, note)
            except Exception:
                # Silently ignore note failure; duplication already succeeded
                pass

            messagebox.showinfo("Duplicated", f"Created {new_job_number} from {src_job}.")
            # Refresh list and select the new job
            self.load_projects()
            try:
                self.preload_job(new_job_number)
                # Recompute planned due dates for the duplicated workflow
                try:
                    prev = getattr(self, 'current_project', None)
                    self.current_project = new_job_number
                    self._recompute_workflow_due_dates_for_current_project()
                    self.current_project = prev
                except Exception:
                    pass
            except Exception:
                pass
        except Exception as e:
            try:
                conn.close()
            except Exception:
                pass
            messagebox.showerror("Error", f"Failed to duplicate project: {str(e)}")
    
    def calculate_duration(self, *args):
        """Calculate project duration"""
        start_date = self.start_date_entry.get()
        completion_date = self.completion_date_entry.get()
        
        if start_date and completion_date:
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d")
                end = datetime.strptime(completion_date, "%Y-%m-%d")
                duration = (end - start).days
                self.duration_var.set(f"{duration} days")
            except ValueError:
                self.duration_var.set("Invalid dates")
        else:
            self.duration_var.set("")
    
    def set_start_date(self, *args):
        """Set start date to assignment date if not already set"""
        if not self.start_date_entry.get() and self.assignment_date_entry.get():
            self.start_date_entry.set(self.assignment_date_entry.get())
    
    def open_dashboard(self):
        """Open the dashboard application"""
        try:
            if os.path.exists('dashboard.py'):
                subprocess.Popen([sys.executable, 'dashboard.py'])
            else:
                messagebox.showerror("Error", "dashboard.py not found in current directory")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to launch Dashboard:\n{str(e)}")
    
    def export_data(self):
        """Export data to JSON"""
        self.db_manager.export_to_json()
        messagebox.showinfo("Success", "Data exported to JSON successfully!")
    
    def import_data(self):
        """Import data from JSON"""
        self.db_manager.import_from_json()
        self.load_projects()
        self.load_dropdown_data()
        messagebox.showinfo("Success", "Data imported from JSON successfully!")
    
    def toggle_fullscreen(self):
        """Toggle fullscreen mode"""
        if self.root.attributes('-fullscreen'):
            self.root.attributes('-fullscreen', False)
        else:
            self.root.attributes('-fullscreen', True)
    
    def exit_fullscreen(self):
        """Exit fullscreen mode"""
        self.root.attributes('-fullscreen', False)
    
    def _bind_mousewheel(self, canvas, frame):
        """Bind mouse wheel scrolling to a canvas"""
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _bind_to_mousewheel(event):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _unbind_from_mousewheel(event):
            canvas.unbind_all("<MouseWheel>")
        
        # Bind when mouse enters the canvas or frame
        canvas.bind('<Enter>', _bind_to_mousewheel)
        frame.bind('<Enter>', _bind_to_mousewheel)
        
        # Unbind when mouse leaves
        canvas.bind('<Leave>', _unbind_from_mousewheel)
        frame.bind('<Leave>', _unbind_from_mousewheel)
    
    def load_job_notes(self, job_number):
        """Load notes for the selected job"""
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            
            # Create notes table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS job_notes (
                    job_number TEXT PRIMARY KEY,
                    notes TEXT
                )
            """)
            
            # Load notes for this job
            cursor.execute("SELECT notes FROM job_notes WHERE job_number = ?", (job_number,))
            result = cursor.fetchone()
            
            if result:
                self.notes_text.delete(1.0, tk.END)
                self.notes_text.insert(1.0, result[0] or "")
            else:
                self.notes_text.delete(1.0, tk.END)
                self.notes_text.insert(1.0, "")
            
            self.current_job_notes = job_number
            conn.close()
            
        except Exception as e:
            print(f"Error loading job notes: {e}")
            self.notes_text.delete(1.0, tk.END)
            self.notes_text.insert(1.0, "")
    
    def save_job_notes(self):
        """Save notes for the current job"""
        if not self.current_job_notes:
            messagebox.showwarning("Warning", "No job selected!")
            return
            
        try:
            conn = sqlite3.connect(self.db_manager.db_path)
            cursor = conn.cursor()
            
            # Create notes table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS job_notes (
                    job_number TEXT PRIMARY KEY,
                    notes TEXT
                )
            """)
            
            # Get notes text
            notes_content = self.notes_text.get(1.0, tk.END).strip()
            
            # Save notes
            cursor.execute("""
                INSERT OR REPLACE INTO job_notes (job_number, notes) 
                VALUES (?, ?)
            """, (self.current_job_notes, notes_content))
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Success", "Job notes saved successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save job notes: {str(e)}")
    
    def create_job_context_menu(self):
        """Create right-click context menu for job numbers"""
        self.job_context_menu = tk.Menu(self.root, tearoff=0)
        
        # Add all dashboard applications
        self.job_context_menu.add_command(label="Open in Projects Management", command=lambda: self.open_app_with_job("projects"))
        self.job_context_menu.add_command(label="Open in Product Configurations", command=lambda: self.open_app_with_job("product_configs"))
        self.job_context_menu.add_command(label="Open in Print Package Management", command=lambda: self.open_app_with_job("print_packages"))
        self.job_context_menu.add_command(label="Open in D365 Import Formatter", command=lambda: self.open_app_with_job("d365_formatter"))
        self.job_context_menu.add_command(label="Open in Project File Monitor", command=lambda: self.open_app_with_job("project_monitor"))
        self.job_context_menu.add_command(label="Open in Drawing Reviews", command=lambda: self.open_app_with_job("drawing_reviews"))
        self.job_context_menu.add_command(label="Open in Drafting Checklist", command=lambda: self.open_app_with_job("drafting_checklist"))
        self.job_context_menu.add_command(label="Open in Resource Allocation", command=lambda: self.open_app_with_job("resource_allocation"))
        self.job_context_menu.add_command(label="Open in Workflow Manager", command=lambda: self.open_app_with_job("workflow_manager"))
        self.job_context_menu.add_command(label="Open in Coil Verification", command=lambda: self.open_app_with_job("coil_verification"))
        self.job_context_menu.add_command(label="Open in Job Notes", command=lambda: self.open_app_with_job("job_notes"))
        self.job_context_menu.add_separator()
        # Common actions
        self.job_context_menu.add_command(label="Duplicate", command=self.duplicate_project)
        self.job_context_menu.add_separator()
        self.job_context_menu.add_command(label="Add New Note…", command=self._add_note_from_context)
        
        # Bind right-click event to treeview
        self.tree.bind("<Button-3>", self.show_job_context_menu)  # Button-3 is right-click on Windows
        self.tree.bind("<Button-2>", self.show_job_context_menu)  # Button-2 is right-click on Mac/Linux
    
    def show_job_context_menu(self, event):
        """Show context menu at cursor position"""
        # Select the item under the cursor
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            # Show context menu
            try:
                self.job_context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.job_context_menu.grab_release()
    
    def open_app_with_job(self, app_name):
        """Open specified app with current job number"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "No job selected!")
            return
        
        # Get job number from selected row
        item = selection[0]
        values = self.tree.item(item, 'values')
        if not values:
            messagebox.showwarning("Warning", "No job data found!")
            return
        
        job_number = values[0]  # Job Number is the first column
        
        # Map app names to actual Python files
        app_files = {
            "projects": "projects.py",
            "product_configs": "product_configurations.py",
            # Correct filename in repo is print_package.py
            "print_packages": "print_package.py",
            "d365_formatter": "d365_import_formatter.py",
            "project_monitor": "project_monitor.py",
            "drawing_reviews": "drawing_reviews.py",
            # Correct checklist script name
            "drafting_checklist": "drafting_items_to_look_for.py",
            "resource_allocation": "resource_allocation.py",
            "workflow_manager": "workflow_manager.py",
            "coil_verification": "coil_verification_tool.py",
            "job_notes": "job_notes.py"
        }
        
        app_file = app_files.get(app_name)
        if not app_file:
            messagebox.showerror("Error", f"Unknown app: {app_name}")
            return
        
        # Check if file exists
        if not os.path.exists(app_file):
            messagebox.showerror("Error", f"Application file not found: {app_file}")
            return
        
        try:
            # Launch the app with job number as parameter
            subprocess.Popen([sys.executable, app_file, "--job", str(job_number)])
            print(f"Launched {app_name} with job number: {job_number}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to launch {app_name}:\n{str(e)}")

    def _add_note_from_context(self):
        selection = self.tree.selection()
        job_number = None
        if selection:
            values = self.tree.item(selection[0], 'values')
            if values:
                job_number = values[0]
        # If job_number is None, dialog will prompt to create a Documentation Only job
        open_add_note_dialog(self.root, str(job_number) if job_number else None)

    def preload_job(self, job_number):
        """Preload a specific job number in the table"""
        try:
            # Find the job in the treeview
            for item in self.tree.get_children():
                values = self.tree.item(item, 'values')
                if values and values[0] == str(job_number):
                    # Select and focus on this item
                    self.tree.selection_set(item)
                    self.tree.focus(item)
                    self.tree.see(item)
                    print(f"Preloaded job number: {job_number}")
                    return
            
            print(f"Job number {job_number} not found in current view")
        except Exception as e:
            print(f"Error preloading job {job_number}: {e}")

    def on_closing(self):
        """Handle application closing"""
        self.db_manager.backup_database()
        self.db_manager.export_to_json()
        self.root.destroy()

def main():
    import argparse
    
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Projects Management Application')
    parser.add_argument('--job', type=str, help='Job number to preload')
    args = parser.parse_args()
    
    root = tk.Tk()
    app = ProjectsApp(root)
    
    # If job number provided, select it in the table
    if args.job:
        app.preload_job(args.job)
    
    root.mainloop()

if __name__ == "__main__":
    main()
