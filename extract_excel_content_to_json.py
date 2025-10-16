import json
import os
from tkinter import Tk, filedialog
from openpyxl import load_workbook

def excel_to_json(file_path, output_path):
    wb = load_workbook(file_path, data_only=False)
    workbook_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {}

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None:
                    cell_ref = cell.coordinate
                    if cell.data_type == "f" or str(cell.value).startswith("="):
                        sheet_data[cell_ref] = {"formula": cell.value}
                    else:
                        sheet_data[cell_ref] = {"value": cell.value}
        workbook_data[sheet_name] = sheet_data

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(workbook_data, f, indent=2, ensure_ascii=False)

    print(f"\n✅ Workbook exported to:\n{output_path}")


def main():
    root = Tk()
    root.withdraw()  # Hide the main window

    print("📂 Select the Excel file to read...")
    excel_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm")]
    )
    if not excel_path:
        print("❌ No file selected. Exiting.")
        return

    print("📁 Select the folder to save the JSON output...")
    save_dir = filedialog.askdirectory(title="Select Save Directory")
    if not save_dir:
        print("❌ No directory selected. Exiting.")
        return

    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    output_path = os.path.join(save_dir, f"{base_name}.json")

    excel_to_json(excel_path, output_path)


if __name__ == "__main__":
    main()
