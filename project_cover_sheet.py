import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from datetime import datetime
import os
from database_setup import DatabaseManager
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ProjectCoverSheet:
    def __init__(self, job_number, db_manager):
        self.job_number = job_number
        self.db_manager = db_manager
        self.project_data = None
        self.workflow_data = None
        
    def load_project_data(self):
        """Load project and workflow data from database"""
        conn = sqlite3.connect(self.db_manager.db_path)
        cursor = conn.cursor()
        
        # Load main project data
        query = """
        SELECT p.job_number, p.job_directory, p.customer_name, p.customer_location,
               d.name as assigned_to, p.assignment_date, p.start_date, p.completion_date, 
               p.total_duration_days, p.released_to_dee
        FROM projects p
        LEFT JOIN designers d ON p.assigned_to_id = d.id
        WHERE p.job_number = ?
        """
        
        cursor.execute(query, (self.job_number,))
        self.project_data = cursor.fetchone()
        
        if not self.project_data:
            conn.close()
            return False
            
        # Get project ID for workflow data
        cursor.execute("SELECT id FROM projects WHERE job_number = ?", (self.job_number,))
        project_id = cursor.fetchone()[0]
        
        # Load workflow data
        self.workflow_data = {}
        
        # Initial redline
        cursor.execute("""
            SELECT ir.redline_date, e.name, ir.is_completed
            FROM initial_redline ir
            LEFT JOIN engineers e ON ir.engineer_id = e.id
            WHERE ir.project_id = ?
        """, (project_id,))
        initial_result = cursor.fetchone()
        self.workflow_data['initial_redline'] = initial_result if initial_result else None
        
        # Redline updates
        cursor.execute("""
            SELECT ru.update_cycle, ru.update_date, e.name, ru.is_completed
            FROM redline_updates ru
            LEFT JOIN engineers e ON ru.engineer_id = e.id
            WHERE ru.project_id = ?
            ORDER BY ru.update_cycle
        """, (project_id,))
        redline_updates_result = cursor.fetchall()
        self.workflow_data['redline_updates'] = redline_updates_result if redline_updates_result else []
        
        # OPS review
        cursor.execute("""
            SELECT review_date, is_completed
            FROM ops_review
            WHERE project_id = ?
        """, (project_id,))
        ops_result = cursor.fetchone()
        self.workflow_data['ops_review'] = ops_result if ops_result else None
        
        # D365 BOM Entry
        cursor.execute("""
            SELECT entry_date, is_completed
            FROM d365_bom_entry
            WHERE project_id = ?
        """, (project_id,))
        d365_bom_result = cursor.fetchone()
        self.workflow_data['d365_bom'] = d365_bom_result if d365_bom_result else None
        
        # Peter Weck review
        cursor.execute("""
            SELECT fixed_errors_date, is_completed
            FROM peter_weck_review
            WHERE project_id = ?
        """, (project_id,))
        peter_result = cursor.fetchone()
        self.workflow_data['peter_weck'] = peter_result if peter_result else None
        
        # Release to Dee
        cursor.execute("""
            SELECT release_date, missing_prints_date, d365_updates_date, 
                   other_notes, other_date, is_completed
            FROM release_to_dee
            WHERE project_id = ?
        """, (project_id,))
        release_result = cursor.fetchone()
        self.workflow_data['release_to_dee'] = release_result if release_result else None
        
        conn.close()
        return True
    
    def get_project_status(self):
        """Determine project status based on dates"""
        if self.project_data[7]:  # completion_date
            return "Completed"
        elif self.project_data[6]:  # start_date
            return "In Progress"
        else:
            return "Assigned"
    
    def format_date(self, date_str):
        """Format date string for display"""
        if not date_str:
            return "Not Set"
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            return date_obj.strftime("%m/%d/%Y")
        except:
            return date_str
    
    def format_duration(self, days):
        """Format duration for display"""
        if not days:
            return "N/A"
        return f"{days} days"
    
    def create_cover_sheet_excel(self, output_path):
        """Create the project cover sheet as an Excel workbook"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project Status Report"
        
        # Set page setup for 8.5 x 11 paper size
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER  # 8.5 x 11
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        
        # Center the content horizontally on the page (CORRECT METHOD)
        ws.print_options.horizontalCentered = True
        
        # Set up styles
        header_font = Font(name='Calibri', size=24, bold=True)
        subheader_font = Font(name='Calibri', size=18, bold=True)
        normal_font = Font(name='Calibri', size=12)
        small_font = Font(name='Calibri', size=10)
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Check if Engineering is processed
        engineering_processed = self.check_engineering_processed()
        
        # Job Number (BIG & no label) - Row 1 - Centered on page
        ws['D1'] = str(self.job_number)
        ws['D1'].font = header_font
        ws['D1'].alignment = center_alignment
        
        # Customer Name - Row 2 - Centered on page
        customer_name = self.project_data[2] or "Not Set"
        ws['D2'] = customer_name
        ws['D2'].font = subheader_font
        ws['D2'].alignment = center_alignment
        
        # Customer Location - Row 3 - Centered on page
        customer_location = self.project_data[3] or "Not Set"
        ws['D3'] = customer_location
        ws['D3'].font = subheader_font
        ws['D3'].alignment = center_alignment
        
        # Status - Row 4 - Centered on page
        status = self.get_project_status()
        start_date = self.format_date(self.project_data[6]) if self.project_data[6] else "Not Set"
        
        if status == "In Progress":
            status_text = f"Status: {status}, Started: {start_date}"
        else:
            status_text = f"Status: {status}"
        
        ws['D4'] = status_text
        ws['D4'].font = subheader_font
        ws['D4'].alignment = center_alignment
        
        # Add some space
        ws.row_dimensions[5].height = 20
        
        # Project Workflow Section - Row 6 - Centered on page
        ws['D6'] = "PROJECT WORKFLOW (Latest Update/Next Up)"
        ws['D6'].font = Font(name='Calibri', size=14, bold=True)
        ws['D6'].alignment = center_alignment
        
        # Add engineering warning if not processed - Row 7
        if not engineering_processed:
            ws['D7'] = "⚠️ ENGINEERING NOT PROCESSED ⚠️"
            ws['D7'].font = Font(name='Calibri', size=16, bold=True, color='FF0000')  # Red color
            ws['D7'].alignment = center_alignment
            current_row = 8  # Start workflow content after warning
        else:
            current_row = 7  # Start workflow content normally
        
        # Build complete workflow list
        completed_steps = []
        next_pending = None
        found_pending = False
        
        # 1. Initial Redline
        initial_redline = self.workflow_data.get('initial_redline')
        if initial_redline and isinstance(initial_redline, (list, tuple)) and len(initial_redline) >= 3:
            if initial_redline[2]:  # is_completed
                completed_steps.append(("1. Drafting Drawing Package to Engineering for Initial Review", initial_redline[0]))
            elif not found_pending:
                next_pending = ("1. Drafting Drawing Package to Engineering for Initial Review", None)
                found_pending = True
        
        # 2. Redline Updates
        redline_updates = self.workflow_data.get('redline_updates', [])
        if redline_updates and isinstance(redline_updates, list):
            for update in redline_updates:
                if isinstance(update, (list, tuple)) and len(update) >= 4:
                    if update[3]:  # is_completed
                        completed_steps.append((f"2. Redline Updates", update[1]))
                    elif not found_pending:
                        next_pending = ("2. Redline Updates", None)
                        found_pending = True
                        break
        
        # 3. OPS Review
        ops_review = self.workflow_data.get('ops_review')
        if ops_review and isinstance(ops_review, (list, tuple)) and len(ops_review) >= 2:
            if ops_review[1]:  # is_completed
                completed_steps.append(("3. To Production ofr OPS Review", ops_review[0]))
            elif not found_pending:
                next_pending = ("3. To Production ofr OPS Review", None)
                found_pending = True
        
        # 4. D365 BOM Entry
        d365_bom = self.workflow_data.get('d365_bom')
        if d365_bom and isinstance(d365_bom, (list, tuple)) and len(d365_bom) >= 2:
            if d365_bom[1]:  # is_completed
                completed_steps.append(("4. D365 BOM Entry", d365_bom[0]))
            elif not found_pending:
                next_pending = ("4. D365 BOM Entry", None)
                found_pending = True
        
        # 5. Peter Weck Review
        peter_weck = self.workflow_data.get('peter_weck')
        if peter_weck and isinstance(peter_weck, (list, tuple)) and len(peter_weck) >= 2:
            if peter_weck[1]:  # is_completed
                completed_steps.append(("5. Pete Weck Review", peter_weck[0]))
            elif not found_pending:
                next_pending = ("5. Pete Weck Review", None)
                found_pending = True
        
        # 6. Release to Dee
        release_to_dee = self.workflow_data.get('release_to_dee')
        if release_to_dee and isinstance(release_to_dee, (list, tuple)) and len(release_to_dee) >= 6:
            if release_to_dee[5]:  # is_completed
                completed_steps.append(("6. Release to DEE", release_to_dee[0] if release_to_dee[0] else None))
            elif not found_pending:
                next_pending = ("6. Release to DEE", None)
                found_pending = True
        
        # Display all completed steps with checkmark
        for step_name, date in completed_steps:
            if date:
                ws[f'D{current_row}'] = f"{step_name} ✅"
                ws[f'D{current_row}'].font = normal_font
                ws[f'D{current_row}'].alignment = center_alignment
                current_row += 1
                ws[f'D{current_row}'] = f"Date: {self.format_date(date)}"
                ws[f'D{current_row}'].font = normal_font
                ws[f'D{current_row}'].alignment = center_alignment
                current_row += 1
            else:
                ws[f'D{current_row}'] = f"{step_name} ✅"
                ws[f'D{current_row}'].font = normal_font
                ws[f'D{current_row}'].alignment = center_alignment
                current_row += 1
        
        # Display next pending with unchecked box
        if next_pending:
            step_name, date = next_pending
            ws[f'D{current_row}'] = f"{step_name}: [PENDING]"
            ws[f'D{current_row}'].font = normal_font
            ws[f'D{current_row}'].alignment = center_alignment
            current_row += 1
        
        # Footer with generation date - Row 20 - Centered on page
        ws['D20'] = f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
        ws['D20'].font = small_font
        ws['D20'].alignment = center_alignment
        
        # Remove custom column widths - let horizontal centering do its job
        # ws.column_dimensions['A'].width = 10  # Left margin
        # ws.column_dimensions['B'].width = 10  # Left margin
        # ws.column_dimensions['C'].width = 20  # Content area
        # ws.column_dimensions['D'].width = 20  # Main content (centered)
        # ws.column_dimensions['E'].width = 20  # Content area
        # ws.column_dimensions['F'].width = 10  # Right margin
        # ws.column_dimensions['G'].width = 10  # Right margin
        
        # Set row heights
        ws.row_dimensions[1].height = 40  # Job number
        ws.row_dimensions[2].height = 30  # Customer name
        ws.row_dimensions[3].height = 30  # Customer location
        ws.row_dimensions[4].height = 30  # Status
        
        # Add watermark if engineering not processed (after all content is added)
        # if not engineering_processed:
        #     self.add_engineering_watermark(ws)
        
        # Save the workbook with error handling
        try:
            wb.save(output_path)
            return True
        except PermissionError:
            print(f"Permission denied saving to {output_path}. File may be open in Excel.")
            # Try saving to a different location
            import tempfile
            temp_path = os.path.join(tempfile.gettempdir(), f"temp_status_report_{self.job_number}.xlsx")
            wb.save(temp_path)
            print(f"Saved to temporary location: {temp_path}")
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False
    
    def check_engineering_processed(self):
        """Check if engineering has been processed (any workflow step completed)"""
        # Check if any workflow step is completed
        if self.workflow_data.get('initial_redline') and isinstance(self.workflow_data['initial_redline'], (list, tuple)) and len(self.workflow_data['initial_redline']) >= 3:
            if self.workflow_data['initial_redline'][2]:  # is_completed
                return True
        
        # Check redline updates
        redline_updates = self.workflow_data.get('redline_updates', [])
        if redline_updates and isinstance(redline_updates, list):
            for update in redline_updates:
                if isinstance(update, (list, tuple)) and len(update) >= 4:
                    if update[3]:  # is_completed
                        return True
        
        # Check OPS review
        if self.workflow_data.get('ops_review') and isinstance(self.workflow_data['ops_review'], (list, tuple)) and len(self.workflow_data['ops_review']) >= 2:
            if self.workflow_data['ops_review'][1]:  # is_completed
                return True
        
        return False
    
    def add_engineering_watermark(self, ws):
        """Add a subtle engineering watermark that doesn't interfere with content"""
        # Add a single subtle watermark at the bottom of the page
        watermark_text = "ENGINEERING NOT PROCESSED"
        
        # Add watermark only in empty cells at the bottom
        for row in range(15, 25):  # Bottom rows only
            for col in range(1, 8):  # All columns
                cell = ws.cell(row=row, column=col)
                # Only add watermark if cell is empty
                if cell.value is None:
                    cell.value = watermark_text
                    cell.font = Font(name='Calibri', size=12, color='E0E0E0', italic=True)  # Very light gray
                    cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=45)
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

def print_project_cover_sheet(job_number, db_manager):
    """Main function to create and print project cover sheet"""
    try:
        print(f"DEBUG: Starting cover sheet creation for job {job_number}")
        cover_sheet = ProjectCoverSheet(job_number, db_manager)
        
        if not cover_sheet.load_project_data():
            messagebox.showerror("Error", f"Project {job_number} not found!")
            return False
        
        print(f"DEBUG: Project data loaded successfully")
        print(f"DEBUG: Workflow data: {cover_sheet.workflow_data}")
        
        # Get job directory from project data
        job_directory = cover_sheet.project_data[1]  # job_directory is at index 1
        
        if not job_directory or not os.path.exists(job_directory):
            # Fallback to current directory if job directory not found
            output_dir = os.getcwd()
            messagebox.showwarning("Warning", f"Job directory not found. Saving to: {output_dir}")
        else:
            output_dir = job_directory
        
        # Create single Status Reports folder
        reports_folder = os.path.join(output_dir, "Status Reports")
        
        # Create the reports folder if it doesn't exist
        os.makedirs(reports_folder, exist_ok=True)
        
        # Create output filename with human-readable timestamp
        now = datetime.now()
        date_str = now.strftime("%m-%d-%Y")  # MM-DD-YYYY format
        time_str = now.strftime("%I%M%p").lower()  # HHMMpm format (no colon)
        output_filename = f"{job_number}-ProjectStatusReport-{date_str}@{time_str}.xlsx"
        output_path = os.path.join(reports_folder, output_filename)
        
        print(f"DEBUG: About to create Excel document at {output_path}")
        
        # Create Excel document
        if cover_sheet.create_cover_sheet_excel(output_path):
            print(f"DEBUG: Excel document created successfully")
            # Record the cover sheet generation date in the database
            try:
                conn = sqlite3.connect(db_manager.db_path)
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE projects 
                    SET last_cover_sheet_date = ? 
                    WHERE job_number = ?
                """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), job_number))
                conn.commit()
                conn.close()
            except Exception as e:
                print(f"Warning: Could not update cover sheet date: {e}")
            
            # Open the Excel document (no modal message as requested)
            import subprocess
            try:
                # Try to open with Excel directly
                subprocess.run(['excel', output_path], check=True)
            except (subprocess.CalledProcessError, FileNotFoundError):
                try:
                    # Fallback to using os.startfile
                    os.startfile(output_path)
                except Exception as e:
                    print(f"Could not open Excel document: {e}")
            return True
        else:
            messagebox.showerror("Error", "Failed to create status report")
            return False
            
    except Exception as e:
        print(f"DEBUG: Exception occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        messagebox.showerror("Error", f"Failed to create status report: {str(e)}")
        return False

if __name__ == "__main__":
    # Test the cover sheet generation
    db_manager = DatabaseManager()
    print_project_cover_sheet("35354", db_manager)